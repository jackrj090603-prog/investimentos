import re
import json
import os
import io
import csv
import requests
import openpyxl
import storage
import llm
import telegram_bot
from config import TELEGRAM_CHAT_ID_CONVERSA

def buscar_documentos_relacionados(pergunta, limit=3):
    """
    RAG Leve: Busca documentos no banco baseando-se em palavras em comum,
    recência (ID maior) e relevância.
    """
    # Extrair palavras-chave significativas da pergunta
    palavras = [p.lower().strip() for p in re.findall(r"\w{4,}", pergunta)]
    if not palavras:
        return storage.get_todos_documentos(limit=limit)
        
    todos_docs = storage.get_todos_documentos(limit=300) # Ler últimos 300 docs
    
    scored_docs = []
    for doc in todos_docs:
        score = 0
        text_to_search = " ".join([
            str(doc.get("company_name", "")),
            str(doc.get("ticker", "")),
            str(doc.get("category", "")),
            str(doc.get("doc_type", "")),
            str(doc.get("description", ""))
        ]).lower()
        
        # Somar pontos para cada palavra em comum
        for p in palavras:
            if p in text_to_search:
                score += 1
                
        # Se for menção exata ao ticker, dá peso extra
        for p in palavras:
            if len(p) <= 6 and p.upper() in str(doc.get("ticker", "")).upper():
                score += 3
                
        if score > 0:
            scored_docs.append((score, doc["id"], doc))
            
    # Ordenar por score descendente e por ID (recência) descendente
    scored_docs.sort(key=lambda x: (x[0], x[1]), reverse=True)
    
    # Retornar apenas a lista de documentos ordenados
    return [item[2] for item in scored_docs[:limit]]

def baixar_e_atualizar_planilhas_mira():
    """
    Pesquisa e atualiza arquivos MONITOR DE VALUATION DO MIRA*.xlsx
    localizados nos caminhos padrão de Downloads e OneDrive/Desktop.
    """
    url = "https://statusinvest.com.br/category/advancedsearchresultexport"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    params = {
        "CategoryType": 1,
        "search": json.dumps({"Sector": "", "SubSector": "", "Segment": "", "my_range": "-20;20"})
    }
    
    try:
        print("[StatusInvest] Iniciando download da planilha de ações...")
        res = requests.get(url, params=params, headers=headers, timeout=25)
        if res.status_code != 200:
            print(f"[StatusInvest] Erro de rede: {res.status_code}")
            return False, f"Erro de rede ao baixar do StatusInvest (HTTP {res.status_code})"
            
        csv_text = res.content.decode("utf-8-sig")
        f = io.StringIO(csv_text)
        reader = csv.reader(f, delimiter=";")
        rows = list(reader)
        if not rows:
            return False, "O arquivo baixado do StatusInvest estava vazio."
            
        print(f"[StatusInvest] Sucesso! Baixado dados de {len(rows) - 1} ativos.")
        
        # Procurar planilhas do Mira no computador do usuário
        search_dirs = [
            "c:/Users/jackr/OneDrive/Desktop/PROJETO_CF",
            "c:/Users/jackr/Downloads",
            "c:/Users/jackr/OneDrive/Desktop",
            "c:/Users/jackr/Desktop"
        ]
        
        updated_files = []
        visited_paths = set()
        
        # Conversão inteligente de tipo de dado
        def clean_val(val, is_header=False):
            if is_header:
                return val
            val = val.strip()
            if not val or val == "-" or val.lower() in ("null", "none"):
                return None
            try:
                # Tratar milhares "." e decimal ","
                cleaned = val.replace(".", "").replace(",", ".")
                if "." in cleaned:
                    return float(cleaned)
                else:
                    return int(cleaned)
            except ValueError:
                return val

        for s_dir in search_dirs:
            if not os.path.exists(s_dir):
                continue
                
            for root, dirs, files in os.walk(s_dir):
                # Ignorar pastas virtuais ou temporárias do Python/Node/Git/Gemini
                if any(x in root.lower() for x in ["venv", "node_modules", ".git", ".gemini"]):
                    continue
                    
                for filename in files:
                    if "monitor de valuation" in filename.lower() and filename.endswith(".xlsx"):
                        file_path = os.path.abspath(os.path.join(root, filename))
                        if file_path in visited_paths:
                            continue
                        visited_paths.add(file_path)
                        
                        print(f"[StatusInvest] Atualizando planilha: {file_path}...")
                        
                        try:
                            wb = openpyxl.load_workbook(file_path)
                            if "Dados Status Invest" not in wb.sheetnames:
                                sheet = wb.create_sheet("Dados Status Invest")
                            else:
                                sheet = wb["Dados Status Invest"]
                                sheet.delete_rows(1, sheet.max_row + 10)
                                
                            # Escrever cabeçalhos e dados
                            for r_idx, row in enumerate(rows):
                                is_h = (r_idx == 0)
                                for c_idx, val in enumerate(row):
                                    cleaned = clean_val(val, is_header=is_h)
                                    sheet.cell(row=r_idx + 1, column=c_idx + 1, value=cleaned)
                                    
                            wb.save(file_path)
                            wb.close()
                            updated_files.append(filename)
                            print(f"[StatusInvest] Planilha {filename} salva com sucesso.")
                        except Exception as err:
                            print(f"[StatusInvest] Erro ao editar {filename}: {err}")
                            
        if updated_files:
            return True, f"Atualizado {len(updated_files)} planilha(s): " + ", ".join(updated_files)
        else:
            return False, "Nenhuma planilha 'MONITOR DE VALUATION DO MIRA' encontrada para atualizar."
            
    except Exception as e:
        print(f"[StatusInvest] Erro durante o processo: {e}")
        return False, f"Erro crítico: {e}"

def responder_updates():
    """Lê novas mensagens do Telegram, aplica RAG e responde."""
    # Buscar o último offset processado para não duplicar respostas
    ultimo_offset_chave = "telegram_chat_offset"
    offset_str = storage.get_ultimo_timestamp() # vamos reusar a tabela de estado
    # Melhor criar um método de estado genérico
    conn = storage.get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT valor FROM estado WHERE chave = ?", (ultimo_offset_chave,))
    row = cursor.fetchone()
    offset = int(row["valor"]) if row else None
    
    updates = telegram_bot.ler_updates(offset=offset)
    
    for update in updates:
        update_id = update.get("update_id")
        
        # Salvar o novo offset (update_id + 1)
        next_offset = update_id + 1
        cursor.execute("INSERT OR REPLACE INTO estado (chave, valor) VALUES (?, ?)", (ultimo_offset_chave, str(next_offset)))
        conn.commit()
        
        message = update.get("message") or update.get("channel_post") or update.get("edited_message") or update.get("edited_channel_post")
        if not message:
            continue
            
        chat = message.get("chat", {})
        chat_id = chat.get("id")
        
        # Auto-download files sent to the target group in real-time
        document = message.get("document")
        if document and chat_id:
            file_id = document.get("file_id")
            file_name = document.get("file_name") or f"arquivo_{message.get('message_id')}"
            if str(chat_id) in ["-2217590850", "-1002217590850"]:
                try:
                    token = getattr(telegram_bot, 'TELEGRAM_BOT_TOKEN', '') or os.getenv("TELEGRAM_BOT_TOKEN", "")
                    if token:
                        file_info_url = f"https://api.telegram.org/bot{token}/getFile?file_id={file_id}"
                        res_info = requests.get(file_info_url).json()
                        if res_info.get("ok"):
                            file_path = res_info["result"]["file_path"]
                            download_url = f"https://api.telegram.org/file/bot{token}/{file_path}"
                            
                            output_dir = "Kenny"
                            # Handle relative paths from different run contexts
                            if not os.path.exists(output_dir) and os.path.exists("../Kenny"):
                                output_dir = "../Kenny"
                            os.makedirs(output_dir, exist_ok=True)
                            
                            dest_path = os.path.join(output_dir, file_name)
                            file_data = requests.get(download_url).content
                            with open(dest_path, "wb") as f_out:
                                f_out.write(file_data)
                            print(f"[Bot Downloader] Novo arquivo salvo na pasta Kenny: {file_name}")
                            telegram_bot.send_message(chat_id, f"📥 <b>[Kenny] Arquivo salvo automaticamente na pasta local:</b> {file_name}")
                except Exception as dl_err:
                    print(f"[Bot Downloader] Erro ao baixar: {dl_err}")
                    
        text = message.get("text", "").strip()
        
        if not text or not chat_id:
            continue
            
        print(f"[Chat] Mensagem recebida de {chat_id}: '{text}'")
        
        # Enviar ação de digitando
        telegram_bot.send_chat_action(chat_id, action="typing")
        
        # Tratar comando /start
        if text.startswith("/start"):
            welcome = (
                "👋 <b>Olá! Eu sou o Agente de Monitoramento da CVM & RI da Ceará Finance.</b>\n\n"
                "Estou monitorando as empresas cadastradas no seu painel. "
                "Você pode me fazer perguntas sobre os fatos relevantes recentes, empresas cadastradas "
                "ou usar o comando /statusinvest para atualizar seus arquivos de valuation automaticamente!\n\n"
                "💡 <i>Exemplo de pergunta: 'Quais os últimos fatos da Direcional?'</i>"
            )
            telegram_bot.send_message(chat_id, welcome)
            storage.salvar_mensagem(chat_id, "user", text)
            storage.salvar_mensagem(chat_id, "assistant", welcome)
            continue
            
        # Tratar comandos de atualização da planilha do Mira
        if text.startswith("/statusinvest") or text.startswith("/atualizar") or "atualizar planilha" in text.lower():
            # Notificar que está baixando
            telegram_bot.send_message(chat_id, "⏳ <b>Iniciando download dos dados no StatusInvest e atualização das planilhas...</b>")
            
            sucesso, msg = baixar_e_atualizar_planilhas_mira()
            if sucesso:
                resposta = f"✅ <b>Planilha Atualizada!</b>\n\n{msg}\nOs dados de 617 ações da B3 foram consolidados e carregados na aba 'Dados Status Invest'."
            else:
                resposta = f"❌ <b>Erro na Atualização!</b>\n\n{msg}\nVerifique se o arquivo está aberto no Excel ou se o caminho está acessível."
                
            telegram_bot.send_message(chat_id, resposta)
            storage.salvar_mensagem(chat_id, "user", text)
            storage.salvar_mensagem(chat_id, "assistant", resposta)
            continue
            
        # 1. RAG Leve: Busca documentos relacionados na base local
        docs_relacionados = buscar_documentos_relacionados(text, limit=3)
        
        # 2. Obter Histórico de Conversação do banco de dados
        historico = storage.get_historico_chat(chat_id, limit=10)
        
        # 3. Chamar a IA (Gemini)
        resposta_ia = llm.perguntar_ao_gemini(text, historico, docs_relacionados)
        
        # 4. Enviar mensagem de resposta
        telegram_bot.send_message(chat_id, resposta_ia)
        
        # 5. Salvar histórico no banco
        storage.salvar_mensagem(chat_id, "user", text)
        storage.salvar_mensagem(chat_id, "assistant", resposta_ia)
        
    conn.close()
