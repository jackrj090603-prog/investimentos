import os
import openpyxl
from dotenv import load_dotenv

# Carregar variáveis do arquivo .env
load_dotenv()

# Caminhos do projeto
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "agente_cvm.db")
EMPRESAS_XLSX = os.path.join(BASE_DIR, "empresas.xlsx")

# Variáveis do Telegram e Gemini
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID_ALERTAS = os.getenv("TELEGRAM_CHAT_ID_ALERTAS", "")
TELEGRAM_CHAT_ID_CONVERSA = os.getenv("TELEGRAM_CHAT_ID_CONVERSA", "")

# Configuração do Heartbeat
HEARTBEAT_INTERVAL = 300  # 5 minutos
MONITOR_INTERVAL = 900    # 15 minutos

# Lista padrão de empresas para criar o empresas.xlsx caso não exista
EMPRESAS_PADRAO = [
    {"TICKER": "DIRR3", "CNPJ": "03.141.011/0001-34", "COD_CVM": "02182-2", "NOME": "DIRECIONAL ENGENHARIA S.A."},
    {"TICKER": "PETR4", "CNPJ": "33.000.167/0001-01", "COD_CVM": "00951-2", "NOME": "PETROLEO BRASILEIRO S.A. PETROBRAS"},
    {"TICKER": "VALE3", "CNPJ": "15.031.206/0001-55", "COD_CVM": "00417-0", "NOME": "VALE S.A."},
    {"TICKER": "WEGE3", "CNPJ": "84.429.695/0001-11", "COD_CVM": "01540-7", "NOME": "WEG S.A."}
]

def inicializar_empresas_xlsx():
    """Cria o arquivo empresas.xlsx se ele não existir."""
    if not os.path.exists(EMPRESAS_XLSX):
        print(f"[Config] Criando arquivo de empresas padrão em: {EMPRESAS_XLSX}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empresas"
        
        # Cabeçalhos
        headers = ["TICKER", "CNPJ", "COD_CVM", "NOME"]
        ws.append(headers)
        
        # Inserir dados
        for emp in EMPRESAS_PADRAO:
            ws.append([emp["TICKER"], emp["CNPJ"], emp["COD_CVM"], emp["NOME"]])
            
        wb.save(EMPRESAS_XLSX)
        wb.close()

def carregar_empresas():
    """Carrega a lista de empresas do excel empresas.xlsx."""
    inicializar_empresas_xlsx()
    
    empresas = []
    try:
        wb = openpyxl.load_workbook(EMPRESAS_XLSX, data_only=True)
        ws = wb.active
        
        # Ler linhas
        header = None
        for row in ws.iter_rows(values_only=True):
            if not header:
                header = [str(cell).upper().strip() for cell in row]
                continue
            
            # Mapear dados
            if not any(row):
                continue
                
            data = dict(zip(header, row))
            ticker = str(data.get("TICKER", "")).strip()
            cnpj = str(data.get("CNPJ", "")).strip()
            cod_cvm = str(data.get("COD_CVM", "")).strip()
            nome = str(data.get("NOME", "")).strip()
            
            if ticker and (cnpj or cod_cvm):
                # Limpar CNPJ de caracteres não numéricos para comparação
                cnpj_limpo = "".join(filter(str.isdigit, cnpj))
                # Limpar código CVM de caracteres não numéricos
                cvm_limpo = "".join(filter(str.isdigit, cod_cvm))
                
                empresas.append({
                    "ticker": ticker,
                    "cnpj": cnpj,
                    "cnpj_limpo": cnpj_limpo,
                    "cvm_code": cod_cvm,
                    "cvm_limpo": cvm_limpo,
                    "nome": nome
                })
        wb.close()
    except Exception as e:
        print(f"[Config] Erro ao carregar empresas.xlsx: {e}")
        
    return empresas
