from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
import json
import socketserver
import threading
import urllib.parse
import os
import sys
import openpyxl
import sqlite3
import pandas as pd
import numpy as np
import yfinance as yf
import requests
from datetime import datetime

# Adicionar caminhos do valuation para importação
sys.path.append(os.path.abspath("../valuation"))
sys.path.append(os.path.abspath("../valuation/src"))
sys.path.append(os.path.abspath("valuation"))
sys.path.append(os.path.abspath("valuation/src"))

import storage
import config
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
import cvm_robot

try:
    from wacc import calcular_beta, calcular_ke, calcular_wacc, desalavancar_beta, realavancar_beta
    from dcf import projetar_fluxos, calcular_dcf
    from monte_carlo import rodar_simulacao, calcular_estatisticas, plotar_histograma
except ImportError:
    print("[Dashboard] Rodando sem dependências de valuation prontas no sys.path. Fallbacks configurados.")

CONFIG_EMPRESAS = {
    "DIRR3": {
        "nome": "Direcional Engenharia S.A.",
        "num_acoes": 172000000,
        "divida_liquida": 613000000,
        "margem_ebit": 0.15,
        "vgv_lancado": [4000000000, 4300000000, 4600000000, 4900000000, 5200000000],
        "vso": [0.60, 0.62, 0.63, 0.64, 0.65],
        "poc": {"ano_1": 0.20, "ano_2": 0.50, "ano_3": 0.30}
    },
    "PETR4": {
        "nome": "Petróleo Brasileiro S.A. - Petrobras",
        "num_acoes": 13044496930,
        "divida_liquida": 320000000000,
        "margem_ebit": 0.32,
        "vgv_lancado": [80000000000, 83000000000, 86000000000, 89000000000, 92000000000],
        "vso": [0.85, 0.85, 0.85, 0.85, 0.85],
        "poc": {"ano_1": 0.33, "ano_2": 0.33, "ano_3": 0.34}
    },
    "VALE3": {
        "nome": "Vale S.A.",
        "num_acoes": 4539000000,
        "divida_liquida": 65000000000,
        "margem_ebit": 0.38,
        "vgv_lancado": [140000000000, 145000000000, 150000000000, 155000000000, 160000000000],
        "vso": [0.90, 0.90, 0.90, 0.90, 0.90],
        "poc": {"ano_1": 0.33, "ano_2": 0.33, "ano_3": 0.34}
    },
    "WEGE3": {
        "nome": "WEG S.A.",
        "num_acoes": 4197317998,
        "divida_liquida": -3500000000,
        "margem_ebit": 0.17,
        "vgv_lancado": [22000000000, 24000000000, 26000000000, 28000000000, 30000000000],
        "vso": [0.95, 0.95, 0.95, 0.95, 0.95],
        "poc": {"ano_1": 0.33, "ano_2": 0.33, "ano_3": 0.34}
    }
}

class DashboardHandler(SimpleHTTPRequestHandler):
    def do_GET(self):
        parsed_path = urllib.parse.urlparse(self.path)
        params = urllib.parse.parse_qs(parsed_path.query)
        
        if parsed_path.path == "/" or parsed_path.path == "/index.html":
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            html_content = self.render_dashboard()
            self.wfile.write(html_content.encode("utf-8"))
            
        elif parsed_path.path == "/api/documentos":
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            
            query = params.get("q", [""])[0].strip()
            if query:
                docs = storage.buscar_documentos_por_termo(query)
            else:
                docs = storage.get_todos_documentos(limit=100)
                
            self.wfile.write(json.dumps(docs).encode("utf-8"))
            
        elif parsed_path.path == "/api/valuation":
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            
            valuation_data = self.carregar_valuation_mira()
            self.wfile.write(json.dumps(valuation_data).encode("utf-8"))
            
        elif parsed_path.path == "/api/relatorio":
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            
            ticker = params.get("ticker", ["DIRR3"])[0].strip().upper()
            report_data = self.gerar_relatorio_quant_poli(ticker)
            self.wfile.write(json.dumps(report_data).encode("utf-8"))
            
        elif parsed_path.path == "/api/demonstrativos":
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            
            ticker = params.get("ticker", ["DIRR3"])[0].strip().upper()
            df_data = self.gerar_demonstrativos_financeiros(ticker)
            self.wfile.write(json.dumps(df_data).encode("utf-8"))
            
        elif parsed_path.path == "/api/cf_tech/config":
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            
            config_path = "../cf_tech_valuation/config/settings.yaml"
            if not os.path.exists(config_path):
                config_path = "cf_tech_valuation/config/settings.yaml"
                
            try:
                import yaml
                with open(config_path, "r", encoding="utf-8") as f:
                    cfg_data = yaml.safe_load(f)
                self.wfile.write(json.dumps({"sucesso": True, "config": cfg_data}).encode("utf-8"))
            except Exception as cfg_err:
                self.wfile.write(json.dumps({"sucesso": False, "erro": str(cfg_err)}).encode("utf-8"))
                
        elif parsed_path.path == "/api/cf_tech/run":
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            
            stage = params.get("stage", ["all"])[0].strip()
            ticker = params.get("ticker", ["DIRR3"])[0].strip().upper()
            preco_atual = params.get("preco_atual", ["11.50"])[0].strip()
            
            import subprocess
            cmd = [
                sys.executable,
                "../cf_tech_valuation/main.py",
                "--stage", stage,
                "--ticker", ticker,
                "--preco-atual", preco_atual
            ]
            if not os.path.exists("../cf_tech_valuation/main.py"):
                cmd[1] = "cf_tech_valuation/main.py"
                
            try:
                res_proc = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
                out_text = res_proc.stdout + "\n" + res_proc.stderr
                success = (res_proc.returncode == 0)
                
                # Check if generated Parquet files exist, load to list
                kpis_data = []
                kpis_path = "../cf_tech_valuation/data/processed/kpis_calculados.parquet"
                if not os.path.exists(kpis_path):
                    kpis_path = "cf_tech_valuation/data/processed/kpis_calculados.parquet"
                if os.path.exists(kpis_path):
                    try:
                        df_res = pd.read_parquet(kpis_path)
                        # Replace infinity and NaNs to avoid JSON encoding errors
                        df_res = df_res.replace([np.inf, -np.inf], 0.0).fillna(0.0)
                        kpis_data = df_res.to_dict(orient="records")
                    except Exception as parquet_err:
                        print(f"Error reading Parquet: {parquet_err}")
                        
                # Also try to load the research agent Markdown report
                report_text = ""
                if stage in ["risk", "all"] and success:
                    try:
                        # Append path to sys.path dynamically
                        sys.path.append(os.path.abspath("../cf_tech_valuation"))
                        sys.path.append(os.path.abspath("cf_tech_valuation"))
                        sys.path.append(os.path.abspath("../cf_tech_valuation/src"))
                        sys.path.append(os.path.abspath("cf_tech_valuation/src"))
                        
                        from cf_tech_valuation.src.ai_agent.research_assistant import run_research_assistant
                        report_text = run_research_assistant(kpis_path=kpis_path)
                    except Exception as ass_err:
                        report_text = f"Erro ao executar Assistente: {ass_err}"
                
                self.wfile.write(json.dumps({
                    "sucesso": success,
                    "log": out_text,
                    "kpis": kpis_data,
                    "report": report_text
                }).encode("utf-8"))
            except Exception as run_err:
                self.wfile.write(json.dumps({
                    "sucesso": False,
                    "log": f"Erro na execução da pipeline: {run_err}",
                    "kpis": [],
                    "report": ""
                }).encode("utf-8"))
                
        elif parsed_path.path == "/LOGO_CF.png":
            self.send_response(200)
            self.send_header("Content-Type", "image/png")
            self.end_headers()
            logo_path = "../Mira/LOGO_CF.png"
            if not os.path.exists(logo_path):
                logo_path = "Mira/LOGO_CF.png"
            if not os.path.exists(logo_path):
                logo_path = "c:/Users/jackr/OneDrive/Desktop/PROJETO_CF/Mira/LOGO_CF.png"
            if os.path.exists(logo_path):
                with open(logo_path, "rb") as f:
                    self.wfile.write(f.read())
            else:
                self.wfile.write(b"")
                
        elif parsed_path.path in ["/monte_carlo_temp.png", "/chart_historico.png", "/chart_retornos.png"]:
            self.send_response(200)
            self.send_header("Content-Type", "image/png")
            self.end_headers()
            image_path = parsed_path.path.lstrip("/")
            if os.path.exists(image_path):
                with open(image_path, "rb") as f:
                    self.wfile.write(f.read())
            else:
                self.wfile.write(b"")

        elif parsed_path.path.startswith("/data/") and parsed_path.path.endswith(".pdf"):
            self.send_response(200)
            self.send_header("Content-Type", "application/pdf")
            self.end_headers()
            pdf_path = parsed_path.path.lstrip("/")
            paths_to_try = [
                pdf_path,
                os.path.join("agente_cvm", pdf_path),
                os.path.join(BASE_DIR, pdf_path),
                os.path.join("../", pdf_path)
            ]
            served = False
            for pt in paths_to_try:
                if os.path.exists(pt):
                    with open(pt, "rb") as f:
                        self.wfile.write(f.read())
                    served = True
                    break
            if not served:
                self.wfile.write(b"")
        elif parsed_path.path == "/workstation.html":
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            work_path = "../Mira/ai_studio_code (28) (3).html"
            if not os.path.exists(work_path):
                work_path = "Mira/ai_studio_code (28) (3).html"
            if not os.path.exists(work_path):
                work_path = "c:/Users/jackr/OneDrive/Desktop/PROJETO_CF/Mira/ai_studio_code (28) (3).html"
            if os.path.exists(work_path):
                with open(work_path, "r", encoding="utf-8") as f:
                    self.wfile.write(f.read().encode("utf-8"))
            else:
                self.wfile.write(b"Workstation file not found.")
                
        else:
            super().do_GET()

    def do_POST(self):
        parsed_path = urllib.parse.urlparse(self.path)
        
        if parsed_path.path == "/api/cf_tech/config":
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            
            try:
                import yaml
                cfg_data = json.loads(post_data.decode("utf-8"))
                
                config_path = "../cf_tech_valuation/config/settings.yaml"
                if not os.path.exists(config_path):
                    config_path = "cf_tech_valuation/config/settings.yaml"
                    
                with open(config_path, "w", encoding="utf-8") as f:
                    yaml.safe_dump(cfg_data, f, default_flow_style=False, allow_unicode=True)
                    
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps({"sucesso": True}).encode("utf-8"))
            except Exception as e:
                self.send_response(400)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps({"sucesso": False, "erro": str(e)}).encode("utf-8"))
        else:
            self.send_response(404)
            self.end_headers()

    def carregar_valuation_mira(self):
        """Carrega e calcula a planilha do Mira diretamente do Excel em tempo real."""
        excel_path = "../Mira/MONITOR DE VALUATION DO MIRA (3).xlsx"
        if not os.path.exists(excel_path):
            excel_path = "Mira/MONITOR DE VALUATION DO MIRA (3).xlsx"
        if not os.path.exists(excel_path):
            excel_path = "c:/Users/jackr/OneDrive/Desktop/PROJETO_CF/Mira/MONITOR DE VALUATION DO MIRA (3).xlsx"
            
        if not os.path.exists(excel_path):
            return []
            
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            # Carregar mapeamento setorial B3
            b3_map = {}
            sheet_b3_name = 'Classificação B3' if 'Classificação B3' in wb.sheetnames else 'Classificacao B3'
            if sheet_b3_name in wb.sheetnames:
                sheet_b3 = wb[sheet_b3_name]
                for r in range(2, sheet_b3.max_row + 1):
                    ticker = sheet_b3.cell(r, 1).value
                    if ticker:
                        ticker_str = str(ticker).strip().upper()
                        b3_map[ticker_str] = {
                            "nome": sheet_b3.cell(r, 2).value,
                            "setor": sheet_b3.cell(r, 4).value,
                            "segmento": sheet_b3.cell(r, 6).value,
                            "ciclico": sheet_b3.cell(r, 7).value
                        }
            
            # Carregar dados brutos do StatusInvest
            sheet_si = wb['Dados Status Invest']
            stocks = []
            
            for r in range(2, sheet_si.max_row + 1):
                ticker = sheet_si.cell(r, 1).value
                if not ticker:
                    continue
                ticker = str(ticker).strip().upper()
                
                def get_val(col_idx):
                    val = sheet_si.cell(r, col_idx).value
                    if val is None or val == "-":
                        return 0.0
                    try:
                        return float(val)
                    except:
                        return 0.0
                
                preco = get_val(2)
                dy_raw = get_val(3)
                pl = get_val(4)
                roe_raw = get_val(18)
                cagr_lucro_raw = get_val(25)
                vpa = get_val(27)
                lpa = get_val(28)
                
                dy = dy_raw / 100.0
                roe = roe_raw / 100.0
                cagr_lucro = cagr_lucro_raw / 100.0
                
                dpa = preco * dy
                payout = (dpa / lpa) if lpa != 0.0 else 0.0
                
                if payout <= 0.40:
                    p_tipo = "CRESCIMENTO"
                elif payout <= 0.60:
                    p_tipo = "INDEFINIDO"
                else:
                    p_tipo = "DIVIDENDOS"
                
                crescimento_damodaran = (1.0 - payout) * roe
                media_crescimento = (cagr_lucro + crescimento_damodaran) / 2.0
                
                preco_justo_bazin = dpa / 0.06
                preco_teto_bazin = preco_justo_bazin / 1.3
                desc_bazin = (preco_justo_bazin - preco) / preco_justo_bazin if preco_justo_bazin > 0 else -10.0
                
                mult = 22.5 * lpa * vpa
                preco_justo_graham = (mult ** 0.5) if mult > 0 else 0.0
                desc_graham = (preco_justo_graham - preco) / preco_justo_graham if preco_justo_graham > 0 else -10.0
                
                info = b3_map.get(ticker, {"nome": ticker, "setor": "-", "segmento": "-", "ciclico": "-"})
                
                stocks.append({
                    "ticker": ticker,
                    "nome": info["nome"] or ticker,
                    "setor": info["setor"] or "-",
                    "segmento": info["segmento"] or "-",
                    "ciclico": info["ciclico"] or "-",
                    "preco": preco,
                    "dy": dy_raw,
                    "pl": pl,
                    "lpa": lpa,
                    "vpa": vpa,
                    "dpa": dpa,
                    "payout": payout * 100.0,
                    "tipo": p_tipo,
                    "roe": roe_raw,
                    "cagr": cagr_lucro_raw,
                    "preco_justo_bazin": preco_justo_bazin,
                    "preco_teto_bazin": preco_teto_bazin,
                    "desc_bazin": desc_bazin * 100.0,
                    "preco_justo_graham": preco_justo_graham,
                    "desc_graham": desc_graham * 100.0
                })
            
            wb.close()
            stocks.sort(key=lambda x: x["ticker"])
            return stocks
        except Exception as e:
            print(f"[Dashboard] Erro ao carregar valuation do Excel: {e}")
            return []

    def gerar_relatorio_quant_poli(self, ticker):
        """Executa os cálculos quantitativos do motor de valuation e gera os 3 gráficos correspondentes."""
        try:
            cfg = CONFIG_EMPRESAS.get(ticker, CONFIG_EMPRESAS["DIRR3"])
            
            # Buscar cotação atual no Excel (se disponível) ou usar default
            cotacao = 11.50
            stocks_mira = self.carregar_valuation_mira()
            for s in stocks_mira:
                if s["ticker"] == ticker:
                    cotacao = s["preco"]
                    break
            
            # Montar premissas em dict dinâmico
            premissas = {
                "ativo": {
                    "ticker": f"{ticker}.SA",
                    "indice": "^BVSP",
                    "cotacao_atual": cotacao,
                    "num_acoes": cfg["num_acoes"],
                    "divida_liquida": cfg["divida_liquida"]
                },
                "wacc": {
                    "taxa_livre_risco": 0.12,
                    "premio_risco_mercado": 0.06,
                    "custo_divida": 0.115,
                    "aliquota_imposto": 0.04,
                    "de_ratio": 0.15
                },
                "projecao": {
                    "anos": 5,
                    "perpetuidade_g": 0.025,
                    "depreciacao_percent": 0.02,
                    "capex_percent": 0.015,
                    "nwc_percent": 0.10,
                    "vgv_lancado": cfg["vgv_lancado"],
                    "vso": cfg["vso"],
                    "poc": cfg["poc"],
                    "margem_ebit": cfg["margem_ebit"]
                },
                "monte_carlo": {
                    "num_simulacoes": 10000,
                    "distribuicoes": {
                        "vso_media": [0.45, 0.62, 0.75],
                        "margem_ebit": [0.12, 0.15, 0.18],
                        "perpetuidade_g": [0.01, 0.025, 0.04],
                        "taxa_livre_risco": [0.10, 0.12, 0.14],
                        "custo_divida": [0.09, 0.115, 0.13]
                    }
                }
            }
            
            # 1. Calcular WACC
            beta_m = calcular_beta(premissas["ativo"]["ticker"], index_ticker="^BVSP", period="2y")
            beta_u = desalavancar_beta(beta_m, premissas["wacc"]["de_ratio"], premissas["wacc"]["aliquota_imposto"])
            beta_l = realavancar_beta(beta_u, premissas["wacc"]["de_ratio"], premissas["wacc"]["aliquota_imposto"])
            ke = calcular_ke(premissas["wacc"]["taxa_livre_risco"], beta_l, premissas["wacc"]["premio_risco_mercado"])
            wacc_final = calcular_wacc(ke, premissas["wacc"]["custo_divida"], premissas["wacc"]["de_ratio"], premissas["wacc"]["aliquota_imposto"])
            
            # 2. Calcular DCF
            fluxos = projetar_fluxos(premissas)
            dcf_res = calcular_dcf(fluxos, wacc_final, premissas["projecao"]["perpetuidade_g"], premissas["ativo"]["divida_liquida"], premissas["ativo"]["num_acoes"])
            
            # 3. Simulação Monte Carlo
            precos_simulados = rodar_simulacao(premissas, beta_l, n_simulacoes=10000)
            stats = calcular_estatisticas(precos_simulados, cotacao)
            
            # 4. Salvar os 3 gráficos analíticos (Monte Carlo, Média Móvel, Distribuição de Retornos)
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            
            # (a) Histograma Monte Carlo
            plotar_histograma(precos_simulados, stats, cotacao, "monte_carlo_temp.png")
            
            # (b) Download histórico de 2 anos para médias móveis e retornos
            df = yf.download(premissas["ativo"]["ticker"], period="2y", progress=False)
            if not df.empty:
                close = df['Close']
                if isinstance(close, pd.DataFrame):
                    close = close.iloc[:, 0]
                    
                sma20 = close.rolling(window=20).mean()
                sma50 = close.rolling(window=50).mean()
                
                # Gráfico Histórico + SMA
                plt.figure(figsize=(9, 4.5))
                plt.plot(close.index, close.values, label="Fechamento Ajustado", color="#5b9bff", linewidth=1.5)
                plt.plot(sma20.index, sma20.values, label="Média Móvel 20d", color="#46e0a0", linestyle="--")
                plt.plot(sma50.index, sma50.values, label="Média Móvel 50d", color="#a472ff", linestyle="--")
                plt.title(f"Evolução do Preço e Médias Móveis — {ticker}", color="white", fontsize=11, fontweight="bold")
                plt.legend(facecolor="#0c0d12", edgecolor="#1b1c24", labelcolor="white")
                ax = plt.gca()
                ax.set_facecolor("#050508")
                ax.tick_params(colors='white')
                ax.grid(color="#1b1c24", linestyle=":")
                plt.tight_layout()
                plt.savefig("chart_historico.png", facecolor="#050508", dpi=150)
                plt.close()
                
                # Gráfico Distribuição de Retornos
                returns = close.pct_change().dropna() * 100.0
                plt.figure(figsize=(9, 4.5))
                plt.hist(returns.values, bins=50, color="#a472ff", edgecolor="#0c0d12", alpha=0.8)
                plt.axvline(returns.mean(), color="#ff6b6b", label=f"Média: {returns.mean():.2f}%")
                plt.title(f"Distribuição de Retornos Diários (%) — {ticker}", color="white", fontsize=11, fontweight="bold")
                plt.legend(facecolor="#0c0d12", edgecolor="#1b1c24", labelcolor="white")
                ax = plt.gca()
                ax.set_facecolor("#050508")
                ax.tick_params(colors='white')
                ax.grid(color="#1b1c24", linestyle=":")
                plt.tight_layout()
                plt.savefig("chart_retornos.png", facecolor="#050508", dpi=150)
                plt.close()
            
            fluxos_detalhe = []
            for i in range(5):
                fluxos_detalhe.append({
                    "ano": i + 1,
                    "fluxo": fluxos[i],
                    "vp": dcf_res["vp_fluxos"][i]
                })
                
            upside = (dcf_res["valor_por_acao"] / cotacao) - 1.0
            recomendacao = "BUY (COMPRA)" if upside > 0.10 else "SELL (VENDA)"
            
            # Tentar extrair texto do PDF local do relatório anual
            texto_pdf = ""
            pdf_path = f"data/Relatorio_Anual_{ticker}.pdf"
            # Verificar caminhos alternativos dependendo de onde o servidor foi executado
            paths_to_try = [
                pdf_path,
                os.path.join("agente_cvm", pdf_path),
                os.path.join(BASE_DIR, pdf_path),
                os.path.join("../", pdf_path)
            ]
            for pt in paths_to_try:
                if os.path.exists(pt):
                    try:
                        from pypdf import PdfReader
                        reader = PdfReader(pt)
                        # Ler as primeiras 10 páginas para obter o sumário operacional
                        for page in reader.pages[:10]:
                            texto_pdf += page.extract_text() or ""
                        break
                    except Exception as pdf_err:
                        print(f"[Dashboard] Erro ao ler PDF local {pt}: {pdf_err}")
            
            # Gerar tese de investimento via Gemini baseada no PDF do relatório anual
            try:
                import llm
                tese_ia = llm.gerar_tese_investimento(ticker, wacc_final * 100.0, dcf_res["valor_por_acao"], upside * 100.0, texto_pdf)
            except Exception as llm_err:
                print(f"[Dashboard] Erro ao chamar llm.gerar_tese_investimento: {llm_err}")
                tese_ia = "Análise fundamentalista baseada no WACC e fluxo de caixa descontado da firma (FCFF)."
                
            return {
                "sucesso": True,
                "ticker": ticker,
                "nome": cfg["nome"],
                "cotacao": cotacao,
                "valor_justo": dcf_res["valor_por_acao"],
                "upside": upside * 100.0,
                "recomendacao": recomendacao,
                "beta_levered": beta_l,
                "beta_unlevered": beta_u,
                "ke": ke * 100.0,
                "wacc": wacc_final * 100.0,
                "ev": dcf_res["enterprise_value"],
                "divida": premissas["ativo"]["divida_liquida"],
                "equity": dcf_res["equity_value"],
                "num_acoes": premissas["ativo"]["num_acoes"],
                "fluxos": fluxos_detalhe,
                "mc_media": stats["media"],
                "mc_mediana": stats["mediana"],
                "mc_p10": stats["p10"],
                "mc_p90": stats["p90"],
                "mc_upside_prob": stats["prob_upside"] * 100.0,
                "tese_ia": tese_ia
            }
        except Exception as e:
            print(f"[Dashboard] Erro no cálculo de valuation: {e}")
            return {"sucesso": False, "erro": str(e)}

    def gerar_demonstrativos_financeiros(self, ticker):
        """Busca demonstrativos na API pública do StatusInvest, ordena e formata os dados."""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }
        
        # Verificar se já temos o link no banco para evitar buscas lentas repetidas
        url_pdf = "#"
        try:
            db_file = "agente_cvm.db"
            if not os.path.exists(db_file) and os.path.exists("agente_cvm/agente_cvm.db"):
                db_file = "agente_cvm/agente_cvm.db"
                
            conn = sqlite3.connect(db_file)
            c = conn.cursor()
            # Criar tabela se não existir
            c.execute("""
                CREATE TABLE IF NOT EXISTS documentos_ri (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ticker TEXT,
                    url_original TEXT,
                    caminho_local TEXT,
                    data_criacao DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """)
            c.execute("SELECT caminho_local, url_original FROM documentos_ri WHERE ticker = ? ORDER BY data_criacao DESC LIMIT 1", (ticker,))
            row = c.fetchone()
            conn.close()
            
            if row:
                local_path = row[0]
                orig_url = row[1]
                
                # Check if file exists locally
                if local_path:
                    # Normalize local_path
                    check_path = local_path.replace("\\", "/")
                    paths_to_try = [
                        check_path,
                        os.path.join("agente_cvm", check_path),
                        os.path.join(BASE_DIR, check_path),
                        os.path.join("../", check_path)
                    ]
                    file_found = False
                    for pt in paths_to_try:
                        if os.path.exists(pt):
                            file_found = True
                            break
                            
                    if file_found:
                        url_pdf = "/" + check_path
                        print(f"[Dashboard] Servindo relatório anual local: {url_pdf}")
                    else:
                        url_pdf = orig_url
                        print(f"[Dashboard] Usando link original (arquivo local não encontrado): {url_pdf}")
                else:
                    url_pdf = orig_url
                    print(f"[Dashboard] Usando link original (sem caminho local): {url_pdf}")
            else:
                # Se não tiver, roda o robô
                rob_res = cvm_robot.buscar_e_baixar_relatorio_anual(ticker)
                url_pdf = rob_res["url_original"]
        except Exception as e:
            print(f"[Dashboard] Erro ao buscar link no banco: {e}")
            url_pdf = "#"

        # Mapeamentos de linhas
        dre_mapping = {
            "Receita Líquida": ["receita l", "receita bruta"],
            "Lucro Bruto": ["lucro bruto"],
            "Custos": ["custos"],
            "Despesas/Receitas Operacionais": ["despesas/receitas operacionais", "despesas operacionais"],
            "EBITDA": ["ebitda"],
            "EBIT": ["ebit -"],
            "Impostos": ["impostos"],
            "Lucro Líquido": ["lucro l"]
        }
        
        balanco_mapping = {
            "Ativo total": ["ativo total"],
            "Ativo Circulante": ["ativo circulante"],
            "Caixa e Equivalentes de Caixa": ["caixa e equivalentes de caixa"],
            "Contas a Receber": ["contas a receber"],
            "Estoques": ["estoque"],
            "Aplicações Financeiras": ["aplica"],
            "Ativo Não Circulante": ["ativo n", "ativo nao circulante", "ativo não circulante"],
            "Ativo Realizável a Longo Prazo": ["ativo realiz"],
            "Investimentos": ["investimentos"],
            "Imobilizado": ["imobilizado"],
            "Intangível": ["intang"],
            "Passivo Total": ["passivo total"],
            "Patrimônio Líquido Consolidado": ["patrim", "consolidado"]
        }
        
        cashflow_mapping = {
            "Caixa Líquido Atividades Operacionais": ["atividades operacionais"],
            "Depreciação e Amortização": ["deprecia", "amortiz"],
            "Caixa Líquido Atividades de Investimento": ["atividades de investimento"],
            "Fluxo de Caixa Livre": ["fluxo de caixa livre"],
            "Caixa Líquido Atividades de Financiamento": ["atividades de financiamento"],
            "Aumento de Caixa e Equivalentes": ["aumento de caixa"],
            "Saldo Inicial de Caixa e Equivalentes": ["saldo inicial"],
            "Saldo Final de Caixa e Equivalentes": ["saldo final"]
        }
        
        def parse_endpoint(url, mapping):
            try:
                res = requests.get(url, headers=headers, timeout=10)
                if res.status_code != 200:
                    return [], {}
                data = res.json()
                grid = data.get("data", {}).get("grid", [])
                header_row = next((r for r in grid if r.get("isHeader")), None)
                if not header_row:
                    return [], {}
                    
                col_indices = {}
                for idx, col in enumerate(header_row.get("columns", [])):
                    val = str(col.get("value", "")).strip()
                    if val.isdigit() and len(val) == 4:
                        col_indices[idx] = val
                    elif "12M" in val or "LTM" in val:
                        col_indices[idx] = "LTM"
                        
                years_ordered = [col_indices[i] for i in sorted(col_indices.keys())]
                
                rows_mapped = {}
                for label, substrings in mapping.items():
                    matched_row = None
                    for row in grid:
                        if row.get("isHeader"):
                            continue
                        row_name = str(row.get("columns", [])[0].get("value", "")).lower()
                        if any(sub in row_name for sub in substrings):
                            matched_row = row
                            break
                    
                    rows_mapped[label] = {}
                    if matched_row:
                        cols = matched_row.get("columns", [])
                        for idx, year in col_indices.items():
                            if idx < len(cols):
                                rows_mapped[label][year] = cols[idx].get("value", "-")
                            else:
                                rows_mapped[label][year] = "-"
                    else:
                        for year in years_ordered:
                            rows_mapped[label][year] = "-"
                            
                return years_ordered, rows_mapped
            except Exception as e:
                print(f"Erro ao parsear endpoint {url}: {e}")
                return [], {}

        # URL base AJAX do StatusInvest
        dre_url = f"https://statusinvest.com.br/acao/getdre?code={ticker}&type=0&future=false"
        bal_url = f"https://statusinvest.com.br/acao/getativos?code={ticker}&type=0&future=false"
        fc_url = f"https://statusinvest.com.br/acao/getfluxocaixa?code={ticker}&type=0&future=false"
        
        years_dre, parsed_dre = parse_endpoint(dre_url, dre_mapping)
        years_bal, parsed_bal = parse_endpoint(bal_url, balanco_mapping)
        years_fc, parsed_fc = parse_endpoint(fc_url, cashflow_mapping)
        
        # Consolidar anos comuns (interseção preservando a ordenação)
        years_all = []
        for y in years_dre:
            if y in years_bal and y in years_fc and y not in years_all:
                years_all.append(y)
                
        # Fallback de anos se vazio
        if not years_all:
            years_all = ["LTM", "2024", "2023", "2022"]
            
        return {
            "sucesso": True,
            "ticker": ticker,
            "years": years_all,
            "dre": parsed_dre,
            "balanco": parsed_bal,
            "fluxo": parsed_fc,
            "pdf_anual_url": url_pdf
        }

    def render_dashboard(self):
        storage.init_db()
        empresas = config.carregar_empresas()
        
        total_docs = len(storage.get_todos_documentos(limit=1000))
        total_empresas = len(empresas)
        tickers_list = ", ".join([emp["ticker"] for emp in empresas])
        
        html = f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Ceará Finance — Hub Integrado de Valuation e RI</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <style>
        :root {{
            --bg-dark: #050508;
            --bg-card: #0c0d12;
            --bg-card-hover: #12131b;
            --accent-primary: #a472ff;
            --accent-glow: #5b9bff;
            --text-main: #ffffff;
            --text-secondary: #8a8d9f;
            --border-color: #1b1c24;
            --success-color: #46e0a0;
            --alert-color: #ff6b6b;
            --glass: rgba(255, 255, 255, 0.02);
            --accent: #3b82f6;
            --border: rgba(255, 255, 255, 0.08);
        }}
        
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{
            font-family: 'Outfit', sans-serif;
            background-color: var(--bg-dark);
            color: var(--text-main);
            padding: 20px;
            min-height: 100vh;
        }}
        
        .container {{ max-width: 1300px; margin: 0 auto; }}
        header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 25px;
            border-bottom: 2px solid var(--border-color);
            padding-bottom: 20px;
        }}
        
        .header-brand {{ display: flex; align-items: center; gap: 16px; }}
        .cf-logo {{ height: 50px; width: auto; object-fit: contain; }}
        h1 {{
            font-size: 26px;
            font-weight: 800;
            letter-spacing: -0.02em;
            background: linear-gradient(135deg, #ffffff 0%, #a472ff 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }}
        
        .status-chip {{
            display: flex;
            align-items: center;
            background-color: var(--bg-card);
            border: 1.5px solid var(--border-color);
            padding: 8px 16px;
            border-radius: 50px;
            font-size: 13px;
            font-weight: 600;
        }}
        
        .pulse-dot {{
            width: 9px;
            height: 9px;
            background-color: var(--success-color);
            border-radius: 50%;
            margin-right: 9px;
            box-shadow: 0 0 0 0 rgba(70, 224, 160, 0.7);
            animation: pulse 1.8s infinite;
        }}
        
        @keyframes pulse {{
            0% {{ transform: scale(0.95); box-shadow: 0 0 0 0 rgba(70, 224, 160, 0.7); }}
            70% {{ transform: scale(1); box-shadow: 0 0 0 8px rgba(70, 224, 160, 0); }}
            100% {{ transform: scale(0.95); box-shadow: 0 0 0 0 rgba(70, 224, 160, 0); }}
        }}
        
        /* Tabs System */
        .tabs-header {{
            display: flex;
            gap: 10px;
            margin-bottom: 25px;
            border-bottom: 1.5px solid var(--border-color);
            padding-bottom: 10px;
        }}
        
        .tab-btn {{
            background: var(--glass);
            border: 1.5px solid var(--border-color);
            color: var(--text-secondary);
            padding: 12px 24px;
            border-radius: 10px;
            font-weight: 700;
            font-size: 15px;
            font-family: inherit;
            cursor: pointer;
            transition: all 0.25s ease;
        }}
        
        .tab-btn:hover {{ border-color: var(--accent-glow); color: #fff; }}
        .tab-btn.active {{
            background: #ffffff;
            color: #000000;
            border-color: #ffffff;
            box-shadow: 0 4px 15px rgba(255, 255, 255, 0.15);
        }}
        
        .tab-content {{ display: none; }}
        .tab-content.active {{ display: block; }}
        
        /* MACRO CARDS BAR */
        .macro-card {{
            background: linear-gradient(145deg, rgba(59, 130, 246, 0.03) 0%, rgba(0, 0, 0, 0) 100%);
            border: 1px solid var(--border);
            border-radius: 1rem;
            padding: 1.25rem;
            transition: all 0.3s;
            position: relative;
            overflow: hidden;
            text-decoration: none;
            display: block;
        }}
        
        .macro-card:hover {{
            background: rgba(59, 130, 246, 0.08);
            border-color: var(--accent);
            transform: translateY(-2px);
        }}
        
        .date-badge {{
            background: rgba(59, 130, 246, 0.1);
            border: 1px solid rgba(59, 130, 246, 0.2);
            color: var(--accent);
            padding: 4px 10px;
            border-radius: 6px;
            font-size: 10px;
            font-weight: 800;
            display: inline-flex;
            align-items: center;
            gap: 5px;
        }}
        
        /* Stats Grid */
        .grid-stats {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        
        .card-stat {{
            background-color: var(--bg-card);
            border: 1.5px solid var(--border-color);
            padding: 22px;
            border-radius: 16px;
        }}
        
        .card-stat h3 {{
            font-size: 13px;
            text-transform: uppercase;
            letter-spacing: 0.06em;
            color: var(--text-secondary);
            margin-bottom: 8px;
            font-weight: 700;
        }}
        
        .card-stat .value {{ font-size: 28px; font-weight: 800; }}
        .card-stat .subtitle {{ font-size: 13px; color: var(--text-secondary); margin-top: 6px; }}
        
        /* Excel style Filters Bar */
        .excel-filters-bar {{
            background-color: var(--bg-card);
            border: 1.5px solid var(--border-color);
            padding: 18px 24px;
            border-radius: 14px;
            margin-bottom: 25px;
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 15px;
            align-items: center;
        }}
        
        .excel-filter-input {{
            background: #111218;
            border: 1.5px solid var(--border-color);
            color: #fff;
            padding: 10px 14px;
            border-radius: 8px;
            font-family: inherit;
            font-size: 13px;
        }}
        
        .excel-filter-label {{
            font-size: 11px;
            text-transform: uppercase;
            font-weight: 700;
            color: var(--text-secondary);
            margin-bottom: 4px;
            display: block;
        }}
        
        /* Search Box */
        .search-container {{ margin-bottom: 25px; }}
        .search-input {{
            width: 100%;
            background-color: var(--bg-card);
            border: 1.5px solid var(--border-color);
            color: #fff;
            padding: 14px 20px;
            border-radius: 12px;
            font-family: inherit;
            font-size: 15px;
            font-weight: 500;
        }}
        
        /* Tables Panel */
        .section-panel {{
            background-color: var(--bg-card);
            border: 1.5px solid var(--border-color);
            border-radius: 18px;
            padding: 24px;
            margin-bottom: 30px;
        }}
        
        .panel-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
        }}
        
        .panel-title {{ font-size: 18px; font-weight: 700; }}
        .table-responsive {{ width: 100%; overflow-x: auto; }}
        table {{ width: 100%; border-collapse: collapse; font-size: 14px; text-align: left; }}
        th {{
            color: var(--text-secondary);
            font-weight: 700;
            padding: 12px 14px;
            border-bottom: 2px solid var(--border-color);
            text-transform: uppercase;
            font-size: 12px;
            cursor: pointer;
            user-select: none;
        }}
        th:hover {{ color: #fff; }}
        td {{ padding: 14px; border-bottom: 1px solid var(--border-color); color: #e2e8f0; }}
        
        .badge-ticker {{
            background-color: rgba(91, 155, 255, 0.12);
            color: var(--accent-glow);
            border: 1px solid rgba(91, 155, 255, 0.25);
            padding: 4px 8px;
            border-radius: 6px;
            font-family: monospace;
            font-weight: 800;
        }}
        
        .badge-tipo {{
            background-color: rgba(255, 255, 255, 0.05);
            color: #fff;
            border: 1px solid var(--border-color);
            padding: 3px 8px;
            border-radius: 6px;
            font-size: 11px;
            font-weight: 700;
        }}
        .badge-tipo.crescimento {{ color: #c09eff; background: rgba(164, 114, 255, 0.12); }}
        .badge-tipo.dividendos {{ color: #8ff0c0; background: rgba(70, 224, 160, 0.12); }}
        
        .discount-badge {{ font-weight: 700; padding: 3px 7px; border-radius: 6px; }}
        .discount-badge.positive {{ background: rgba(70, 224, 160, 0.15); color: var(--success-color); }}
        .discount-badge.negative {{ background: rgba(255, 107, 107, 0.15); color: var(--alert-color); }}
        
        .text-truncate {{ max-width: 220px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
        .btn-icon {{ font-size: 15px; text-decoration: none; }}
        .btn-expand {{
            background-color: rgba(164, 114, 255, 0.1);
            color: var(--accent-primary);
            border: 1.5px solid rgba(164, 114, 255, 0.25);
            padding: 6px 12px;
            border-radius: 8px;
            cursor: pointer;
            font-family: inherit;
            font-size: 12px;
            font-weight: 600;
        }}
        
        /* SLIDESHOW PITCH DECK (Poli USP format) */
        .rel-setup-box {{
            background: var(--bg-card);
            border: 1.5px solid var(--border-color);
            padding: 24px;
            border-radius: 18px;
            margin-bottom: 25px;
            display: flex;
            gap: 20px;
            align-items: center;
        }}
        
        .select-ticker {{
            background: #1b1c24;
            border: 1.5px solid var(--border-color);
            color: #fff;
            padding: 12px 18px;
            border-radius: 10px;
            font-family: inherit;
            font-size: 15px;
            font-weight: 600;
        }}
        
        .btn-gerar {{
            background: var(--accent-primary);
            color: #000;
            font-weight: 800;
            padding: 12px 24px;
            border: none;
            border-radius: 10px;
            font-family: inherit;
            font-size: 15px;
            cursor: pointer;
            box-shadow: 0 4px 15px rgba(164, 114, 255, 0.3);
        }}
        .btn-gerar:hover {{ filter: brightness(1.1); }}
        
        .deck-container {{
            background: #000000;
            border: 2px solid #fff;
            border-radius: 16px;
            padding: 40px;
            margin-top: 20px;
            min-height: 520px;
            position: relative;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
        }}
        
        .deck-header {{
            display: flex;
            justify-content: space-between;
            border-bottom: 2px solid rgba(255, 255, 255, 0.2);
            padding-bottom: 15px;
            margin-bottom: 25px;
        }}
        .deck-header h2 {{ font-size: 24px; font-weight: 900; color: #fff; }}
        .deck-subtitle {{ font-size: 11px; text-transform: uppercase; letter-spacing: 0.15em; color: var(--text-secondary); }}
        
        .deck-footer {{
            display: flex;
            justify-content: space-between;
            border-top: 1px solid rgba(255, 255, 255, 0.15);
            padding-top: 15px;
            margin-top: 25px;
            font-size: 12px;
            color: var(--text-secondary);
        }}
        
        .deck-nav {{
            display: flex;
            gap: 10px;
            justify-content: center;
            margin-top: 20px;
        }}
        .deck-nav-btn {{
            background: var(--bg-card);
            border: 1.5px solid var(--border-color);
            color: #fff;
            padding: 10px 18px;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 700;
        }}
        .deck-nav-btn:hover {{ border-color: #fff; }}
        
        .slide-page {{ display: none; }}
        .slide-page.active {{ display: block; }}
        
        /* Capa */
        .capa-slide {{
            text-align: center;
            padding: 80px 0;
        }}
        .capa-slide h1 {{
            font-size: 46px;
            font-weight: 900;
            color: #fff;
            letter-spacing: -0.03em;
            margin-bottom: 20px;
            background: none;
            -webkit-text-fill-color: initial;
        }}
        .capa-slide p {{ font-size: 16px; color: var(--text-secondary); }}
        
        /* Metricas */
        .val-badge-rec {{
            display: inline-block;
            font-size: 20px;
            font-weight: 900;
            padding: 8px 18px;
            border-radius: 8px;
            margin-bottom: 20px;
        }}
        .val-badge-rec.buy {{ background: rgba(70, 224, 160, 0.2); color: var(--success-color); border: 2px solid var(--success-color); }}
        .val-badge-rec.sell {{ background: rgba(255, 107, 107, 0.2); color: var(--alert-color); border: 2px solid var(--alert-color); }}
        
        .grid-val-deck {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 30px;
        }}
        
        .val-card-left {{
            border-right: 1.5px solid rgba(255, 255, 255, 0.15);
            padding-right: 30px;
        }}
        
        /* Resumos e parágrafos */
        .summary-container p {{ color: #e2e8f0; }}
        
    </style>
    <script>
        let cvmDocs = [];
        let valuationStocks = [];
        let relatorioData = null;
        let activeSlideIdx = 0;
        let sortColIndex = -1;
        let sortAsc = true;

        async function fetchCVMDocs(query = "") {{
            const res = await fetch(`/api/documentos?q=${{encodeURIComponent(query)}}`);
            cvmDocs = await res.json();
            renderCVMDocs();
        }}

        async function fetchValuation() {{
            const res = await fetch('/api/valuation');
            valuationStocks = await res.json();
            renderValuation();
        }}

        function toggleSummary(id) {{
            const row = document.getElementById("summary-row-" + id);
            row.style.display = row.style.display === "none" ? "table-row" : "none";
        }}

        function renderCVMDocs() {{
            const tbody = document.getElementById("cvm-tbody");
            if (!cvmDocs || cvmDocs.length === 0) {{
                tbody.innerHTML = `
                <tr>
                    <td colspan="7" style="text-align: center; color: var(--text-secondary); padding: 40px;">
                        Nenhum documento encontrado.
                    </td>
                </tr>`;
                return;
            }}
            
            let html = "";
            cvmDocs.forEach((doc, idx) => {{
                const ticker = doc.ticker || "B3";
                const company = doc.company_name || "-";
                const cat = doc.category || "-";
                const type = doc.doc_type || "-";
                const desc = doc.description || "-";
                const date = doc.delivery_date || "-";
                const link = doc.link || "#";
                const rawSummary = doc.resumo_ia || "Sem resumo gerado.";
                
                const summaryHtml = rawSummary
                    .replace(/\\n/g, "<br>")
                    .replace(/&/g, "&amp;")
                    .replace(/</g, "&lt;")
                    .replace(/>/g, "&gt;")
                    .replace(/&lt;br&gt;/g, "<br>")
                    .replace(/&lt;b&gt;/g, "<b>")
                    .replace(/&lt;\\/b&gt;/g, "</b>");

                html += `
                <tr class="doc-row">
                    <td><span class="badge-ticker">${{ticker}}</span></td>
                    <td class="text-truncate" title="${{company}}">${{company}}</td>
                    <td>${{cat}}</td>
                    <td>${{type}}</td>
                    <td class="text-truncate" title="${{desc}}">${{desc}}</td>
                    <td>${{date}}</td>
                    <td style="text-align: center;">
                        <a href="${{link}}" target="_blank" class="btn-icon" title="Ver PDF Original">🔗</a>
                        <button onclick="toggleSummary(${{idx}})" class="btn-expand" title="Ver Resumo Executivo">💡 Resumo</button>
                    </td>
                </tr>
                <tr id="summary-row-${{idx}}" class="summary-row" style="display: none;">
                    <td colspan="7">
                        <div class="summary-container">
                            <h4>💡 Resumo Executivo (Agente de RI)</h4>
                            <p>${{summaryHtml}}</p>
                        </div>
                    </td>
                </tr>
                `;
            }});
            tbody.innerHTML = html;
        }}

        // Excel-like Column Sorting
        function sortValTable(colIdx, type = 'num') {{
            if (sortColIndex === colIdx) {{
                sortAsc = !sortAsc;
            }} else {{
                sortColIndex = colIdx;
                sortAsc = true;
            }}
            
            valuationStocks.sort((a, b) => {{
                let valA, valB;
                
                // Mapear coluna para o campo
                switch(colIdx) {{
                    case 0: valA = a.ticker; valB = b.ticker; break;
                    case 1: valA = a.nome; valB = b.nome; break;
                    case 2: valA = a.preco; valB = b.preco; break;
                    case 3: valA = a.dy; valB = b.dy; break;
                    case 4: valA = a.pl; valB = b.pl; break;
                    case 5: valA = a.lpa; valB = b.lpa; break;
                    case 6: valA = a.vpa; valB = b.vpa; break;
                    case 7: valA = a.tipo; valB = b.tipo; break;
                    case 8: valA = a.preco_justo_bazin; valB = b.preco_justo_bazin; break;
                    case 9: valA = a.preco_teto_bazin; valB = b.preco_teto_bazin; break;
                    case 10: valA = a.desc_bazin; valB = b.desc_bazin; break;
                    case 11: valA = a.preco_justo_graham; valB = b.preco_justo_graham; break;
                    case 12: valA = a.desc_graham; valB = b.desc_graham; break;
                    default: return 0;
                }}
                
                if (typeof valA === 'string') {{
                    return sortAsc ? valA.localeCompare(valB) : valB.localeCompare(valA);
                }} else {{
                    return sortAsc ? (valA - valB) : (valB - valA);
                }}
            }});
            
            renderValuation();
        }}

        // Excel-like Dynamic Filters
        function renderValuation() {{
            const tbody = document.getElementById("val-tbody");
            
            // Carregar critérios de filtro
            const filTicker = document.getElementById("filt-ticker").value.toLowerCase().trim();
            const filSetor = document.getElementById("filt-setor").value.toLowerCase().trim();
            const filTipo = document.getElementById("filt-tipo").value;
            const plMin = parseFloat(document.getElementById("filt-pl-min").value) || -9999;
            const plMax = parseFloat(document.getElementById("filt-pl-max").value) || 9999;
            const dyMin = parseFloat(document.getElementById("filt-dy-min").value) || -9999;
            const dyMax = parseFloat(document.getElementById("filt-dy-max").value) || 9999;
            
            const filtered = valuationStocks.filter(s => {{
                const matchTicker = s.ticker.toLowerCase().includes(filTicker) || s.nome.toLowerCase().includes(filTicker);
                const matchSetor = s.setor.toLowerCase().includes(filSetor);
                const matchTipo = filTipo === "TODOS" ? true : s.tipo === filTipo;
                const matchPL = s.pl >= plMin && s.pl <= plMax;
                const matchDY = s.dy >= dyMin && s.dy <= dyMax;
                
                return matchTicker && matchSetor && matchTipo && matchPL && matchDY;
            }});
            
            if (filtered.length === 0) {{
                tbody.innerHTML = `
                <tr>
                    <td colspan="13" style="text-align: center; color: var(--text-secondary); padding: 40px;">
                        Nenhuma ação corresponde aos filtros configurados.
                    </td>
                </tr>`;
                return;
            }}
            
            let html = "";
            filtered.forEach(s => {{
                const badgeClass = s.tipo === "DIVIDENDOS" ? "dividendos" : (s.tipo === "CRESCIMENTO" ? "crescimento" : "");
                
                const bzFair = s.preco_justo_bazin ? s.preco_justo_bazin.toFixed(2) : "0.00";
                const bzCeil = s.preco_teto_bazin ? s.preco_teto_bazin.toFixed(2) : "0.00";
                const bzDiscClass = s.desc_bazin >= 0 ? "positive" : "negative";
                const bzDiscStr = s.desc_bazin ? (s.desc_bazin >= 0 ? "+" : "") + s.desc_bazin.toFixed(1) + "%" : "-";
                
                const grFair = s.preco_justo_graham ? s.preco_justo_graham.toFixed(2) : "0.00";
                const grDiscClass = s.desc_graham >= 0 ? "positive" : "negative";
                const grDiscStr = s.desc_graham ? (s.desc_graham >= 0 ? "+" : "") + s.desc_graham.toFixed(1) + "%" : "-";

                html += `
                <tr class="val-row">
                    <td><span class="badge-ticker">${{s.ticker}}</span></td>
                    <td class="text-truncate" title="${{s.nome}}">${{s.nome}}</td>
                    <td><b>R$ ${{s.preco.toFixed(2)}}</b></td>
                    <td>${{s.dy.toFixed(2)}}%</td>
                    <td>${{s.pl.toFixed(2)}}</td>
                    <td>${{s.lpa.toFixed(2)}}</td>
                    <td>${{s.vpa.toFixed(2)}}</td>
                    <td><span class="badge-tipo ${{badgeClass}}">${{s.tipo}}</span></td>
                    <td>R$ ${{bzFair}}</td>
                    <td>R$ ${{bzCeil}}</td>
                    <td><span class="discount-badge ${{bzDiscClass}}">${{bzDiscStr}}</span></td>
                    <td>R$ ${{grFair}}</td>
                    <td><span class="discount-badge ${{grDiscClass}}">${{grDiscStr}}</span></td>
                </tr>
                `;
            }});
            tbody.innerHTML = html;
        }}

        // Gerador de Relatório Poli USP
        async function dispararCalculoRelatorio() {{
            const ticker = document.getElementById("select-ticker").value;
            const statusArea = document.getElementById("rel-status");
            
            statusArea.style.display = "block";
            statusArea.innerHTML = "⏳ <b>Processando WACC, DCF e Simulação de Monte Carlo (10.000 runs) no servidor...</b>";
            
            const res = await fetch(`/api/relatorio?ticker=${{ticker}}`);
            relatorioData = await res.json();
            
            if (relatorioData.sucesso) {{
                statusArea.innerHTML = "✅ <b>Relatório calculado com sucesso! Navegue pelos slides abaixo.</b>";
                montarSlideshow();
                document.getElementById("deck-box").style.display = "block";
            }} else {{
                statusArea.innerHTML = `❌ <b>Erro no processamento:</b> ${{relatorioData.erro}}`;
            }}
        }}

        function montarSlideshow() {{
            const d = relatorioData;
            activeSlideIdx = 0;
            
            // Slide 1: Capa (Estilo Poli USP / UFC)
            document.getElementById("slide-1").innerHTML = `
            <div class="capa-slide">
                <div style="margin-bottom: 25px;">
                    <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAMgAAADICAIAAAAiOjnJAAAAtGVYSWZJSSoACAAAAAYAEgEDAAEAAAABAAAAGgEFAAEAAABWAAAAGwEFAAEAAABeAAAAKAEDAAEAAAACAAAAEwIDAAEAAAABAAAAaYcEAAEAAABmAAAAAAAAAGAAAAABAAAAYAAAAAEAAAAGAACQBwAEAAAAMDIxMAGRBwAEAAAAAQIDAACgBwAEAAAAMDEwMAGgAwABAAAA//8AAAKgBAABAAAAyAAAAAOgBAABAAAAyAAAAAAAAACKGshfAAAACXBIWXMAAA7EAAAOxAGVKw4bAAAFTmlUWHRYTUw6Y29tLmFkb2JlLnhtcAAAAAAAPD94cGFja2V0IGJlZ2luPSfvu78nIGlkPSdXNU0wTXBDZWhpSHpyZVN6TlRjemtjOWQnPz4KPHg6eG1wbWV0YSB4bWxuczp4PSdhZG9iZTpuczptZXRhLyc+CjxyZGY6UkRGIHhtbG5zOnJkZj0naHR0cDovL3d3dy53My5vcmcvMTk5OS8wMi8yMi1yZGYtc3ludGF4LW5zIyc+CgogPHJkZjpEZXNjcmlwdGlvbiByZGY6YWJvdXQ9JycKICB4bWxuczpBdHRyaWI9J2h0dHA6Ly9ucy5hdHRyaWJ1dGlvbi5jb20vYWRzLzEuMC8nPgogIDxBdHRyaWI6QWRzPgogICA8cmRmOlNlcT4KICAgIDxyZGY6bGkgcmRmOnBhcnNlVHlwZT0nUmVzb3VyY2UnPgogICAgIDxBdHRyaWI6Q3JlYXRlZD4yMDI2LTA3LTI3PC9BdHRyaWI6Q3JlYXRlZD4KICAgICA8QXR0cmliOkRhdGE+eyZxdW90O2RvYyZxdW90OzomcXVvdDtEQUd2MHZRTU85RSZxdW90OywmcXVvdDt1c2VyJnF1b3Q7OiZxdW90O1VBR2tiLUhraS13JnF1b3Q7LCZxdW90O2JyYW5kJnF1b3Q7OiZxdW90O0JBR2tiM2VKbU13JnF1b3Q7fTwvQXR0cmliOkRhdGE+CiAgICAgPEF0dHJpYjpFeHRJZD43YmYzODMxOS05MzMzLTQ5ZDMtYjg1MC03ZjNlYWE2ZmIxNDM8L0F0dHJpYjpFeHRJZD4KICAgICA8QXR0cmliOkZiSWQ+NTI1MjY1OTE0MTc5NTgwPC9BdHRyaWI6RmJJZD4KICAgICA8QXR0cmliOlRvdWNoVHlwZT4yPC9BdHRyaWI6VG91Y2hUeXBlPgogICAgPC9yZGY6bGk+CiAgIDwvcmRmOlNlcT4KICA8L0F0dHJpYjpBZHM+CiA8L3JkZjpEZXNjcmlwdGlvbj4KCiA8cmRmOkRlc2NyaXB0aW9uIHJkZjphYm91dD0nJwogIHhtbG5zOmRjPSdodHRwOi8vcHVybC5vcmcvZGMvZWxlbWVudHMvMS4xLyc+CiAgPGRjOnRpdGxlPgogICA8cmRmOkFsdD4KICAgIDxyZGY6bGkgeG1sOmxhbmc9J3gtZGVmYXVsdCc+RGVzaWduIHNlbSBub21lIC0gMTwvcmRmOmxpPgogICA8L3JkZjpBbHQ+CiAgPC9kYzp0aXRsZT4KIDwvcmRmOkRlc2NyaXB0aW9uPgoKIDxyZGY6RGVzY3JpcHRpb24gcmRmOmFib3V0PScnCiAgeG1sbnM6cGRmPSdodHRwOi8vbnMuYWRvYmUuY29tL3BkZi8xLjMvJz4KICA8cGRmOkF1dGhvcj5BbmFsaXN0YSBDZWFyw6EgRmluYW5jZTwvcGRmOkF1dGhvcj4KIDwvcmRmOkRlc2NyaXB0aW9uPgoKIDxyZGY6RGVzY3JpcHRpb24gcmRmOmFib3V0PScnCiAgeG1sbnM6eG1wPSdodHRwOi8vbnMuYWRvYmUuY29tL3hhcC8xLjAvJz4KICA8eG1wOkNyZWF0b3JUb29sPkNhbnZhIGRvYz1EQUd2MHZRTU85RSB1c2VyPVVBR2tiLUhraS13IGJyYW5kPUJBR2tiM2VKbU13PC94bXA6Q3JlYXRvclRvb2w+CiA8L3JkZjpEZXNjcmlwdGlvbj4KPC9yZGY6UkRGPgo8L3g6eG1wbWV0YT4KPD94cGFja2V0IGVuZD0ncic/PuXEA3cAACAASURBVHic7V13nFTV9b/WpQRpspTQl116pPeyAgGCgQSEDQYIPQooCChVglJESEAioUjoTYFQBHRd2m6oIhABEaR3lL40WSx7f+c33887n8uUN++9aW925vvHfGZnZ94t53vPOffec88VMoooAgAR6gpEkTURJVYUAUGUWP+PTA2/OPDzzz//9NNPPxoAfY2+jF/xQ0LdGlsgEonFHAKB6NW/z+fHMtv8+/ywQKQQi5nkScykfq5fv37o0KHU1NTVq1fPmTNn4sSJb+ti0qRJc+fOpS/TT+iH9HN6iKfSmWeBbKWNkMWJRbIk5UGvTh/evHlz3759H3300dixY3v27Nm8efOyZcvmypXrySefFJZAP6Sf00PoUfRAeiw9nIqgglxLd61S1kMWJBbUAwlP/TAjI+PAgQNTpkxJSkqqWrVqnjx5Hn/8cU8seeyxx5544gniylO6oC/Q1+jLnp5DRVBBVBwV+v7775NWe/jwoVorWMwsqcayDrHAJ9VhIq1w/PjxWbNm/fnPfy5durQrk4gTTzpA/KD/PqbBrLriH9JDwEiC63PoX3FxcZ07d6YqUcVUpYWaZyWGZQViwXmSDm7R6507d9avX//GG2/UqFGD9IqTaEnkzCGzBLJGOCoO3FX/RRWj6lElqap3797lysMVC21/+gVhTCySBJkSHuXkOKelpb3++uulSpWCUFUyBYdJ+kAdVJKhSlThgQMHUuXZ93dqWjgiLIkFq8d/fvPNNyNGjEhISFBF6NUBCi3YjVNrSL7/yJEjjxw5wk0LX/sYZsRSlwxu3bo1b968pk2bsr0Dn3S8cnsCtpIZRs1p1qwZNS09PV0qSxWh7ntzCBtiqZSiOTzN70qUKCE0awLvO5Ts8Afg+wutUdRAaiY1VoYhvcKAWCqlLl++PHr06MKFCwvNmtjZ3lmDU7uosWPHjr1y5YoMK3rZmljoR6bU8OHD8+XLx12fBVSUPqDAQK8CBQqMGjWKOsGpW2wL+xKLh+aFCxcGDx6cN29eETGUUqHSi8bVkCFDLl26JJVFFnvCjsRCsIB0rCDMmDGDBquISEqpUOkVGxs7c+ZMbC1wX9kN9iKWquR37txZp04dEfGUUqHSq27dutRF0mXxxSawEbFYt1+9evWvf/0r5kdRSrmCJ4/0Sh117do1ab8le1sQi8ccvVmwYEGRIkXU7ovCLXjIUXctXbpU2kx1hZ5Y3B0nT55s3ry50GxfFltECATUjmrZsiV1oLTNYn2IicWxbwsXLsyTJ4+I2j7zYNVOE+fFixdLLaoxtJINJbGgqG7evNmpUyehjb9QiylcwaqLOhOL9aE1i6EhFnbv6c0XX3wRFxen9ksUlsEjs0yZMtSx0hFIGCqzGAJisVNF5i9nzpzCEdcbaqFkHaAzc+TIsWTJEhk6lyvYxOKDK4MHDxYO/yDqUfkd3KvDhw/nI0lBFnRQiQVFdefOnXbt2omo+Qsk2CxSVyNCNcguV/CIhYZdunSpXr16wqGxo6wKKBCdRm+ow7/77jsZXG4FiVhw1Q8dOlS6dGkRdaqCCHR1fHz84cOHWRBBQDCIhcbs3LkzNjZWRFkVdKDDqfOxtxgcbgWcWGjG5s2bc+fOLRx+Vaj7ORKBbs+TJ09aWpoMCrcCSyw0YMOGDb/61a+EY7YS6h6OXKDzSRAkDhl4bgWQWKj6J598EhMTI6KssgEgAhIHCUUGmFuBIhYqnZKSgiXQKKtsAgiChELOiQwktwJCLExrt27dmitXLhFllc0AcZDLu23bNhmwNQj/EwsV3bdvHw4+RFllQ0AoJCASkwwMt/xMLFTx1KlTJUuWFNE5oI0B0ZCYSFgyANzyJ7GwIXX9+vVq1aqJKKtsDwiIhIUwG//uJ/qNWNhCf/jw4W9/+1sRXQUNE0BML7zwAgJs/BgH4TdiQZf2799fRFkVVoCwBg4cKP1qEP1DLFRo8eLFImoBwxAQ2aJFi6T/uOUHYqEq//vf/7BpE41ZCDtAZHny5CEhSj9xy1diwSqnp6dXqVJF2FhdIekZ53F0SuqHvHtIJ8TJIxEuZoeMbb7D63kCduTV9IKhJBbY3adPH2FL14ozt/j4HPukBbQAg9VGL7366qvSHzNEn4gFVq1Zs0bYbyHU6bxrtmzZypQpQzPW3r17jxkzZu7cuStXrvz000+THVi1atWSJUtmzpw5YcKEoUOHvvLKKx06dKhXr15cXBwZCCep4Ml2a68noBNiYmJGjx793HPPCV1J4V/YSfTRIFonFkh97ty5YsWK6Vc3yFAplT179saNG3/wwQcHDhy4deuWKQ1PPXv9+vUTJ05s3ryZntCpU6fSpUtjQ51h5wRdsOzCMaiWL19O8ipevLjQdYIhxBIlSpw/f176prd8JRaNbGEb10o9GUzKZsqUKYcPH1ZHHg4I8R04jJ80qNfjuDaZXMl9+/YtXbqUVBoNJ1VCaroOO4AdgAoVKmzfvp0q/+9//1sYGP/4FYlVhoRYkBZ1sZG6BgesqGhcTps2jTMsWr5uhC9dAvOcepn039atW//xj3+0bdsWubu4GpgZ+Ngca09QfUpSrgMHDrx69Sra0qxZM2FMBUCgy5Ytkz4YRCvEgoQuXLgA1Ro0Yuk4N+gv0vn9+vXDwQHX+wR8B85RkVZzeuzFixfJaevSpcuvf/1rtaqmyMGTVrWBPFFFw9VbDlSgODV/ad26dREYgxTfNAbwuZEqoQIlS5akdkmrM0QrxMLY7dy5swiKEfQqJNShbNmyn3/+uXS5rMa/xzX5aU6aDJ+TT0Y6rGrVqkwOr9rLXxlQeC2KNNOcOXPAJ7b4iYmJwowKQJd27dpVWjWIpomFwUoiFIFfC1X1E43a+Pj4IUOG0JxOXdfAF9q1a4c0USG5/8iJyhkZGTTfbNOmDcLRoFFcW6euLdEXihYt2qJFC5qTLlq0aNOmTTt27CA1Q87G3/72NxrDNJ+tXr069QApEtKLBQsWzJcvH9Eof/789CcNqkaNGlHn0FRDrRLoNXnyZGFeBUC4GKsWFL9FH6t+/foW6mq8Seogpr7r06dPamoq3KaXX35ZaHxCBbp3746D5JBuCJP4qBkJ6fXo0aNUN9VCMfiw8jPPPNOrV6+UlJTLly/r15xYcvv2bfKZyEKdPn2aOHTs2LGTJ0/Sn3fv3uXfIn8dTDb9uWvXLua3KSmg2iRoa11hjlhgrsH5hQU4Wb2GDRvS8MXUV10z49zu9NqkSZMHDx6wRO2QGgoU51G+c+dOMo5qj+ENeYSkn4giUrOk9IqLW8EMnq4ayb/APiU7gvQhMRsXwFgTFn41b948aT6I2QSx0DayOAji868dVCmVPXv2l156KTk5me+WoTdU+o0bNypWrCiU4V6mTJlz587JUKfs8QTYI3ozYsQIoY0E9ghxEkslhD57Mh+F033B/Fu+Um/Pnj3UP8IHFcCdTFNgadJbNUEsCO+dd94RfjWC6vycXIfevXvv379fKlMwFs+ECROElvEB5nLdunVSo53xhgQNqNXZs2exhszOVs2aNaGGoZz8VXnOoEyvixcvfvbZZ4XPhgWCHjdunDQ5eo0Si7P4FypUSPhJXampZooXL06+J1YKpDbW1bFI9s5JXQ0fPlxqKaDsSSxIgrxvoUxEqBVI1O6v5FWqzqM/yffCqrVfMvlA0CR0vrvAYK2MEgt9NHr0aOEPdaVO9+Li4v75z3/SRB0FuSYuR9GrVq0SykzqN7/5zZ07d6S/A2r9CB4VNF8TipixE+fp9mgjD4Qud3W/SPbDhg2j2YAwsMxhSlj0SqKXZpSWIWL5UV2xIaP3lSpVWrZsGeZ60vNKAUpv3bq1UBgJj9LpJlxbATIgR+rpp59m8ZARvHfvnjQ89DMNXJ5DX6CpX8+ePZFtGgbXjx6wNaVliFjoozFjxghL6opdItZSiYmJK1asgMrRn9ChT8ltx8VMeEJsbOyFCxekjdWV1DrtlVdeEY5FOKy90TRQGpthObnkeCD5AzR5orleamrq0qVLx44d27179+rVq4O7wpKiMmIuIXQigDSstLwTK1O7HNDaZFClFHVuixYtyKhB07AjpVO60yoDnjNq1Chpb1ZJrXrdunVDw1FzZDU2Ihv8nGZ2gwYN6tKlS8uWLevUqZOQkFCgQIGYmBgnKSAVlgUtBcbkyJFDP5YOTyYCGJ8eeicWhtfMmTOFSXWlhq9Q1X/3u9/RJA51clqq9lr6gAED6CEYlzTZwRKDPR12hrrxBbGR90PKRhobEiAfuQrCg1Lh/UHLhg/SqV279o4dO7Aw4XXfbMaMGdKYxvVCLAiPPE2vMWJOleAq5s2bNykp6csvv+QHmlrJxNeIlEIjVseOHaVdF65UoIassYQj0wuynxnUtfgaTfWFthutbkJbYBKDk/01aNCA6rl+/Xp9VglN9EQDsMqr+LwQC71D/pAwwCqnrZjSpUtTp+C6BGnpGjS2wlhowBW9H374obQ0qwoy0NgePXqAWBjx06dPl2ZWsfEQ+hVfgeYLn1hMYFWzZs0wGW/fvr0wYI5QOnky0sDA9kIsDBqvBTttxZBvTsXzdM/gyrIr8H0yfOhW4Rj0Bw8eNNKwkAM1pMmaUJz3/v37SzOV59FI47NChQrC5wVP9k9I8cNhOnHiRP78+YUB1uKHRAZpQOnqEQs/PnXqFHlIngpWHals2bKRS7Fp0yb45kamyl67FRXAwTLhyNNqnz1BfUAtYXWUNVaTJk3MLrVjA1E6Mo0J5ViHZUqR70/zu4yMDNRw2rRpwpj3jEJz5sx55swZ6Y1besRSN1KcZg1OywfFihUj/xqn0qS2C+v7rA1P+Pbbb5EQUDgOg0vlBh47A5U/e/askz7YsmWLNK9xIYuxY8dyn8OiqeuCntigOvjVqlVTAwAvX76s3tnuFeDf3//+d+nNoHskFiRH7UeGD3VnXo0iqlKlyvjx4xEL5bSf6rvsIZsjR46QykTLybJI21xwpQ+sj9Pr888/L5QgdFLq0nwTMrVLADZs2NCuXbuyZcs6DXV4909o4OOT/IX4+PhJkyaRopIOToBYEydOFGYm+3ggUeJn7RpATxX2SCw0Y9euXRz5pBZPKoSUx9q1a3/44QcZsNvVUe/z588XLFgQ5aakpMhwcLAA1yUDvE6ZMkVqe+em6MU9TO7Rtm3b3nrrrYYNGxYqVEiHGXny5CGXd/78+bdv35baQg8qRs5r0aJFhZk5Ab5JrcBdPToS90gsKDqqulDsIDWAJpyTJ08+evQoh+QGTn/wvhgia0lvkVmUtl/BUoGqJiUlCeXOBHJGkWHWQlyGGumFP2ngEckWLFgwevTo7t27d+jQgfzrLl26jBw5cvny5ceOHeMZNBsTPAGXrpndSgEZRowYIXWtoXtiZWo5icqVKweGli9f/p133tmxYwdUlDS2bu47UHVM2slZQfhDWPhYAJ++TEhIEI4VEwiS1Mx///tfqe2QGl/VUzehDfKSo4/wW/Dsgw8+EJbmmByjged4qoB7YqE70tLSaIo7dOhQMkAPHjzg/wbC6nkCaj9kyBDhCI5DNcKFVQDUw/Hjx2vUqAFrgpXevHnzYitdatF5FtrlNthB/USVFEeWJicnW4tX5p/QK64j8MQEPY1FUwbsE0ttohf8++8gld27d1Nj6tSpox6JCRdkapczkmP05ptv8iQREiKb9c0330iNIoGrBsdLEqtQB8tLYrCG2LH1ZA31ZoWQIp/2DMnqEUqk2hcvXpxmQzLcWCUVpYL+PH369LBhw0h7kcZihfHiiy+mpqYa3C2xUAH2g0lHYlXSl4VW/LZBgwb8fNdC9Yhlk1k96vCnP/3J71nngoxM5d5AsunkVpPymD17NrWrfv36lSpVotmifwewutl/8+bN3r17C8/H0YyDjyacPXtWWiCWTeJSUA2a+GCX0Ca1sgw2SU6gWdGVK1esZfTPVIBdIFUp0Cc0Z8S9a/6KAQQ158+fLz1YQ711LDuoKwZ5ncgJZjdYG4HMAGueK6++6v/w/PnzU6dOJd+UV9H8FVkKN6tXr17Sgw0JG2KFxMMzAn/NJ0w10Omb0IKk8y5evLhr1y7STyNHjmzVqhVnK/HLwQoVeFq5cuV4X9iphh6XG4wMiCDDVpUBoDnMZt7yvVDpODhP7nPz5s2bNm1Kb2rWrBkXF/fMM884EShA+ZXwwJiYGBzqd9XZesSStpSlTcDr1yNGjMCswhfnz/gYRinXrl3DNpdbxnAyo4Am1wCDFy5cKN25WXrrWFHoAKzasGEDde5//vMf6fN01fXcm6ev0WtqaiqnPXLNcBQ4MqmAmzVgwAAZJZa/gP65c+dO5cqVqXNx1ZHvkWfSQJwt/rt27VphOGttgIC9qcaNG6t9wnBDLPssNNgWbASFYzkHgWjWOg2/WrZsWd++fW/cuCFdTjY7AbqBdKQIdUp9lF64cOH79+9LI8QyqJMjFryAjuQI2bJlw56MtU4DR2fNmiUcWWWJrHv37pXKgoLT99Wwz6DlUtQh1lNPPXXkyBHX5rvXWFFT6AmZWvxjr1690L+ksYyE6noCfkU6j+O/iS7dunVDmkanZ7IxefHFF4UNcgqD2R9//LF0cTHda6wosTwBcj106BAfPq5QoYKPMRf4IQJN6bEcpclxuU7fJKVVq1YtEWqNJTT//e2335Yu/nuUWOaA7luyZAl3a/369X30HPBMBEhxSDFpI9hE9eGQS0ZGRtmyZYVtiMXx4mqjosQyB4i5bdu23K1FihSxkJfM9Znp6ekcHU+vw4YNky6m8Gftsm3jKZADClS1efPmro2K+lhW0KRJE+5Wct5xKNcXvQXG3L59m3ThhAkTkpOT3Trv+BoSDtjh5iKozPLly7tGkzoTK7rWoA/0nZqMPyYmxnhGBp3HOk3G3Ybp4hNE0IfccxeaysydO7erznYmFhboohrLE5yIhZ5dvXq19G3lXY144Sh4Jyngz++++w7XFITcDgrdFQc3xIpqLB04EQtqo2/fvjLwgWK/aImNYIDsQyx63bVrl9QnVtTB0gdmcMgAw2fBa9asaS1AzxQClFzYR4Bbn332mXxUZ7snVpRbnvCzdk+xei0A+dE0U5OBDJvO1PKK1a5dW9hgoYGBmuA+aXV0uTeFUWJ5ArT9gwcPXnrpJaHcCUjzxIC6pyj3xIkTOtEyIQEG2L/+9S8ZJZYvyNTOch06dEjdhKE3c+bMkQFL3IXHTp06VdhjoYGBykyYMEHqEys6JfQKnsEhzyAfnI+Njf3qq69kYC6KgpG1z0IDA9nwyPOTjw4qNxorSiyvAG+OHz+OGxvZiy9VqhQm3v69LIPtIHK428cOCo1Y2C7UI1Z0gdQgoPbJZ8+ZM6dQsjuVLFly9+7dUslj6PsKDtTV6tWrhc1YJUwRK6I0li+NBbcWLFiALuZbM/LmzTt9+nSnGaIvpgCPGjRokLCZHRRRYnmC5fbihxD5hx9+qCajg1ls3Lgx6Zjz589fvnx5586dli/RyNQSVyMM2j4LDUCUWG6Qnp7+/fffS9/0FnLkjRo1SmhTJPUMO85j9enTx3IpMKP79u2DCMPbFGZ5bmVq6XQGDhyoZmiy8Bz0ZseOHcWjdoqXuBITE/VTSekDBnf8+PHCfnZQGCdWMHNfhRa/aJnGeQ3GrOB5TQuXLrvqEnWL2to88RcNiACzmx0UxpcbQkUsT2cHAodftLtu4uLiwA9TDWddtXjxYrf5sXnn35czPKjYlClThC1ZJYwvkIZq5Z0PXgetaJRI3g/1y+zZs6XJa4/x5XXr1nlKN8XbiK4RxmYr2bJlS2FLOyi0Wk2bNk3qEyskPhbmVlu2bEHu2uAwWyVWjRo13B6O8wRUeNu2bfny5RMedAnfiL5nzx5piVj4yf79+3Fww25uO4C2L126VBqJbjDbBb4gUztV3KFDB+SuDU4FIDbcJ0jiR1p9r+EJ7FcRq9QrFF3Bmat5vdRU9bigvn37CruqK8bnn38u9YkVfFOI2sycOZPUBj4JPrGEdqmYV70CYS9atIgX3HW6G/+dOnWqNGlnpdYtly5d4ntTA08P0zAR6BcqH6t69epVqlTB+5AQq3Tp0l4P24Ac5K0jKNmrN40v0OQAZ+eNW0OeGQwdOlTYWF2BWGSpjx07JvWJFeS9QoxLciOofnXr1g1m7iSU9fLLLwttFxmJ1z01nxNZ42Ifg3M0fG3IkCHSwUv9RQ12cLHounHjRjWc0IZAxcjR9H6YQgbXzYIOGDlypHDc7OXLQqJZqMTCYoxrGKRTPQ8cOAC/yuxNs8Kx0oMS1Wzs/HwOxXn48CEG244dO2AE7bnKAKBuZGpcr9YJ8YFV9DW57cJxqNxT3sHAFY0UDJh2IZ2/K7HQa4cPH0Z+WAsJzSCANm3aEDXd8gkfZmoHnefPn68z37QPMMCoXa5Sc6+x/CtCHaCs1q1bC4dG9SW7hlmglK5duwpNY82dO1c+SiyemtHMDoeurLk77JDFxMTQLIHs6cWLF522pe/fv0+eysKFCzkvg20tIAOro/369ZO2OmLP5qBFixaoKN9lHegK8BpH+fLluYNciYX35OsgS6yPTjSrOiINmbmaNWu2atWqffv2pLBbtmxJBkXNRWt/Vgmt39577z1pJClI0HwsLggry4RBgwZJf4dfugX7MSxp4UIsvPn0009xv6tfpmZO12a7Qr0O0v5AQ9auXSsNaqygTQydNBbyhgdaY7GBQ148jqBCnlZ41phG0DwxQL4OjOMTCsJFSzF4rYH3S9RODnGqSCdiNWrUCKwKDrFwqRhHrC9ZskRqIcXSMQcsVqyYsPEyUmgBYpHraTRVpOuXAgRmMI67UEVz5sz59ddfywD77yAucYu8HKG4PsuWLZNa4N7+/fvBKludtbIVdHIYSR1iBYFbXArNV4U25588ebI0vwFiClBIR44cwVInq3RcRysd3jryi9p8th9aYMjBLTaRjjuYi0k0MxIascgsygCrTBDrL3/5i1DsYHx8fHp6unTsA4JwUQuoD50EpFKHWMFxs8D06dOnCyWPWVpamgxYHgRQ9saNGzgSyLMwLMaMGjUKTjRfsR5VWm7Bl1ufPn1aGr/yhIkVaL2FUi5cuICbZLFQmZSUJLW5oX8rkKmdq5kzZ45QlotIq5P5e/XVV/lD5lZok/TbFuicypUre7q8M/S3f6GU3//+90Lzo0nMrLT8TizpyF1WqlQpoWzuUokFChQQGqtYUdWvXx+fR7nlBOvXyqnaIqAuF6o1Y8YMoaQya9q0qQwAsbA65SnFlHqpH/137Nix9OX169fzoa4gC8/OsH4RplSuVg8osfDk69evc4ppiHDNmjXSr54WHrVnzx7OEqP2FEcS0ys57x999JHUugzJZKO+PMNvV/ci81PgjCPkRxpCKP5yuXLlbt686aneZoGHnD9/XidDOqhTuHBhZFFDfAu9uXjxYqFChURUaWmwftk4KPXFF18kJydjXZUfge72L8lQ3MmTJ7F6BDeL3owcOVL6wyCyz96zZ0/hTvdwiYmJibjZ0WnHEPcxRZUWgL6iGbT0vOLonlgQw6ZNm+j3ZcqUIXmQD7R//3410uMX7YZ03xnGe3MTJ04USsYpms26Jre0ADycrJvwoKvQUy1atMAtI2pnoeijR48i9CCqtLgHEHDrSTR6Wzokj0qVKglNGDRea9asSQZr8+bNV65cUb/vow5jjfL999/Hx8cLZWO4ePHip06dkj5s8jAz3B6q4RQxf/zjH+/cuSNdhiDXjYOYgypG+wEdWLFiRf1wX48+lho0HBMTo8qDOrdIkSKdOnUiHXDmzBn1YDWnKTfLM94YXrlyJdceiqR79+7SErGYEzQzqFq1qvDMqq5du5Iy5u87AUWnpaXZ5KKR0AJCId9A6u686a1j0euuXbvUdUJeDmDkz5+/YcOGNIE/ePAgtm/VJyBSwDjJUCgpD6HsDVMR1pQW+E0V4AeqNeeFUNLBmbrp0VD5S5cuQedFMrE4fA37qjoS0ZsVSocGqlatmnh0rEMeuKSKPyR/qFatWuSNTZs27cCBA64pXDK1i0NBNbeA8vvqq69wao81yoIFC6T5AECMJ7erVvgzNjaWFzW8spa+07x5c9dHRRQg8erVq7uennCC3joWBDNhwgThOXrErRp76qmnEhISkpKS6Ldr164l/+aHH34wQgW3vh0VvWXLFmnYheeoGHr/8ccfC3e6SjiO1SNEx0hcIZ42evRona6IBKAnaY4lvUWg6BELg5jMEOsPnSLVkEinf5GLRlPL1q1bv/nmmzS73LBhw86dO/fu3Uua6ZAGek+f7NixY/ny5R07dlQjhsmfM5UeTT0BQWbUqeZ4/4c//AETQHKtjFhqdaUtYomFriMyGDnzokcs+WhYi0ETwDu4IJlbOsKSPq2AlJyTbeUSsb1jkFJSU2wXL17EWqhabVSmUKFCfKmQwbRNINaYMWNEBBMLPUlk8Moq6ZVYEJI6U7MAlWdeI7t5oYH/TE5OlsbsIJ/RI8tLdBSeBwNxq02bNrwSE9VYRmDqqjMvxMrULiF+7rnnhP8iKh/zDP4Oe0Kmjkejwf369RO6KpajRpExzCu3osSCOIgGnuJkzBGL+5SjD4LWEshv+PDh0likMrtWM2fOFN7GgBobgyUZfd0eJRZET31rUBzeiQVu3rp1q2TJkiJYqzgopWjRoufPn/cqdQCKjaYFnlLsuQJml8pKTU3V7y/866233hIRSSyIgwhg/PZr78SSmn2B6xocpaXGkRlZDgDztm/fTtNPU5VEQe3atZO69EUP9O7dW0QksdCf48aNk4YXfQwRK1NLXh3k6BF22404QOvWrbOQCR1fLlCggE5oET6kCYHrWnEkgKfSpvItGiKW1HiKFcJAKy08v23bttJwir1vv/3WWtIOzvCks2uErrxw4UJsbKyIvC0ddCmJXpoJMzFKXNiMqAAACdRJREFUrOArLSN3luK/x44dM2sBGayxsOjndjiilFWrVonIYxXaW6RIEbPpYY0Si/v37bffFoFUWnhyYmKi9HagA//NyMho1KiRsOr6oLjq1at7yprMIQ89evSwXEr4Av3DizLG2WKCWBycHhcXJwLmamCIuGbhda2M77GdHDjqmqacAeN49epVZF2LKI0FEZM1sBAjboJYLGlcpBYIYoEfOAytn/TGqSbW5A1nv3bt2jqZBNXN7Ehz29HexYsXS/NBvOaIxXj++edFAAwiWoITMp5awrpq165davIFI1D3MVEWufwHDx70VBzvEcHaRhSxIFwStDWGmCYWBLB9+3a/t4TPfuhf7Qc1duXKFdc9Zv2HO4WUCUemFNc80q6NTU5OFhFmBAHqsR07dkjz6kpa01gQQ//+/YVflRYehexnnhbB2ZXGtdsWXOkcOXKQ7Rs5cuSWLVsQ8qrTa2gpUqRGVHwfGjtgwABp9bSBFWKxF5+QkCD8ZCA4RP/27dvSs58IEiD20FSm9caNG69atWrz5s2nTp1iJumHI+Nru3fvjrST0OgxEi6JWEcW+rDoY/klnEYFhsiYMWOk5z0ckGDdunWmLi1SbxHinXlO2+cWmUqKbDvfvBUgoMdIuNKSEQQsEktaigH0BFCEfKa7d+9KD7oXH545c8ZsWn11akO2D4zRH4V8znH+/Pm+ty68gMYavFlIB74S69y5c0ip6IveQmN69uyp0xidkxFGHm7qRkIwj7hVv359H5sWXkBLS5QogaASX44iWyeW1PTkmjVrhD96PyUlReqqq8OHDyMztil3B8QydQUX2oUgmchhldAa+8knn0gfjCDgE7G4eFwn6cumCs28pOchAmJhU8WsYUKtcBu2EWLBCG7fvt3O108GAmpmQ9/zOfpKLFAhPT0dR40tuCNqpiVPq5T0evr0aQvqSmjEGjp0qDRALD7pj1DsyPGu0NJq1arBzfU9H4evxJIauw8ePOh61sornGJ93A4UCPvdd98Vjk0Ys8RClyH9pP5A5FONfpmRhBEwtvPly4dNCL+kn/UDseSjgSWmnBII77XXXpPeVhk2btxYtGhR4bKG7hUgYq5cuXQW2eGtoxVUGRFJrOIzLAaP3xiEf4glH40ENOhscT4Ir/sGnDuUPAAkObLArXXr1kkP9OXT/RF4chCNNRV2bAR+I5bUqtW5c2dhTDbQCnXr1jVyoogDlNevX1+wYEFhRjWiMggqck0AwatWyAoednfa+AL0DIlM+pVV0r/EgsDu37/foEEDYYBbIBY5T9JYq1iv7N692xS3nKJx1KS9yLtE72fPnm0hZD6sAQGRsDwFOfoCfxJLKqumCAb0emS0cOHCly5dkmZaBW6tWbPm6aefdjrjql9WTEzM/v37peJmQQvSn++99x6yAUQOqyCa+Ph44wfsTMHPxJKa7tm3bx8miZ6UCj6vWbMmQgyME4stF7kFOs9324/Dhg2TDmryXiHRulOnTsLShbzhC3QaCQgjLRCXgPifWFKr6NatW3HfhE6KYthBU1cyYQaHfB64NszIDI4vzEH0BAboZ599hlO4Eciq3Llzb9u2TQaGVTJAxJIaV1JSUpACyZNeQeIrONRuE8qrH6r/5Wvls2XLJow5RqjDG2+8ce/ePSq3ffv2ZBxFhM0B0QkkFGygBYhVMnDEkprsaRIH2bsGcNKgwTFRVwPvNQBBap1ialsa6xTPPvssyMRH7AMhQhsCIiBxkFCkSUNhFgEkllSuVUZwOnMLsnzmmWdWrFhx+vRpcnTu3LlDzpbroWdVk2EGRyAN9/DhwwcPHtAbU54WF81JKCMH6CISBIlDBphVMtDEkloDNm/e7PbKbmpt9uzZiWFFihSpUKFCnTp1GjVq1LRp05YOJCYmNlZA/yVnv1atWlWqVKlYsWLlypVpUmlhKhdRsz8A3Z4nTx7cfhVoVskgEEtqzdi5cyeOqKs+TaQJOCRAh1PnkwhkUFglg0MsqTXm66+/xn1uzC015drjGp5Q8Lguorz0CnQ1TYc5k29wJB4kYkklNWi9evWEcq9JFAECn/OuX78+sgMHbg7oiuARS2oNIz+9Xbt2IsJWj4IMnp20b98eIVbBZJUMMrGktlVHr8i5EIGzsyCAR+zo0aN5PTnIgg42saRyZG/t2rW4Ry5qFv0FNn/UsdS90tilG4FACIglFW6dOnUKLlf0vnjfgXkPvaEuRR45v999bByhIRYAbmVkZAwdOvRx7ea6qOqyAPYoSF2Rj+E1dUAQEEpiSeWQ++bNmznSJqq6TIEVVenSpfnSoZCYPxUhJpZUzOLNmze7desmtPEXVV1eoU59evTocePGDRlS86ci9MQCuDs2bNiARVQeiFG4Bd+HRZoeFxzrp6oLMuxCLKmorvT09H79+qHXopbRFTzkyKN67bXXEGFmE0XFsBGxAO6gvXv34qIl9XqSCAcoBSehSZMmX375pVQGpK1gO2JJRaVTl61cuRKZ+yKcXiqlEhISqFsytTv0bKWoGHYkFsD0unfv3vvvv4/TqhFIL5VS1AnUFdQh0mYelSvsSyyARyQ5XhMmTChQoICIGHqplCpYsCA1n+9IsqHtc4LdiSW1fgS9rly58u6775YoUUIoB6mz2MIENwrtojkyNfnq1atOXWFzhAGxABzLQZ/evXt33rx5NWrUYGFkDQXmtMJCDaRm8mlSOyx7GkfYEAsAvfCe3qSkpCQlJeGCQqFtwYYdw6jC6jZ8zpw5O3bsuHHjRrWlYUQpIMyIBThZhDNnzowfP75cuXIsKjDMzlYS9s4prIOaQA3hC+7CyPC5IiyJxVCH8o8//piWljZw4EDsOTJYfiEnGeoAxqufU4Vff/11qjxOeMtHFXOYIryJBTiJgcRDdmTw4MHkoyDPhypayDU48fJ8vYqrgaaKUfUGDRpEVWU+yfC0em6RFYgFuGZvJwkdP3581qxZnTt3Ll++PHKKOgmeBKxSzZpiU8+DgEZuMw/GxMRUqFChS5cuVCWqmEog1DxMrZ5bZB1iMdzeD/DTTz+dPHly5cqVw4cPf+GFF0qUKOGkzJyIwvzQgVc3jr5DBbVq1WrYsGErVqw4deqU0yGZrMcnRhYklgrSCiRLV+OSkZFx7Nix1atXjxs3rmfPns2bNy9btmyuXLks53GgH9LP6SH0KHogPXbVqlVUBGLujFQpiyGLE4uBSRZO6LvVEOToXLt27dChQ6mpqUS4uXPnTpo06W1dTJw4cc6cOfRl+gn98Pr166q3pBbN5WZJ5eQWkUIsFZwMAvL2+/yLc0y4zZ8TIYhEYrmCGYCdXTDjRwMAL/GrSKaRK6LEiiIgiBIrioAgSqwoAoL/A6GO23N48ta+AAAAAElFTkSuQmCC" alt="Ceará Finance" style="height: 60px; width: auto; object-fit: contain; margin: 0 auto;">
                </div>
                <div class="cover-subtitle" style="color: var(--accent-primary); font-size: 14px; font-weight: 800; letter-spacing: 0.25em; margin-bottom: 15px;">
                    CONSTELLATION CHALLENGE
                </div>
                <h1>${{d.nome}}</h1>
                <h2 style="font-weight: 800; font-size: 26px; border:none; margin-bottom: 40px; color: var(--text-secondary);">Ticker: ${{d.ticker}}.SA</h2>
                <p>Relatório Institucional de Valuation e Análise Multivariável de Risco</p>
            </div>
            `;
            
            // Slide 2: Resumo e Tese + Gráfico de Preços/SMA
            document.getElementById("slide-2").innerHTML = `
            <div class="grid-val-deck">
                <div class="val-card-left">
                    <h3 style="color: var(--accent-primary); font-size: 16px; margin-bottom: 10px;">📋 Tese de Investimento</h3>
                    <p style="font-size: 12px; line-height: 1.4; margin-bottom: 15px; text-align: justify; max-height: 160px; overflow-y: auto; padding-right: 5px;" class="custom-scrollbar">${{d.tese_ia}}</p>
                    <div class="card-stat" style="padding: 12px; margin-bottom: 10px;">
                        <h3 style="font-size: 10px;">Alvo Estimado</h3>
                        <div style="font-size: 20px; font-weight: 800; color: var(--success-color);">R$ ${{d.valor_justo.toFixed(2)}}</div>
                    </div>
                    <div class="val-badge-rec ${{d.recomendacao.includes('BUY') ? 'buy' : 'sell'}}" style="font-size: 13px; padding: 4px 10px; margin: 0;">
                        ${{d.recomendacao}}
                    </div>
                </div>
                <div style="text-align: center;">
                    <h3 style="margin-bottom: 10px; font-size: 14px;">Preço Histórico e Médias Móveis</h3>
                    <img src="/chart_historico.png?t=${{new Date().getTime()}}" style="max-width: 100%; border: 1.5px solid var(--border-color); border-radius: 8px; max-height: 200px;" alt="Média Móvel Chart">
                </div>
            </div>
            `;
            
            // Slide 3: Valuation e Ke
            document.getElementById("slide-3").innerHTML = `
            <div class="grid-val-deck">
                <div class="val-card-left">
                    <h3 style="margin-bottom: 15px; font-size: 16px;">🎯 Resumo de Preços e Alvos</h3>
                    <p>Cotação de Referência: <b>R$ ${{d.cotacao.toFixed(2)}}</b></p>
                    <p>Valor Justo (Cenário Base): <b>R$ ${{d.valor_justo.toFixed(2)}}</b></p>
                    <p>Upside do Modelo: <b style="color: var(--success-color);">${{d.upside.toFixed(1)}}%</b></p>
                    <div class="val-badge-rec ${{d.recomendacao.includes('BUY') ? 'buy' : 'sell'}}" style="margin-top: 15px;">
                        ${{d.recomendacao}}
                    </div>
                </div>
                <div>
                    <h3 style="margin-bottom: 15px; font-size: 16px;">📉 Estrutura do Custo de Capital</h3>
                    <p>Beta Alavancado: <b>${{d.beta_levered.toFixed(2)}}</b></p>
                    <p>Beta Desalavancado (Hamada): <b>${{d.beta_unlevered.toFixed(2)}}</b></p>
                    <p>Custo do Capital Próprio ($K_e$): <b>${{d.ke.toFixed(2)}}%</b></p>
                    <p>WACC Final: <b style="color: var(--accent-primary);">${{d.wacc.toFixed(2)}}%</b></p>
                </div>
            </div>
            `;
            
            // Slide 4: Fluxos DCF
            let rowsHtml = "";
            d.fluxos.forEach(f => {{
                rowsHtml += `
                <tr>
                    <td><b>Ano ${{f.ano}}</b></td>
                    <td>R$ ${{f.fluxo.toLocaleString('pt-BR', {{minimumFractionDigits: 2}})}}</td>
                    <td>R$ ${{f.vp.toLocaleString('pt-BR', {{minimumFractionDigits: 2}})}}</td>
                </tr>`;
            }});
            
            document.getElementById("slide-4").innerHTML = `
            <h3 style="margin-bottom: 15px; font-size: 16px;">📊 Demonstração dos Fluxos de Caixa da Firma (FCFF)</h3>
            <div class="table-responsive" style="margin-bottom: 20px;">
                <table>
                    <thead>
                        <tr>
                            <th>Período</th>
                            <th>Fluxo Proj. (FCFF)</th>
                            <th>Valor Presente (VP) no WACC de ${{d.wacc.toFixed(1)}}%</th>
                        </tr>
                    </thead>
                    <tbody>
                        ${{rowsHtml}}
                    </tbody>
                </table>
            </div>
            <p style="font-size: 12px; color: var(--text-secondary);">Enterprise Value: <b>R$ ${{d.ev.toLocaleString('pt-BR')}}</b> | Dívida Líquida: <b>R$ ${{d.divida.toLocaleString('pt-BR')}}</b> | Equity Value: <b>R$ ${{d.equity.toLocaleString('pt-BR')}}</b></p>
            `;
            
            // Slide 5: Monte Carlo & Retornos
            document.getElementById("slide-5").innerHTML = `
            <div class="grid-val-deck">
                <div style="text-align: center;" class="val-card-left">
                    <h3 style="margin-bottom: 10px; font-size: 14px;">Distribuição de Probabilidades</h3>
                    <img src="/monte_carlo_temp.png?t=${{new Date().getTime()}}" style="max-width: 100%; border: 1.5px solid var(--border-color); border-radius: 8px; max-height: 180px;" alt="Monte Carlo Histogram">
                    <p style="font-size: 11px; margin-top: 5px; color: var(--text-secondary);">P10 (Pessimista): <b>R$ ${{d.mc_p10.toFixed(2)}}</b> | P90 (Otimista): <b>R$ ${{d.mc_p90.toFixed(2)}}</b></p>
                </div>
                <div style="text-align: center;">
                    <h3 style="margin-bottom: 10px; font-size: 14px;">Perfil Histórico de Retornos Diários</h3>
                    <img src="/chart_retornos.png?t=${{new Date().getTime()}}" style="max-width: 100%; border: 1.5px solid var(--border-color); border-radius: 8px; max-height: 180px;" alt="Retornos Histogram">
                    <p style="font-size: 11px; margin-top: 5px; color: var(--text-secondary);">Probabilidade de ganho real (Upside): <b style="color: var(--success-color);">${{d.mc_upside_prob.toFixed(1)}}%</b></p>
                </div>
            </div>
            `;
            
            showSlide(0);
        }}

        function showSlide(idx) {{
            if (idx < 0) idx = 4;
            if (idx > 4) idx = 0;
            activeSlideIdx = idx;
            
            document.querySelectorAll('.slide-page').forEach((slide, sIdx) => {{
                if (sIdx === idx) {{
                    slide.classList.add('active');
                }} else {{
                    slide.classList.remove('active');
                }}
            }});
            
            document.getElementById("slide-counter").innerText = `Slide ${{idx + 1}} de 5`;
        }}

        function changeSlide(direction) {{
            showSlide(activeSlideIdx + direction);
        }}

        // Nova Aba de Demonstrações Financeiras
        async function fetchDemonstrativos() {{
            const ticker = document.getElementById("select-ticker-df").value;
            const statusArea = document.getElementById("df-status");
            
            statusArea.style.display = "block";
            statusArea.innerHTML = "⏳ <b>Robô consultando CVM e carregando demonstrativos históricos...</b>";
            
            const res = await fetch(`/api/demonstrativos?ticker=${{ticker}}`);
            const dfData = await res.json();
            
            if (dfData.sucesso) {{
                statusArea.innerHTML = `✅ <b>Demonstrativos carregados!</b> Link do PDF CVM: <a href="${{dfData.pdf_anual_url}}" target="_blank" style="color: var(--accent-primary); font-weight: 700;">Abrir Relatório Anual / DFP Oficial</a>`;
                renderTabelasDF(dfData);
                document.getElementById("df-tables-box").style.display = "block";
            }} else {{
                statusArea.innerHTML = "❌ Erro ao obter dados de demonstrativos.";
            }}
        }}

        function renderTabelasDF(data) {{
            const years = data.years;
            
            // Render Headers
            let headerHtml = "<th>Conta / Rubrica</th>";
            years.forEach(y => {{
                headerHtml += `<th>${{y}}</th>`;
            }});
            
            document.getElementById("df-thead-dre").innerHTML = `<tr>${{headerHtml}}</tr>`;
            document.getElementById("df-thead-balanco").innerHTML = `<tr>${{headerHtml}}</tr>`;
            document.getElementById("df-thead-fluxo").innerHTML = `<tr>${{headerHtml}}</tr>`;
            
            // Render DRE
            let dreHtml = "";
            for (const [key, yearVals] of Object.entries(data.dre)) {{
                dreHtml += `<tr><td><b>${{key}}</b></td>`;
                years.forEach(y => {{
                    dreHtml += `<td>${{yearVals[y] || "-"}}</td>`;
                }});
                dreHtml += "</tr>";
            }}
            document.getElementById("df-tbody-dre").innerHTML = dreHtml;
            
            // Render Balanço
            let balHtml = "";
            for (const [key, yearVals] of Object.entries(data.balanco)) {{
                balHtml += `<tr><td><b>${{key}}</b></td>`;
                years.forEach(y => {{
                    balHtml += `<td>${{yearVals[y] || "-"}}</td>`;
                }});
                balHtml += "</tr>";
            }}
            document.getElementById("df-tbody-balanco").innerHTML = balHtml;
            
            // Render Fluxo
            let fcHtml = "";
            for (const [key, yearVals] of Object.entries(data.fluxo)) {{
                fcHtml += `<tr><td><b>${{key}}</b></td>`;
                years.forEach(y => {{
                    fcHtml += `<td>${{yearVals[y] || "-"}}</td>`;
                }});
                fcHtml += "</tr>";
            }}
            document.getElementById("df-tbody-fluxo").innerHTML = fcHtml;
        }}

        function switchTab(tabId) {{
            document.querySelectorAll('.tab-btn').forEach(btn => btn.classList.remove('active'));
            document.querySelectorAll('.tab-content').forEach(cnt => cnt.classList.remove('active'));
            
            document.getElementById('btn-' + tabId).classList.add('active');
            document.getElementById('content-' + tabId).classList.add('active');
            
            if (tabId === 'val' && valuationStocks.length === 0) {{
                fetchValuation();
            }}
            if (tabId === 'df') {{
                fetchDemonstrativos();
            }}
            if (tabId === 'pipeline') {{
                loadCFConfig();
            }}
            if (tabId === 'work') {{
                const iframe = document.querySelector('#content-work iframe');
                if (!iframe.src || iframe.src === window.location.origin + '/' || iframe.src === '') {{
                    iframe.src = iframe.getAttribute('data-src');
                }}
            }}
        }}
        
        let cfConfig = {{}};
        
        async function loadCFConfig() {{
            try {{
                const res = await fetch("/api/cf_tech/config");
                const data = await res.json();
                if (data.sucesso) {{
                    cfConfig = data.config;
                    renderConfigForm();
                }} else {{
                    alert("Erro ao carregar configurações: " + data.erro);
                }}
            }} catch (err) {{
                console.error("Error loading config:", err);
            }}
        }}
        
        function renderConfigForm() {{
            if (!cfConfig.valuation) return;
            const w = cfConfig.valuation.wacc;
            const d = cfConfig.valuation.dcf;
            const mc = cfConfig.valuation.monte_carlo;
            
            document.getElementById("cfg-rf").value = w.taxa_livre_risco;
            document.getElementById("cfg-premium").value = w.premio_risco_mercado;
            document.getElementById("cfg-kd").value = w.custo_divida;
            document.getElementById("cfg-tax").value = w.aliquota_imposto;
            document.getElementById("cfg-de").value = w.de_ratio;
            
            document.getElementById("cfg-dcf-years").value = d.anos_projecao;
            document.getElementById("cfg-dcf-g").value = d.perpetuidade_g;
            
            document.getElementById("cfg-mc-sims").value = mc.num_simulacoes;
            document.getElementById("cfg-mc-vso").value = mc.vso_media;
            document.getElementById("cfg-mc-ebit").value = mc.margem_ebit_media;
        }}
        
        async function saveCFConfig() {{
            cfConfig.valuation.wacc.taxa_livre_risco = parseFloat(document.getElementById("cfg-rf").value);
            cfConfig.valuation.wacc.premio_risco_mercado = parseFloat(document.getElementById("cfg-premium").value);
            cfConfig.valuation.wacc.custo_divida = parseFloat(document.getElementById("cfg-kd").value);
            cfConfig.valuation.wacc.aliquota_imposto = parseFloat(document.getElementById("cfg-tax").value);
            cfConfig.valuation.wacc.de_ratio = parseFloat(document.getElementById("cfg-de").value);
            
            cfConfig.valuation.dcf.anos_projecao = parseInt(document.getElementById("cfg-dcf-years").value);
            cfConfig.valuation.dcf.perpetuidade_g = parseFloat(document.getElementById("cfg-dcf-g").value);
            
            cfConfig.valuation.monte_carlo.num_simulacoes = parseInt(document.getElementById("cfg-mc-sims").value);
            cfConfig.valuation.monte_carlo.vso_media = parseFloat(document.getElementById("cfg-mc-vso").value);
            cfConfig.valuation.monte_carlo.margem_ebit_media = parseFloat(document.getElementById("cfg-mc-ebit").value);
            
            try {{
                const res = await fetch("/api/cf_tech/config", {{
                    method: "POST",
                    headers: {{ "Content-Type": "application/json" }},
                    body: JSON.stringify(cfConfig)
                }});
                const data = await res.json();
                if (data.sucesso) {{
                    alert("Configurações salvas com sucesso!");
                }} else {{
                    alert("Erro ao salvar: " + data.erro);
                }}
            }} catch (err) {{
                alert("Erro de conexão ao salvar: " + err);
            }}
        }}
        
        async function runPipeline() {{
            const stage = document.getElementById("pipeline-stage").value;
            const ticker = document.getElementById("pipeline-ticker").value;
            const price = document.getElementById("pipeline-price").value;
            const consoleBox = document.getElementById("pipeline-console");
            const btn = document.getElementById("btn-run-pipeline");
            
            consoleBox.innerHTML = "⏳ [Pipeline] Iniciando execução do estágio " + stage.toUpperCase() + "...\n";
            btn.disabled = true;
            btn.innerHTML = "Rodando...";
            
            try {{
                const res = await fetch(`/api/cf_tech/run?stage=${{stage}}&ticker=${{ticker}}&preco_atual=${{price}}`);
                const data = await res.json();
                
                consoleBox.innerHTML += data.log;
                
                if (data.sucesso) {{
                    consoleBox.innerHTML += "\n✅ Pipeline executada com sucesso!\n";
                    if (data.kpis && data.kpis.length > 0) {{
                        renderPipelineKpis(data.kpis);
                    }}
                    if (data.report) {{
                        document.getElementById("pipeline-report-container").style.display = "block";
                        let rHtml = data.report.replace(/\n/g, "<br>").replace(/\*\*(.*?)\*\*/g, "<b>$1</b>");
                        document.getElementById("pipeline-ai-report").innerHTML = rHtml;
                    }}
                }} else {{
                    consoleBox.innerHTML += "\n❌ Falha na execução da pipeline.\n";
                }}
            }} catch (err) {{
                consoleBox.innerHTML += "\n❌ Erro de rede: " + err + "\n";
            }} finally {{
                btn.disabled = false;
                btn.innerHTML = "Executar Estágio";
            }}
        }}
        
        function renderPipelineKpis(kpis) {{
            document.getElementById("pipeline-kpis-container").style.display = "block";
            const tbody = document.getElementById("pipeline-kpis-tbody");
            let html = "";
            kpis.forEach(k => {{
                html += `
                <tr>
                    <td><b>${{k.denom_cia}}</b></td>
                    <td>R$ ${{(k.receita_ltm / 1e6).toFixed(1)}}M</td>
                    <td>R$ ${{(k.ebit_ltm / 1e6).toFixed(1)}}M</td>
                    <td>${{(k.margem_ebit * 100).toFixed(1)}}%</td>
                    <td>R$ ${{(k.divida_liquida / 1e6).toFixed(1)}}M</td>
                    <td>${{k.alavancagem_ebit.toFixed(2)}}x</td>
                    <td style="font-size: 10px; color: var(--text-secondary); max-width: 150px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;">${{k.source_file}}</td>
                </tr>
                `;
            }});
            tbody.innerHTML = html;
        }}

        window.addEventListener('DOMContentLoaded', () => {{
            fetchCVMDocs();
            
            document.getElementById('search-cvm').addEventListener('input', (e) => {{
                fetchCVMDocs(e.target.value);
            }});
        }});
    </script>
</head>
<body>
    <div class="container">
        <header>
            <div class="header-brand">
                <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAMgAAADICAIAAAAiOjnJAAAAtGVYSWZJSSoACAAAAAYAEgEDAAEAAAABAAAAGgEFAAEAAABWAAAAGwEFAAEAAABeAAAAKAEDAAEAAAACAAAAEwIDAAEAAAABAAAAaYcEAAEAAABmAAAAAAAAAGAAAAABAAAAYAAAAAEAAAAGAACQBwAEAAAAMDIxMAGRBwAEAAAAAQIDAACgBwAEAAAAMDEwMAGgAwABAAAA//8AAAKgBAABAAAAyAAAAAOgBAABAAAAyAAAAAAAAACKGshfAAAACXBIWXMAAA7EAAAOxAGVKw4bAAAFTmlUWHRYTUw6Y29tLmFkb2JlLnhtcAAAAAAAPD94cGFja2V0IGJlZ2luPSfvu78nIGlkPSdXNU0wTXBDZWhpSHpyZVN6TlRjemtjOWQnPz4KPHg6eG1wbWV0YSB4bWxuczp4PSdhZG9iZTpuczptZXRhLyc+CjxyZGY6UkRGIHhtbG5zOnJkZj0naHR0cDovL3d3dy53My5vcmcvMTk5OS8wMi8yMi1yZGYtc3ludGF4LW5zIyc+CgogPHJkZjpEZXNjcmlwdGlvbiByZGY6YWJvdXQ9JycKICB4bWxuczpBdHRyaWI9J2h0dHA6Ly9ucy5hdHRyaWJ1dGlvbi5jb20vYWRzLzEuMC8nPgogIDxBdHRyaWI6QWRzPgogICA8cmRmOlNlcT4KICAgIDxyZGY6bGkgcmRmOnBhcnNlVHlwZT0nUmVzb3VyY2UnPgogICAgIDxBdHRyaWI6Q3JlYXRlZD4yMDI2LTA3LTI3PC9BdHRyaWI6Q3JlYXRlZD4KICAgICA8QXR0cmliOkRhdGE+eyZxdW90O2RvYyZxdW90OzomcXVvdDtEQUd2MHZRTU85RSZxdW90OywmcXVvdDt1c2VyJnF1b3Q7OiZxdW90O1VBR2tiLUhraS13JnF1b3Q7LCZxdW90O2JyYW5kJnF1b3Q7OiZxdW90O0JBR2tiM2VKbU13JnF1b3Q7fTwvQXR0cmliOkRhdGE+CiAgICAgPEF0dHJpYjpFeHRJZD43YmYzODMxOS05MzMzLTQ5ZDMtYjg1MC03ZjNlYWE2ZmIxNDM8L0F0dHJpYjpFeHRJZD4KICAgICA8QXR0cmliOkZiSWQ+NTI1MjY1OTE0MTc5NTgwPC9BdHRyaWI6RmJJZD4KICAgICA8QXR0cmliOlRvdWNoVHlwZT4yPC9BdHRyaWI6VG91Y2hUeXBlPgogICAgPC9yZGY6bGk+CiAgIDwvcmRmOlNlcT4KICA8L0F0dHJpYjpBZHM+CiA8L3JkZjpEZXNjcmlwdGlvbj4KCiA8cmRmOkRlc2NyaXB0aW9uIHJkZjphYm91dD0nJwogIHhtbG5zOmRjPSdodHRwOi8vcHVybC5vcmcvZGMvZWxlbWVudHMvMS4xLyc+CiAgPGRjOnRpdGxlPgogICA8cmRmOkFsdD4KICAgIDxyZGY6bGkgeG1sOmxhbmc9J3gtZGVmYXVsdCc+RGVzaWduIHNlbSBub21lIC0gMTwvcmRmOmxpPgogICA8L3JkZjpBbHQ+CiAgPC9kYzp0aXRsZT4KIDwvcmRmOkRlc2NyaXB0aW9uPgoKIDxyZGY6RGVzY3JpcHRpb24gcmRmOmFib3V0PScnCiAgeG1sbnM6cGRmPSdodHRwOi8vbnMuYWRvYmUuY29tL3BkZi8xLjMvJz4KICA8cGRmOkF1dGhvcj5BbmFsaXN0YSBDZWFyw6EgRmluYW5jZTwvcGRmOkF1dGhvcj4KIDwvcmRmOkRlc2NyaXB0aW9uPgoKIDxyZGY6RGVzY3JpcHRpb24gcmRmOmFib3V0PScnCiAgeG1sbnM6eG1wPSdodHRwOi8vbnMuYWRvYmUuY29tL3hhcC8xLjAvJz4KICA8eG1wOkNyZWF0b3JUb29sPkNhbnZhIGRvYz1EQUd2MHZRTU85RSB1c2VyPVVBR2tiLUhraS13IGJyYW5kPUJBR2tiM2VKbU13PC94bXA6Q3JlYXRvclRvb2w+CiA8L3JkZjpEZXNjcmlwdGlvbj4KPC9yZGY6UkRGPgo8L3g6eG1wbWV0YT4KPD94cGFja2V0IGVuZD0ncic/PuXEA3cAACAASURBVHic7V13nFTV9b/WpQRpspTQl116pPeyAgGCgQSEDQYIPQooCChVglJESEAioUjoTYFQBHRd2m6oIhABEaR3lL40WSx7f+c33887n8uUN++9aW925vvHfGZnZ94t53vPOffec88VMoooAgAR6gpEkTURJVYUAUGUWP+PTA2/OPDzzz//9NNPPxoAfY2+jF/xQ0LdGlsgEonFHAKB6NW/z+fHMtv8+/ywQKQQi5nkScykfq5fv37o0KHU1NTVq1fPmTNn4sSJb+ti0qRJc+fOpS/TT+iH9HN6iKfSmWeBbKWNkMWJRbIk5UGvTh/evHlz3759H3300dixY3v27Nm8efOyZcvmypXrySefFJZAP6Sf00PoUfRAeiw9nIqgglxLd61S1kMWJBbUAwlP/TAjI+PAgQNTpkxJSkqqWrVqnjx5Hn/8cU8seeyxx5544gniylO6oC/Q1+jLnp5DRVBBVBwV+v7775NWe/jwoVorWMwsqcayDrHAJ9VhIq1w/PjxWbNm/fnPfy5durQrk4gTTzpA/KD/PqbBrLriH9JDwEiC63PoX3FxcZ07d6YqUcVUpYWaZyWGZQViwXmSDm7R6507d9avX//GG2/UqFGD9IqTaEnkzCGzBLJGOCoO3FX/RRWj6lElqap3797lysMVC21/+gVhTCySBJkSHuXkOKelpb3++uulSpWCUFUyBYdJ+kAdVJKhSlThgQMHUuXZ93dqWjgiLIkFq8d/fvPNNyNGjEhISFBF6NUBCi3YjVNrSL7/yJEjjxw5wk0LX/sYZsRSlwxu3bo1b968pk2bsr0Dn3S8cnsCtpIZRs1p1qwZNS09PV0qSxWh7ntzCBtiqZSiOTzN70qUKCE0awLvO5Ts8Afg+wutUdRAaiY1VoYhvcKAWCqlLl++PHr06MKFCwvNmtjZ3lmDU7uosWPHjr1y5YoMK3rZmljoR6bU8OHD8+XLx12fBVSUPqDAQK8CBQqMGjWKOsGpW2wL+xKLh+aFCxcGDx6cN29eETGUUqHSi8bVkCFDLl26JJVFFnvCjsRCsIB0rCDMmDGDBquISEqpUOkVGxs7c+ZMbC1wX9kN9iKWquR37txZp04dEfGUUqHSq27dutRF0mXxxSawEbFYt1+9evWvf/0r5kdRSrmCJ4/0Sh117do1ab8le1sQi8ccvVmwYEGRIkXU7ovCLXjIUXctXbpU2kx1hZ5Y3B0nT55s3ry50GxfFltECATUjmrZsiV1oLTNYn2IicWxbwsXLsyTJ4+I2j7zYNVOE+fFixdLLaoxtJINJbGgqG7evNmpUyehjb9QiylcwaqLOhOL9aE1i6EhFnbv6c0XX3wRFxen9ksUlsEjs0yZMtSx0hFIGCqzGAJisVNF5i9nzpzCEdcbaqFkHaAzc+TIsWTJEhk6lyvYxOKDK4MHDxYO/yDqUfkd3KvDhw/nI0lBFnRQiQVFdefOnXbt2omo+Qsk2CxSVyNCNcguV/CIhYZdunSpXr16wqGxo6wKKBCdRm+ow7/77jsZXG4FiVhw1Q8dOlS6dGkRdaqCCHR1fHz84cOHWRBBQDCIhcbs3LkzNjZWRFkVdKDDqfOxtxgcbgWcWGjG5s2bc+fOLRx+Vaj7ORKBbs+TJ09aWpoMCrcCSyw0YMOGDb/61a+EY7YS6h6OXKDzSRAkDhl4bgWQWKj6J598EhMTI6KssgEgAhIHCUUGmFuBIhYqnZKSgiXQKKtsAgiChELOiQwktwJCLExrt27dmitXLhFllc0AcZDLu23bNhmwNQj/EwsV3bdvHw4+RFllQ0AoJCASkwwMt/xMLFTx1KlTJUuWFNE5oI0B0ZCYSFgyANzyJ7GwIXX9+vVq1aqJKKtsDwiIhIUwG//uJ/qNWNhCf/jw4W9/+1sRXQUNE0BML7zwAgJs/BgH4TdiQZf2799fRFkVVoCwBg4cKP1qEP1DLFRo8eLFImoBwxAQ2aJFi6T/uOUHYqEq//vf/7BpE41ZCDtAZHny5CEhSj9xy1diwSqnp6dXqVJF2FhdIekZ53F0SuqHvHtIJ8TJIxEuZoeMbb7D63kCduTV9IKhJBbY3adPH2FL14ozt/j4HPukBbQAg9VGL7366qvSHzNEn4gFVq1Zs0bYbyHU6bxrtmzZypQpQzPW3r17jxkzZu7cuStXrvz000+THVi1atWSJUtmzpw5YcKEoUOHvvLKKx06dKhXr15cXBwZCCep4Ml2a68noBNiYmJGjx793HPPCV1J4V/YSfTRIFonFkh97ty5YsWK6Vc3yFAplT179saNG3/wwQcHDhy4deuWKQ1PPXv9+vUTJ05s3ryZntCpU6fSpUtjQ51h5wRdsOzCMaiWL19O8ipevLjQdYIhxBIlSpw/f176prd8JRaNbGEb10o9GUzKZsqUKYcPH1ZHHg4I8R04jJ80qNfjuDaZXMl9+/YtXbqUVBoNJ1VCaroOO4AdgAoVKmzfvp0q/+9//1sYGP/4FYlVhoRYkBZ1sZG6BgesqGhcTps2jTMsWr5uhC9dAvOcepn039atW//xj3+0bdsWubu4GpgZ+Ngca09QfUpSrgMHDrx69Sra0qxZM2FMBUCgy5Ytkz4YRCvEgoQuXLgA1Ro0Yuk4N+gv0vn9+vXDwQHX+wR8B85RkVZzeuzFixfJaevSpcuvf/1rtaqmyMGTVrWBPFFFw9VbDlSgODV/ad26dREYgxTfNAbwuZEqoQIlS5akdkmrM0QrxMLY7dy5swiKEfQqJNShbNmyn3/+uXS5rMa/xzX5aU6aDJ+TT0Y6rGrVqkwOr9rLXxlQeC2KNNOcOXPAJ7b4iYmJwowKQJd27dpVWjWIpomFwUoiFIFfC1X1E43a+Pj4IUOG0JxOXdfAF9q1a4c0USG5/8iJyhkZGTTfbNOmDcLRoFFcW6euLdEXihYt2qJFC5qTLlq0aNOmTTt27CA1Q87G3/72NxrDNJ+tXr069QApEtKLBQsWzJcvH9Eof/789CcNqkaNGlHn0FRDrRLoNXnyZGFeBUC4GKsWFL9FH6t+/foW6mq8Seogpr7r06dPamoq3KaXX35ZaHxCBbp3746D5JBuCJP4qBkJ6fXo0aNUN9VCMfiw8jPPPNOrV6+UlJTLly/r15xYcvv2bfKZyEKdPn2aOHTs2LGTJ0/Sn3fv3uXfIn8dTDb9uWvXLua3KSmg2iRoa11hjlhgrsH5hQU4Wb2GDRvS8MXUV10z49zu9NqkSZMHDx6wRO2QGgoU51G+c+dOMo5qj+ENeYSkn4giUrOk9IqLW8EMnq4ayb/APiU7gvQhMRsXwFgTFn41b948aT6I2QSx0DayOAji868dVCmVPXv2l156KTk5me+WoTdU+o0bNypWrCiU4V6mTJlz587JUKfs8QTYI3ozYsQIoY0E9ghxEkslhD57Mh+F033B/Fu+Um/Pnj3UP8IHFcCdTFNgadJbNUEsCO+dd94RfjWC6vycXIfevXvv379fKlMwFs+ECROElvEB5nLdunVSo53xhgQNqNXZs2exhszOVs2aNaGGoZz8VXnOoEyvixcvfvbZZ4XPhgWCHjdunDQ5eo0Si7P4FypUSPhJXampZooXL06+J1YKpDbW1bFI9s5JXQ0fPlxqKaDsSSxIgrxvoUxEqBVI1O6v5FWqzqM/yffCqrVfMvlA0CR0vrvAYK2MEgt9NHr0aOEPdaVO9+Li4v75z3/SRB0FuSYuR9GrVq0SykzqN7/5zZ07d6S/A2r9CB4VNF8TipixE+fp9mgjD4Qud3W/SPbDhg2j2YAwsMxhSlj0SqKXZpSWIWL5UV2xIaP3lSpVWrZsGeZ60vNKAUpv3bq1UBgJj9LpJlxbATIgR+rpp59m8ZARvHfvnjQ89DMNXJ5DX6CpX8+ePZFtGgbXjx6wNaVliFjoozFjxghL6opdItZSiYmJK1asgMrRn9ChT8ltx8VMeEJsbOyFCxekjdWV1DrtlVdeEY5FOKy90TRQGpthObnkeCD5AzR5orleamrq0qVLx44d27179+rVq4O7wpKiMmIuIXQigDSstLwTK1O7HNDaZFClFHVuixYtyKhB07AjpVO60yoDnjNq1Chpb1ZJrXrdunVDw1FzZDU2Ihv8nGZ2gwYN6tKlS8uWLevUqZOQkFCgQIGYmBgnKSAVlgUtBcbkyJFDP5YOTyYCGJ8eeicWhtfMmTOFSXWlhq9Q1X/3u9/RJA51clqq9lr6gAED6CEYlzTZwRKDPR12hrrxBbGR90PKRhobEiAfuQrCg1Lh/UHLhg/SqV279o4dO7Aw4XXfbMaMGdKYxvVCLAiPPE2vMWJOleAq5s2bNykp6csvv+QHmlrJxNeIlEIjVseOHaVdF65UoIassYQj0wuynxnUtfgaTfWFthutbkJbYBKDk/01aNCA6rl+/Xp9VglN9EQDsMqr+LwQC71D/pAwwCqnrZjSpUtTp+C6BGnpGjS2wlhowBW9H374obQ0qwoy0NgePXqAWBjx06dPl2ZWsfEQ+hVfgeYLn1hMYFWzZs0wGW/fvr0wYI5QOnky0sDA9kIsDBqvBTttxZBvTsXzdM/gyrIr8H0yfOhW4Rj0Bw8eNNKwkAM1pMmaUJz3/v37SzOV59FI47NChQrC5wVP9k9I8cNhOnHiRP78+YUB1uKHRAZpQOnqEQs/PnXqFHlIngpWHals2bKRS7Fp0yb45kamyl67FRXAwTLhyNNqnz1BfUAtYXWUNVaTJk3MLrVjA1E6Mo0J5ViHZUqR70/zu4yMDNRw2rRpwpj3jEJz5sx55swZ6Y1besRSN1KcZg1OywfFihUj/xqn0qS2C+v7rA1P+Pbbb5EQUDgOg0vlBh47A5U/e/askz7YsmWLNK9xIYuxY8dyn8OiqeuCntigOvjVqlVTAwAvX76s3tnuFeDf3//+d+nNoHskFiRH7UeGD3VnXo0iqlKlyvjx4xEL5bSf6rvsIZsjR46QykTLybJI21xwpQ+sj9Pr888/L5QgdFLq0nwTMrVLADZs2NCuXbuyZcs6DXV4909o4OOT/IX4+PhJkyaRopIOToBYEydOFGYm+3ggUeJn7RpATxX2SCw0Y9euXRz5pBZPKoSUx9q1a3/44QcZsNvVUe/z588XLFgQ5aakpMhwcLAA1yUDvE6ZMkVqe+em6MU9TO7Rtm3b3nrrrYYNGxYqVEiHGXny5CGXd/78+bdv35baQg8qRs5r0aJFhZk5Ab5JrcBdPToS90gsKDqqulDsIDWAJpyTJ08+evQoh+QGTn/wvhgia0lvkVmUtl/BUoGqJiUlCeXOBHJGkWHWQlyGGumFP2ngEckWLFgwevTo7t27d+jQgfzrLl26jBw5cvny5ceOHeMZNBsTPAGXrpndSgEZRowYIXWtoXtiZWo5icqVKweGli9f/p133tmxYwdUlDS2bu47UHVM2slZQfhDWPhYAJ++TEhIEI4VEwiS1Mx///tfqe2QGl/VUzehDfKSo4/wW/Dsgw8+EJbmmByjged4qoB7YqE70tLSaIo7dOhQMkAPHjzg/wbC6nkCaj9kyBDhCI5DNcKFVQDUw/Hjx2vUqAFrgpXevHnzYitdatF5FtrlNthB/USVFEeWJicnW4tX5p/QK64j8MQEPY1FUwbsE0ttohf8++8gld27d1Nj6tSpox6JCRdkapczkmP05ptv8iQREiKb9c0330iNIoGrBsdLEqtQB8tLYrCG2LH1ZA31ZoWQIp/2DMnqEUqk2hcvXpxmQzLcWCUVpYL+PH369LBhw0h7kcZihfHiiy+mpqYa3C2xUAH2g0lHYlXSl4VW/LZBgwb8fNdC9Yhlk1k96vCnP/3J71nngoxM5d5AsunkVpPymD17NrWrfv36lSpVotmifwewutl/8+bN3r17C8/H0YyDjyacPXtWWiCWTeJSUA2a+GCX0Ca1sgw2SU6gWdGVK1esZfTPVIBdIFUp0Cc0Z8S9a/6KAQQ158+fLz1YQ711LDuoKwZ5ncgJZjdYG4HMAGueK6++6v/w/PnzU6dOJd+UV9H8FVkKN6tXr17Sgw0JG2KFxMMzAn/NJ0w10Omb0IKk8y5evLhr1y7STyNHjmzVqhVnK/HLwQoVeFq5cuV4X9iphh6XG4wMiCDDVpUBoDnMZt7yvVDpODhP7nPz5s2bNm1Kb2rWrBkXF/fMM884EShA+ZXwwJiYGBzqd9XZesSStpSlTcDr1yNGjMCswhfnz/gYRinXrl3DNpdbxnAyo4Am1wCDFy5cKN25WXrrWFHoAKzasGEDde5//vMf6fN01fXcm6ev0WtqaiqnPXLNcBQ4MqmAmzVgwAAZJZa/gP65c+dO5cqVqXNx1ZHvkWfSQJwt/rt27VphOGttgIC9qcaNG6t9wnBDLPssNNgWbASFYzkHgWjWOg2/WrZsWd++fW/cuCFdTjY7AbqBdKQIdUp9lF64cOH79+9LI8QyqJMjFryAjuQI2bJlw56MtU4DR2fNmiUcWWWJrHv37pXKgoLT99Wwz6DlUtQh1lNPPXXkyBHX5rvXWFFT6AmZWvxjr1690L+ksYyE6noCfkU6j+O/iS7dunVDmkanZ7IxefHFF4UNcgqD2R9//LF0cTHda6wosTwBcj106BAfPq5QoYKPMRf4IQJN6bEcpclxuU7fJKVVq1YtEWqNJTT//e2335Yu/nuUWOaA7luyZAl3a/369X30HPBMBEhxSDFpI9hE9eGQS0ZGRtmyZYVtiMXx4mqjosQyB4i5bdu23K1FihSxkJfM9Znp6ekcHU+vw4YNky6m8Gftsm3jKZADClS1efPmro2K+lhW0KRJE+5Wct5xKNcXvQXG3L59m3ThhAkTkpOT3Trv+BoSDtjh5iKozPLly7tGkzoTK7rWoA/0nZqMPyYmxnhGBp3HOk3G3Ybp4hNE0IfccxeaysydO7erznYmFhboohrLE5yIhZ5dvXq19G3lXY144Sh4Jyngz++++w7XFITcDgrdFQc3xIpqLB04EQtqo2/fvjLwgWK/aImNYIDsQyx63bVrl9QnVtTB0gdmcMgAw2fBa9asaS1AzxQClFzYR4Bbn332mXxUZ7snVpRbnvCzdk+xei0A+dE0U5OBDJvO1PKK1a5dW9hgoYGBmuA+aXV0uTeFUWJ5ArT9gwcPXnrpJaHcCUjzxIC6pyj3xIkTOtEyIQEG2L/+9S8ZJZYvyNTOch06dEjdhKE3c+bMkQFL3IXHTp06VdhjoYGBykyYMEHqEys6JfQKnsEhzyAfnI+Njf3qq69kYC6KgpG1z0IDA9nwyPOTjw4qNxorSiyvAG+OHz+OGxvZiy9VqhQm3v69LIPtIHK428cOCo1Y2C7UI1Z0gdQgoPbJZ8+ZM6dQsjuVLFly9+7dUslj6PsKDtTV6tWrhc1YJUwRK6I0li+NBbcWLFiALuZbM/LmzTt9+nSnGaIvpgCPGjRokLCZHRRRYnmC5fbihxD5hx9+qCajg1ls3Lgx6Zjz589fvnx5586dli/RyNQSVyMM2j4LDUCUWG6Qnp7+/fffS9/0FnLkjRo1SmhTJPUMO85j9enTx3IpMKP79u2DCMPbFGZ5bmVq6XQGDhyoZmiy8Bz0ZseOHcWjdoqXuBITE/VTSekDBnf8+PHCfnZQGCdWMHNfhRa/aJnGeQ3GrOB5TQuXLrvqEnWL2to88RcNiACzmx0UxpcbQkUsT2cHAodftLtu4uLiwA9TDWddtXjxYrf5sXnn35czPKjYlClThC1ZJYwvkIZq5Z0PXgetaJRI3g/1y+zZs6XJa4/x5XXr1nlKN8XbiK4RxmYr2bJlS2FLOyi0Wk2bNk3qEyskPhbmVlu2bEHu2uAwWyVWjRo13B6O8wRUeNu2bfny5RMedAnfiL5nzx5piVj4yf79+3Fww25uO4C2L126VBqJbjDbBb4gUztV3KFDB+SuDU4FIDbcJ0jiR1p9r+EJ7FcRq9QrFF3Bmat5vdRU9bigvn37CruqK8bnn38u9YkVfFOI2sycOZPUBj4JPrGEdqmYV70CYS9atIgX3HW6G/+dOnWqNGlnpdYtly5d4ntTA08P0zAR6BcqH6t69epVqlTB+5AQq3Tp0l4P24Ac5K0jKNmrN40v0OQAZ+eNW0OeGQwdOlTYWF2BWGSpjx07JvWJFeS9QoxLciOofnXr1g1m7iSU9fLLLwttFxmJ1z01nxNZ42Ifg3M0fG3IkCHSwUv9RQ12cLHounHjRjWc0IZAxcjR9H6YQgbXzYIOGDlypHDc7OXLQqJZqMTCYoxrGKRTPQ8cOAC/yuxNs8Kx0oMS1Wzs/HwOxXn48CEG244dO2AE7bnKAKBuZGpcr9YJ8YFV9DW57cJxqNxT3sHAFY0UDJh2IZ2/K7HQa4cPH0Z+WAsJzSCANm3aEDXd8gkfZmoHnefPn68z37QPMMCoXa5Sc6+x/CtCHaCs1q1bC4dG9SW7hlmglK5duwpNY82dO1c+SiyemtHMDoeurLk77JDFxMTQLIHs6cWLF522pe/fv0+eysKFCzkvg20tIAOro/369ZO2OmLP5qBFixaoKN9lHegK8BpH+fLluYNciYX35OsgS6yPTjSrOiINmbmaNWu2atWqffv2pLBbtmxJBkXNRWt/Vgmt39577z1pJClI0HwsLggry4RBgwZJf4dfugX7MSxp4UIsvPn0009xv6tfpmZO12a7Qr0O0v5AQ9auXSsNaqygTQydNBbyhgdaY7GBQ148jqBCnlZ41phG0DwxQL4OjOMTCsJFSzF4rYH3S9RODnGqSCdiNWrUCKwKDrFwqRhHrC9ZskRqIcXSMQcsVqyYsPEyUmgBYpHraTRVpOuXAgRmMI67UEVz5sz59ddfywD77yAucYu8HKG4PsuWLZNa4N7+/fvBKludtbIVdHIYSR1iBYFbXArNV4U25588ebI0vwFiClBIR44cwVInq3RcRysd3jryi9p8th9aYMjBLTaRjjuYi0k0MxIascgsygCrTBDrL3/5i1DsYHx8fHp6unTsA4JwUQuoD50EpFKHWMFxs8D06dOnCyWPWVpamgxYHgRQ9saNGzgSyLMwLMaMGjUKTjRfsR5VWm7Bl1ufPn1aGr/yhIkVaL2FUi5cuICbZLFQmZSUJLW5oX8rkKmdq5kzZ45QlotIq5P5e/XVV/lD5lZok/TbFuicypUre7q8M/S3f6GU3//+90Lzo0nMrLT8TizpyF1WqlQpoWzuUokFChQQGqtYUdWvXx+fR7nlBOvXyqnaIqAuF6o1Y8YMoaQya9q0qQwAsbA65SnFlHqpH/137Nix9OX169fzoa4gC8/OsH4RplSuVg8osfDk69evc4ppiHDNmjXSr54WHrVnzx7OEqP2FEcS0ys57x999JHUugzJZKO+PMNvV/ci81PgjCPkRxpCKP5yuXLlbt686aneZoGHnD9/XidDOqhTuHBhZFFDfAu9uXjxYqFChURUaWmwftk4KPXFF18kJydjXZUfge72L8lQ3MmTJ7F6BDeL3owcOVL6wyCyz96zZ0/hTvdwiYmJibjZ0WnHEPcxRZUWgL6iGbT0vOLonlgQw6ZNm+j3ZcqUIXmQD7R//3410uMX7YZ03xnGe3MTJ04USsYpms26Jre0ADycrJvwoKvQUy1atMAtI2pnoeijR48i9CCqtLgHEHDrSTR6Wzokj0qVKglNGDRea9asSQZr8+bNV65cUb/vow5jjfL999/Hx8cLZWO4ePHip06dkj5s8jAz3B6q4RQxf/zjH+/cuSNdhiDXjYOYgypG+wEdWLFiRf1wX48+lho0HBMTo8qDOrdIkSKdOnUiHXDmzBn1YDWnKTfLM94YXrlyJdceiqR79+7SErGYEzQzqFq1qvDMqq5du5Iy5u87AUWnpaXZ5KKR0AJCId9A6u686a1j0euuXbvUdUJeDmDkz5+/YcOGNIE/ePAgtm/VJyBSwDjJUCgpD6HsDVMR1pQW+E0V4AeqNeeFUNLBmbrp0VD5S5cuQedFMrE4fA37qjoS0ZsVSocGqlatmnh0rEMeuKSKPyR/qFatWuSNTZs27cCBA64pXDK1i0NBNbeA8vvqq69wao81yoIFC6T5AECMJ7erVvgzNjaWFzW8spa+07x5c9dHRRQg8erVq7uennCC3joWBDNhwgThOXrErRp76qmnEhISkpKS6Ldr164l/+aHH34wQgW3vh0VvWXLFmnYheeoGHr/8ccfC3e6SjiO1SNEx0hcIZ42evRona6IBKAnaY4lvUWg6BELg5jMEOsPnSLVkEinf5GLRlPL1q1bv/nmmzS73LBhw86dO/fu3Uua6ZAGek+f7NixY/ny5R07dlQjhsmfM5UeTT0BQWbUqeZ4/4c//AETQHKtjFhqdaUtYomFriMyGDnzokcs+WhYi0ETwDu4IJlbOsKSPq2AlJyTbeUSsb1jkFJSU2wXL17EWqhabVSmUKFCfKmQwbRNINaYMWNEBBMLPUlk8Moq6ZVYEJI6U7MAlWdeI7t5oYH/TE5OlsbsIJ/RI8tLdBSeBwNxq02bNrwSE9VYRmDqqjMvxMrULiF+7rnnhP8iKh/zDP4Oe0Kmjkejwf369RO6KpajRpExzCu3osSCOIgGnuJkzBGL+5SjD4LWEshv+PDh0likMrtWM2fOFN7GgBobgyUZfd0eJRZET31rUBzeiQVu3rp1q2TJkiJYqzgopWjRoufPn/cqdQCKjaYFnlLsuQJml8pKTU3V7y/866233hIRSSyIgwhg/PZr78SSmn2B6xocpaXGkRlZDgDztm/fTtNPU5VEQe3atZO69EUP9O7dW0QksdCf48aNk4YXfQwRK1NLXh3k6BF22404QOvWrbOQCR1fLlCggE5oET6kCYHrWnEkgKfSpvItGiKW1HiKFcJAKy08v23bttJwir1vv/3WWtIOzvCks2uErrxw4UJsbKyIvC0ddCmJXpoJMzFKXNiMqAAACdRJREFUrOArLSN3luK/x44dM2sBGayxsOjndjiilFWrVonIYxXaW6RIEbPpYY0Si/v37bffFoFUWnhyYmKi9HagA//NyMho1KiRsOr6oLjq1at7yprMIQ89evSwXEr4Av3DizLG2WKCWBycHhcXJwLmamCIuGbhda2M77GdHDjqmqacAeN49epVZF2LKI0FEZM1sBAjboJYLGlcpBYIYoEfOAytn/TGqSbW5A1nv3bt2jqZBNXN7Ehz29HexYsXS/NBvOaIxXj++edFAAwiWoITMp5awrpq165davIFI1D3MVEWufwHDx70VBzvEcHaRhSxIFwStDWGmCYWBLB9+3a/t4TPfuhf7Qc1duXKFdc9Zv2HO4WUCUemFNc80q6NTU5OFhFmBAHqsR07dkjz6kpa01gQQ//+/YVflRYehexnnhbB2ZXGtdsWXOkcOXKQ7Rs5cuSWLVsQ8qrTa2gpUqRGVHwfGjtgwABp9bSBFWKxF5+QkCD8ZCA4RP/27dvSs58IEiD20FSm9caNG69atWrz5s2nTp1iJumHI+Nru3fvjrST0OgxEi6JWEcW+rDoY/klnEYFhsiYMWOk5z0ckGDdunWmLi1SbxHinXlO2+cWmUqKbDvfvBUgoMdIuNKSEQQsEktaigH0BFCEfKa7d+9KD7oXH545c8ZsWn11akO2D4zRH4V8znH+/Pm+ty68gMYavFlIB74S69y5c0ip6IveQmN69uyp0xidkxFGHm7qRkIwj7hVv359H5sWXkBLS5QogaASX44iWyeW1PTkmjVrhD96PyUlReqqq8OHDyMztil3B8QydQUX2oUgmchhldAa+8knn0gfjCDgE7G4eFwn6cumCs28pOchAmJhU8WsYUKtcBu2EWLBCG7fvt3O108GAmpmQ9/zOfpKLFAhPT0dR40tuCNqpiVPq5T0evr0aQvqSmjEGjp0qDRALD7pj1DsyPGu0NJq1arBzfU9H4evxJIauw8ePOh61sornGJ93A4UCPvdd98Vjk0Ys8RClyH9pP5A5FONfpmRhBEwtvPly4dNCL+kn/UDseSjgSWmnBII77XXXpPeVhk2btxYtGhR4bKG7hUgYq5cuXQW2eGtoxVUGRFJrOIzLAaP3xiEf4glH40ENOhscT4Ir/sGnDuUPAAkObLArXXr1kkP9OXT/RF4chCNNRV2bAR+I5bUqtW5c2dhTDbQCnXr1jVyoogDlNevX1+wYEFhRjWiMggqck0AwatWyAoednfa+AL0DIlM+pVV0r/EgsDu37/foEEDYYBbIBY5T9JYq1iv7N692xS3nKJx1KS9yLtE72fPnm0hZD6sAQGRsDwFOfoCfxJLKqumCAb0emS0cOHCly5dkmZaBW6tWbPm6aefdjrjql9WTEzM/v37peJmQQvSn++99x6yAUQOqyCa+Ph44wfsTMHPxJKa7tm3bx8miZ6UCj6vWbMmQgyME4stF7kFOs9324/Dhg2TDmryXiHRulOnTsLShbzhC3QaCQgjLRCXgPifWFKr6NatW3HfhE6KYthBU1cyYQaHfB64NszIDI4vzEH0BAboZ599hlO4Eciq3Llzb9u2TQaGVTJAxJIaV1JSUpACyZNeQeIrONRuE8qrH6r/5Wvls2XLJow5RqjDG2+8ce/ePSq3ffv2ZBxFhM0B0QkkFGygBYhVMnDEkprsaRIH2bsGcNKgwTFRVwPvNQBBap1ialsa6xTPPvssyMRH7AMhQhsCIiBxkFCkSUNhFgEkllSuVUZwOnMLsnzmmWdWrFhx+vRpcnTu3LlDzpbroWdVk2EGRyAN9/DhwwcPHtAbU54WF81JKCMH6CISBIlDBphVMtDEkloDNm/e7PbKbmpt9uzZiWFFihSpUKFCnTp1GjVq1LRp05YOJCYmNlZA/yVnv1atWlWqVKlYsWLlypVpUmlhKhdRsz8A3Z4nTx7cfhVoVskgEEtqzdi5cyeOqKs+TaQJOCRAh1PnkwhkUFglg0MsqTXm66+/xn1uzC015drjGp5Q8Lguorz0CnQ1TYc5k29wJB4kYkklNWi9evWEcq9JFAECn/OuX78+sgMHbg7oiuARS2oNIz+9Xbt2IsJWj4IMnp20b98eIVbBZJUMMrGktlVHr8i5EIGzsyCAR+zo0aN5PTnIgg42saRyZG/t2rW4Ry5qFv0FNn/UsdS90tilG4FACIglFW6dOnUKLlf0vnjfgXkPvaEuRR45v999bByhIRYAbmVkZAwdOvRx7ea6qOqyAPYoSF2Rj+E1dUAQEEpiSeWQ++bNmznSJqq6TIEVVenSpfnSoZCYPxUhJpZUzOLNmze7desmtPEXVV1eoU59evTocePGDRlS86ci9MQCuDs2bNiARVQeiFG4Bd+HRZoeFxzrp6oLMuxCLKmorvT09H79+qHXopbRFTzkyKN67bXXEGFmE0XFsBGxAO6gvXv34qIl9XqSCAcoBSehSZMmX375pVQGpK1gO2JJRaVTl61cuRKZ+yKcXiqlEhISqFsytTv0bKWoGHYkFsD0unfv3vvvv4/TqhFIL5VS1AnUFdQh0mYelSvsSyyARyQ5XhMmTChQoICIGHqplCpYsCA1n+9IsqHtc4LdiSW1fgS9rly58u6775YoUUIoB6mz2MIENwrtojkyNfnq1atOXWFzhAGxABzLQZ/evXt33rx5NWrUYGFkDQXmtMJCDaRm8mlSOyx7GkfYEAsAvfCe3qSkpCQlJeGCQqFtwYYdw6jC6jZ8zpw5O3bsuHHjRrWlYUQpIMyIBThZhDNnzowfP75cuXIsKjDMzlYS9s4prIOaQA3hC+7CyPC5IiyJxVCH8o8//piWljZw4EDsOTJYfiEnGeoAxqufU4Vff/11qjxOeMtHFXOYIryJBTiJgcRDdmTw4MHkoyDPhypayDU48fJ8vYqrgaaKUfUGDRpEVWU+yfC0em6RFYgFuGZvJwkdP3581qxZnTt3Ll++PHKKOgmeBKxSzZpiU8+DgEZuMw/GxMRUqFChS5cuVCWqmEog1DxMrZ5bZB1iMdzeD/DTTz+dPHly5cqVw4cPf+GFF0qUKOGkzJyIwvzQgVc3jr5DBbVq1WrYsGErVqw4deqU0yGZrMcnRhYklgrSCiRLV+OSkZFx7Nix1atXjxs3rmfPns2bNy9btmyuXLks53GgH9LP6SH0KHogPXbVqlVUBGLujFQpiyGLE4uBSRZO6LvVEOToXLt27dChQ6mpqUS4uXPnTpo06W1dTJw4cc6cOfRl+gn98Pr166q3pBbN5WZJ5eQWkUIsFZwMAvL2+/yLc0y4zZ8TIYhEYrmCGYCdXTDjRwMAL/GrSKaRK6LEiiIgiBIrioAgSqwoAoL/A6GO23N48ta+AAAAAElFTkSuQmCC" alt="Ceará Finance Logo" class="cf-logo">
                <div>
                    <h1>CEARÁ FINANCE</h1>
                    <p style="color: var(--text-secondary); font-size: 13px; font-weight: 500;">Liga de Mercado Financeiro da UFC | Hub Integrado de RI & Valuation</p>
                </div>
            </div>
            <div class="status-chip">
                <span class="pulse-dot"></span>
                Sistemas CVM & Valuation Ativos
            </div>
        </header>
        
        <main>
            <div class="tabs-header">
                <button id="btn-cvm" class="tab-btn active" onclick="switchTab('cvm')">🤖 Alertas e Consultas CVM</button>
                <button id="btn-val" class="tab-btn" onclick="switchTab('val')">📊 Monitor de Valuation (Mira)</button>
                <button id="btn-rel" class="tab-btn" onclick="switchTab('rel')">📋 Relatório Acadêmico (Poli USP)</button>
                <button id="btn-df" class="tab-btn" onclick="switchTab('df')">📈 Demonstrações Financeiras</button>
                <button id="btn-work" class="tab-btn" onclick="switchTab('work')">🏛️ Gregori Markets</button>
                <button id="btn-pipeline" class="tab-btn" onclick="switchTab('pipeline')">⚙️ CF Tech Pipelines</button>
            </div>
            
            <!-- TAB 1: ALERTAS E CONSULTAS CVM -->
            <div id="content-cvm" class="tab-content active">
                <section class="grid-stats">
                    <div class="card-stat">
                        <h3>Fatos Relevantes Processados</h3>
                        <div class="value">{total_docs}</div>
                        <div class="subtitle">Cadastrados no banco de dados SQLite local</div>
                    </div>
                    <div class="card-stat">
                        <h3>Empresas sob Monitoramento</h3>
                        <div class="value">{total_empresas}</div>
                        <div class="subtitle" style="font-family: monospace; font-size: 11px;">{tickers_list}</div>
                    </div>
                    <div class="card-stat">
                        <h3>Frequência do Vigia</h3>
                        <div class="value">A cada 15 min</div>
                        <div class="subtitle">Varredura automática e envio de alertas no Telegram</div>
                    </div>
                </section>
                
                <div class="search-container">
                    <input type="text" id="search-cvm" class="search-input" placeholder="Buscar documentos por ticker, nome da empresa, assunto ou palavra-chave do resumo..." autocomplete="off">
                </div>
                
                <section class="section-panel">
                    <div class="panel-header">
                        <h2 class="panel-title">Documentos de RI Monitorados na CVM</h2>
                        <span style="color: var(--text-secondary); font-size: 13px;">Pesquisa em tempo real na base de dados</span>
                    </div>
                    
                    <div class="table-responsive">
                        <table>
                            <thead>
                                <tr>
                                    <th>Ticker</th>
                                    <th>Empresa</th>
                                    <th>Categoria</th>
                                    <th>Tipo</th>
                                    <th>Descrição / Assunto</th>
                                    <th>Divulgação</th>
                                    <th style="text-align: center; width: 180px;">Ações</th>
                                </tr>
                            </thead>
                            <tbody id="cvm-tbody">
                                <tr>
                                    <td colspan="7" style="text-align: center; color: var(--text-secondary); padding: 40px;">
                                        Carregando documentos...
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                    </div>
                </section>
            </div>
            
            <!-- TAB 2: MONITOR DE VALUATION (MIRA) -->
            <div id="content-val" class="tab-content">


                <!-- EXCEL STYLE FILTER BAR -->
                <div class="excel-filters-bar">
                    <div>
                        <label class="excel-filter-label">Busca Ativo</label>
                        <input type="text" id="filt-ticker" class="excel-filter-input" placeholder="Ex: PETR4, Vale..." onkeyup="renderValuation()">
                    </div>
                    <div>
                        <label class="excel-filter-label">Setor B3</label>
                        <input type="text" id="filt-setor" class="excel-filter-input" placeholder="Ex: Consumo, Energia..." onkeyup="renderValuation()">
                    </div>
                    <div>
                        <label class="excel-filter-label">Tipo Carteira</label>
                        <select id="filt-tipo" class="excel-filter-input" onchange="renderValuation()">
                            <option value="TODOS">TODOS</option>
                            <option value="DIVIDENDOS">DIVIDENDOS</option>
                            <option value="CRESCIMENTO">CRESCIMENTO</option>
                            <option value="INDEFINIDO">INDEFINIDO</option>
                        </select>
                    </div>
                    <div>
                        <label class="excel-filter-label">Faixa de P/L</label>
                        <div style="display: flex; gap: 5px;">
                            <input type="number" id="filt-pl-min" class="excel-filter-input" style="padding: 6px; width: 50%;" placeholder="Min" onkeyup="renderValuation()" onchange="renderValuation()">
                            <input type="number" id="filt-pl-max" class="excel-filter-input" style="padding: 6px; width: 50%;" placeholder="Max" onkeyup="renderValuation()" onchange="renderValuation()">
                        </div>
                    </div>
                    <div>
                        <label class="excel-filter-label">Faixa de DY (%)</label>
                        <div style="display: flex; gap: 5px;">
                            <input type="number" id="filt-dy-min" class="excel-filter-input" style="padding: 6px; width: 50%;" placeholder="Min" onkeyup="renderValuation()" onchange="renderValuation()">
                            <input type="number" id="filt-dy-max" class="excel-filter-input" style="padding: 6px; width: 50%;" placeholder="Max" onkeyup="renderValuation()" onchange="renderValuation()">
                        </div>
                    </div>
                </div>
                
                <section class="section-panel">
                    <div class="panel-header">
                        <h2 class="panel-title">Monitor de Valuation do Mira — Planilha Integrada</h2>
                        <span style="color: var(--text-secondary); font-size: 13px;">Dica: Clique nos cabeçalhos das colunas para ordenar a tabela</span>
                    </div>
                    
                    <div class="table-responsive">
                        <table>
                            <thead>
                                <tr>
                                    <th onclick="sortValTable(0)">Ticker</th>
                                    <th onclick="sortValTable(1)">Empresa</th>
                                    <th onclick="sortValTable(2)">Preço Atual</th>
                                    <th onclick="sortValTable(3)">DY (%)</th>
                                    <th onclick="sortValTable(4)">P/L</th>
                                    <th onclick="sortValTable(5)">LPA</th>
                                    <th onclick="sortValTable(6)">VPA</th>
                                    <th onclick="sortValTable(7)">Tipo Ação</th>
                                    <th onclick="sortValTable(8)">Justo Bazin</th>
                                    <th onclick="sortValTable(9)">Teto Bazin</th>
                                    <th onclick="sortValTable(10)">Margem Bazin</th>
                                    <th onclick="sortValTable(11)">Justo Graham</th>
                                    <th onclick="sortValTable(12)">Margem Graham</th>
                                </tr>
                            </thead>
                            <tbody id="val-tbody">
                                <tr>
                                    <td colspan="13" style="text-align: center; color: var(--text-secondary); padding: 40px;">
                                        Carregando planilha de valuation e calculando fórmulas...
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                    </div>
                </section>
            </div>
            
            <!-- TAB 3: RELATÓRIO ACADÊMICO POLI USP -->
            <div id="content-rel" class="tab-content">
                <div class="rel-setup-box">
                    <label for="select-ticker" style="font-weight: 700; font-size: 15px;">Escolha a Empresa de Análise:</label>
                    <select id="select-ticker" class="select-ticker">
                        <option value="DIRR3">Direcional Engenharia (DIRR3)</option>
                        <option value="PETR4">Petrobras (PETR4)</option>
                        <option value="VALE3">Vale (VALE3)</option>
                        <option value="WEGE3">WEG (WEGE3)</option>
                    </select>
                    <button class="btn-gerar" onclick="dispararCalculoRelatorio()">📋 Gerar Relatório de Valuation Acadêmico</button>
                    
                    <div style="font-size: 13px; color: var(--text-secondary); max-width: 320px; line-height: 1.4;">
                        Roda a regressão do Beta histórico, projeta fluxos POC (incorporadoras) e simula 10.000 caminhos estatísticos.
                    </div>
                </div>
                
                <div id="rel-status" style="display: none; background: rgba(255,255,255,0.03); border: 1.5px solid var(--border-color); padding: 16px; border-radius: 12px; font-size: 14px;"></div>
                
                <!-- Deck de Slides formatado estilo Poli USP / Oss Capital -->
                <div id="deck-box" style="display: none;">
                    <div class="deck-container">
                        <div>
                            <div class="deck-header">
                                <h2>CONSTELLATION CHALLENGE</h2>
                                <div class="deck-subtitle">POLI USP & UFC CO-PROJECT</div>
                            </div>
                            
                            <!-- Os 5 Slides da Apresentação -->
                            <div id="slide-1" class="slide-page active"></div>
                            <div id="slide-2" class="slide-page"></div>
                            <div id="slide-3" class="slide-page"></div>
                            <div id="slide-4" class="slide-page"></div>
                            <div id="slide-5" class="slide-page"></div>
                        </div>
                        
                        <div class="deck-footer">
                            <div>Jack Gregori Rodriguez Cachi | Oss Capital</div>
                            <div id="slide-counter">Slide 1 de 5</div>
                            <div>Ceará Finance 2026</div>
                        </div>
                    </div>
                    
                    <div class="deck-nav">
                        <button class="deck-nav-btn" onclick="changeSlide(-1)">◀ Anterior</button>
                        <button class="deck-nav-btn" onclick="changeSlide(1)">Próximo ▶</button>
                        <button class="deck-nav-btn" onclick="window.print()" style="margin-left: auto; background: #fff; color: #000;">🖨️ Imprimir PDF</button>
                    </div>
                </div>
            </div>
            
            <!-- TAB 4: DEMONSTRAÇÕES FINANCEIRAS -->
            <div id="content-df" class="tab-content">
                <div class="rel-setup-box">
                    <label for="select-ticker-df" style="font-weight: 700; font-size: 15px;">Escolha o Ativo:</label>
                    <select id="select-ticker-df" class="select-ticker" onchange="fetchDemonstrativos()">
                        <option value="DIRR3">Direcional Engenharia (DIRR3)</option>
                        <option value="PETR4">Petrobras (PETR4)</option>
                        <option value="VALE3">Vale (VALE3)</option>
                        <option value="WEGE3">WEG (WEGE3)</option>
                    </select>
                    <button class="btn-gerar" onclick="fetchDemonstrativos()">🔄 Recarregar Demonstrativos</button>
                </div>
                
                <div id="df-status" style="margin-bottom: 20px; background: rgba(255,255,255,0.03); border: 1.5px solid var(--border-color); padding: 16px; border-radius: 12px; font-size: 14px;"></div>
                
                <div id="df-tables-box" style="display: none;">
                    <!-- 1. DRE -->
                    <section class="section-panel">
                        <h3 class="panel-title" style="margin-bottom: 15px; color: var(--accent-primary);">📊 Demonstração do Resultado do Exercício (DRE)</h3>
                        <div class="table-responsive">
                            <table>
                                <thead id="df-thead-dre"></thead>
                                <tbody id="df-tbody-dre"></tbody>
                            </table>
                        </div>
                    </section>
                    
                    <!-- 2. Balanço Patrimonial -->
                    <section class="section-panel">
                        <h3 class="panel-title" style="margin-bottom: 15px; color: var(--accent-glow);">🏦 Balanço Patrimonial (Ativo e Passivo)</h3>
                        <div class="table-responsive">
                            <table>
                                <thead id="df-thead-balanco"></thead>
                                <tbody id="df-tbody-balanco"></tbody>
                            </table>
                        </div>
                    </section>
                    
                    <!-- 3. DFC -->
                    <section class="section-panel">
                        <h3 class="panel-title" style="margin-bottom: 15px; color: var(--success-color);">💸 Demonstração dos Fluxos de Caixa (DFC)</h3>
                        <div class="table-responsive">
                            <table>
                                <thead id="df-thead-fluxo"></thead>
                                <tbody id="df-tbody-fluxo"></tbody>
                            </table>
                        </div>
                    </section>
                </div>
            </div>
            
            <!-- TAB 5: GREGORI MARKETS WORKSTATION -->
            <div id="content-work" class="tab-content">
                <iframe data-src="/workstation.html" src="" style="width: 100%; height: 85vh; border: none; border-radius: 12px; background: #030712;"></iframe>
            </div>
            
            <!-- TAB 6: CF TECH VALUATION & EQUITY RESEARCH PIPELINES -->
            <div id="content-pipeline" class="tab-content">
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 20px;">
                    <!-- LEFT COLUMN: SETTINGS EDITOR -->
                    <section class="section-panel" style="margin-bottom: 0;">
                        <h3 class="panel-title" style="color: var(--accent-primary);"><i class="fas fa-sliders-h"></i> Painel de Premissas (settings.yaml)</h3>
                        <p style="font-size: 11px; color: var(--text-secondary); margin-bottom: 15px;">Atualize as taxas, impostos e restrições globais do sistema sem alterar código.</p>
                        
                        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px; margin-bottom: 15px;">
                            <div>
                                <label style="display:block; font-size:11px; margin-bottom:5px; color:var(--text-secondary);">Taxa Livre de Risco (Rf)</label>
                                <input type="number" id="cfg-rf" step="0.001" style="width:100%; padding:8px; background:#1b1c24; border:1px solid var(--border-color); border-radius:6px; color:#fff;">
                            </div>
                            <div>
                                <label style="display:block; font-size:11px; margin-bottom:5px; color:var(--text-secondary);">Prêmio de Risco (CAPM)</label>
                                <input type="number" id="cfg-premium" step="0.001" style="width:100%; padding:8px; background:#1b1c24; border:1px solid var(--border-color); border-radius:6px; color:#fff;">
                            </div>
                            <div>
                                <label style="display:block; font-size:11px; margin-bottom:5px; color:var(--text-secondary);">Custo da Dívida (Kd)</label>
                                <input type="number" id="cfg-kd" step="0.001" style="width:100%; padding:8px; background:#1b1c24; border:1px solid var(--border-color); border-radius:6px; color:#fff;">
                            </div>
                            <div>
                                <label style="display:block; font-size:11px; margin-bottom:5px; color:var(--text-secondary);">Alíquota IR (T)</label>
                                <input type="number" id="cfg-tax" step="0.01" style="width:100%; padding:8px; background:#1b1c24; border:1px solid var(--border-color); border-radius:6px; color:#fff;">
                            </div>
                            <div>
                                <label style="display:block; font-size:11px; margin-bottom:5px; color:var(--text-secondary);">Estrutura D/E</label>
                                <input type="number" id="cfg-de" step="0.01" style="width:100%; padding:8px; background:#1b1c24; border:1px solid var(--border-color); border-radius:6px; color:#fff;">
                            </div>
                            <div>
                                <label style="display:block; font-size:11px; margin-bottom:5px; color:var(--text-secondary);">Anos Projeção DCF</label>
                                <input type="number" id="cfg-dcf-years" step="1" style="width:100%; padding:8px; background:#1b1c24; border:1px solid var(--border-color); border-radius:6px; color:#fff;">
                            </div>
                            <div>
                                <label style="display:block; font-size:11px; margin-bottom:5px; color:var(--text-secondary);">Perpetuidade dy g</label>
                                <input type="number" id="cfg-dcf-g" step="0.001" style="width:100%; padding:8px; background:#1b1c24; border:1px solid var(--border-color); border-radius:6px; color:#fff;">
                            </div>
                            <div>
                                <label style="display:block; font-size:11px; margin-bottom:5px; color:var(--text-secondary);">Simulações Monte Carlo</label>
                                <input type="number" id="cfg-mc-sims" step="1000" style="width:100%; padding:8px; background:#1b1c24; border:1px solid var(--border-color); border-radius:6px; color:#fff;">
                            </div>
                        </div>
                        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px; margin-bottom: 15px;">
                            <div>
                                <label style="display:block; font-size:11px; margin-bottom:5px; color:var(--text-secondary);">VSO Média (MC)</label>
                                <input type="number" id="cfg-mc-vso" step="0.01" style="width:100%; padding:8px; background:#1b1c24; border:1px solid var(--border-color); border-radius:6px; color:#fff;">
                            </div>
                            <div>
                                <label style="display:block; font-size:11px; margin-bottom:5px; color:var(--text-secondary);">Margem EBIT Média (MC)</label>
                                <input type="number" id="cfg-mc-ebit" step="0.01" style="width:100%; padding:8px; background:#1b1c24; border:1px solid var(--border-color); border-radius:6px; color:#fff;">
                            </div>
                        </div>
                        
                        <button onclick="saveCFConfig()" class="action-btn-main" style="width:100%; padding:10px; background:var(--success-color); font-weight:bold;"><i class="fas fa-save"></i> Salvar Alterações</button>
                    </section>
                    
                    <!-- RIGHT COLUMN: PIPELINE CONTROLLER -->
                    <section class="section-panel" style="margin-bottom: 0; display: flex; flex-direction: column; justify-content: space-between;">
                        <div>
                            <h3 class="panel-title" style="color: var(--accent-primary);"><i class="fas fa-cogs"></i> Executar Pipelines de Dados</h3>
                            <p style="font-size: 11px; color: var(--text-secondary); margin-bottom: 20px;">Selecione o estágio e clique para processar a fila.</p>
                            
                            <div style="margin-bottom: 15px;">
                                <label style="display:block; font-size:11px; margin-bottom:5px; color:var(--text-secondary);">Estágio da Pipeline</label>
                                <select id="pipeline-stage" style="width:100%; padding:10px; background:#1b1c24; border:1px solid var(--border-color); border-radius:6px; color:#fff;">
                                    <option value="all">Pipeline Completa (End-to-End)</option>
                                    <option value="ingest">Ingestão CVM (Coletor DFP/ITR)</option>
                                    <option value="clean">Isolamento e Limpeza LTM</option>
                                    <option value="kpi">Cálculo de KPIs de Endividamento/Margem</option>
                                    <option value="wacc">Custo de Capital (Regressão Beta OLS)</option>
                                    <option value="dcf">Fluxo de Caixa Descontado (FCFF)</option>
                                    <option value="mc">Simulação de Monte Carlo</option>
                                    <option value="risk">Auditoria de Riscos & Anomalias</option>
                                </select>
                            </div>
                            
                            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px; margin-bottom: 20px;">
                                <div>
                                    <label style="display:block; font-size:11px; margin-bottom:5px; color:var(--text-secondary);">Ticker do Ativo</label>
                                    <select id="pipeline-ticker" style="width:100%; padding:10px; background:#1b1c24; border:1px solid var(--border-color); border-radius:6px; color:#fff;">
                                        <option value="DIRR3">DIRR3 (Direcional)</option>
                                        <option value="PETR4">PETR4 (Petrobras)</option>
                                        <option value="VALE3">VALE3 (Vale)</option>
                                        <option value="WEGE3">WEGE3 (Weg)</option>
                                    </select>
                                </div>
                                <div>
                                    <label style="display:block; font-size:11px; margin-bottom:5px; color:var(--text-secondary);">Preço Atual B3 (R$)</label>
                                    <input type="number" id="pipeline-price" step="0.01" value="11.50" style="width:100%; padding:10px; background:#1b1c24; border:1px solid var(--border-color); border-radius:6px; color:#fff;">
                                </div>
                            </div>
                        </div>
                        
                        <button id="btn-run-pipeline" onclick="runPipeline()" class="action-btn-main" style="width:100%; padding:12px; background:var(--accent-glow); font-weight:bold; font-size:14px;"><i class="fas fa-play"></i> Executar Estágio da Pipeline</button>
                    </section>
                </div>
                
                <!-- TERMINAL CONSOLE LOGGER -->
                <section class="section-panel" style="margin-bottom: 20px;">
                    <h3 class="panel-title" style="color: #46e0a0; font-family: monospace;"><i class="fas fa-terminal"></i> Console de Execução (Linhagem & Logs)</h3>
                    <pre id="pipeline-console" style="background:#050508; border: 1.5px solid #1b1c24; border-radius: 8px; padding:15px; height: 180px; overflow-y: auto; color:#46e0a0; font-family: 'Courier New', Courier, monospace; font-size:11px; white-space: pre-wrap; word-break: break-all; margin:0;" class="custom-scrollbar">Console ocioso. Selecione um estágio e execute a pipeline...</pre>
                </section>
                
                <!-- PARQUET KPIs RESULT TABLE -->
                <section id="pipeline-kpis-container" class="section-panel" style="margin-bottom: 20px; display: none;">
                    <h3 class="panel-title" style="color: var(--success-color);"><i class="fas fa-database"></i> Tabela Processada (Parquet Lineage: kpis_calculados.parquet)</h3>
                    <div class="table-responsive">
                        <table>
                            <thead>
                                <tr>
                                    <th>Empresa / Denominação</th>
                                    <th>Receita LTM</th>
                                    <th>EBIT LTM</th>
                                    <th>Margem EBIT</th>
                                    <th>Dívida Líquida</th>
                                    <th>Alavancagem</th>
                                    <th>Arquivo CVM de Origem</th>
                                </tr>
                            </thead>
                            <tbody id="pipeline-kpis-tbody"></tbody>
                        </table>
                    </div>
                </section>
                
                <!-- RESEARCH ASSISTANT REPORT -->
                <section id="pipeline-report-container" class="section-panel" style="margin-bottom: 0; display: none;">
                    <h3 class="panel-title" style="color: var(--warning-color);"><i class="fas fa-file-invoice-dollar"></i> Relatório de Auditoria e Riscos (IA Agent)</h3>
                    <div id="pipeline-ai-report" style="background:#0c0d12; border:1px solid var(--border-color); border-radius:8px; padding:20px; font-size:13px; line-height:1.6; color:var(--text-primary); text-align:justify;"></div>
                </section>
            </div>
        </main>
    </div>
</body>
</html>
"""
        return html

def iniciar_servidor_dashboard(port=8000):
    server_address = ("", port)
    httpd = ThreadingHTTPServer(server_address, DashboardHandler)
    print(f"[Dashboard] Servidor HTTP ativo em http://localhost:{port}")
    httpd.serve_forever()

def rodar_dashboard_async(port=8000):
    t = threading.Thread(target=iniciar_servidor_dashboard, args=(port,), daemon=True)
    t.start()
    return t

if __name__ == "__main__":
    iniciar_servidor_dashboard()
