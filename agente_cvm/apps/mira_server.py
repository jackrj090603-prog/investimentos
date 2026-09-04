import os
import sys
import json
import openpyxl
import urllib.parse
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
AGENTE_DIR = os.path.abspath(os.path.join(BASE_DIR, ".."))
PROJECT_DIR = os.path.abspath(os.path.join(BASE_DIR, "..", ".."))

PORT = 8002

_VALUATION_CACHE = None
_CACHE_MTIME = 0

def obter_caminho_excel():
    paths = [
        os.path.join(PROJECT_DIR, "Mira", "MONITOR DE VALUATION DO MIRA (3).xlsx"),
        os.path.join(PROJECT_DIR, "MONITOR DE VALUATION DO MIRA (3).xlsx"),
        os.path.join(AGENTE_DIR, "Mira", "MONITOR DE VALUATION DO MIRA (3).xlsx")
    ]
    for p in paths:
        if os.path.exists(p):
            return p
    return paths[0]

def carregar_dados_mira_cached():
    global _VALUATION_CACHE, _CACHE_MTIME
    
    excel_path = obter_caminho_excel()
    if not os.path.exists(excel_path):
        return []
        
    mtime_atual = os.path.getmtime(excel_path)
    if _VALUATION_CACHE is not None and mtime_atual == _CACHE_MTIME:
        return _VALUATION_CACHE
        
    print(f"[Mira Server] Carregando e parseando Excel: {excel_path}...")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        b3_map = {}
        sheet_b3_name = 'Classificação B3' if 'Classificação B3' in wb.sheetnames else 'Classificacao B3'
        if sheet_b3_name in wb.sheetnames:
            sheet_b3 = wb[sheet_b3_name]
            for r in range(2, sheet_b3.max_row + 1):
                ticker = sheet_b3.cell(r, 1).value
                if ticker:
                    b3_map[str(ticker).strip().upper()] = {
                        "nome": sheet_b3.cell(r, 2).value or "",
                        "setor": sheet_b3.cell(r, 4).value or "Outros",
                        "segmento": sheet_b3.cell(r, 6).value or "Outros",
                        "ciclico": sheet_b3.cell(r, 7).value or ""
                    }
        
        sheet_si = wb['Dados Status Invest']
        stocks = []
        
        for r in range(2, sheet_si.max_row + 1):
            ticker = sheet_si.cell(r, 1).value
            if not ticker:
                continue
            ticker = str(ticker).strip().upper()
            
            def get_val(col_idx):
                val = sheet_si.cell(r, col_idx).value
                if val is None or val == "-":
                    return 0.0
                try:
                    return float(val)
                except:
                    return 0.0
            
            preco = get_val(2)
            dy_raw = get_val(3)
            pl = get_val(4)
            roe_raw = get_val(18)
            cagr_lucro_raw = get_val(25)
            vpa = get_val(27)
            lpa = get_val(28)
            
            # Fórmulas de Valuation
            preco_teto_graham = 0.0
            if vpa > 0 and lpa > 0:
                try:
                    preco_teto_graham = round((22.5 * vpa * lpa) ** 0.5, 2)
                except:
                    preco_teto_graham = 0.0
                    
            dpa_estimado = round(preco * dy_raw, 2) if dy_raw > 0 else 0.0
            preco_teto_bazin = round(dpa_estimado / 0.06, 2) if dpa_estimado > 0 else 0.0
            
            b3_info = b3_map.get(ticker, {
                "nome": ticker,
                "setor": "Outros",
                "segmento": "Outros",
                "ciclico": ""
            })
            
            stocks.append({
                "ticker": ticker,
                "nome": b3_info["nome"],
                "setor": b3_info["setor"],
                "segmento": b3_info["segmento"],
                "preco": preco,
                "dy": round(dy_raw * 100, 2),
                "pl": pl,
                "roe": round(roe_raw * 100, 2),
                "cagr_lucro": round(cagr_lucro_raw * 100, 2),
                "vpa": vpa,
                "lpa": lpa,
                "graham": preco_teto_graham,
                "bazin": preco_teto_bazin,
                "margem_graham": round(((preco_teto_graham / preco) - 1) * 100, 1) if preco > 0 and preco_teto_graham > 0 else 0.0,
                "margem_bazin": round(((preco_teto_bazin / preco) - 1) * 100, 1) if preco > 0 and preco_teto_bazin > 0 else 0.0
            })
            
        wb.close()
        _VALUATION_CACHE = stocks
        _CACHE_MTIME = mtime_atual
        print(f"[Mira Server] Cache atualizado com sucesso! {len(stocks)} ações carregadas.")
        return stocks
    except Exception as e:
        print(f"[Mira Server] Erro ao carregar planilha: {e}")
        return []

class MiraHandler(SimpleHTTPRequestHandler):
    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        
        if parsed.path in ["/", "/index.html"]:
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(self.render_mira_dashboard().encode("utf-8"))
            
        elif parsed.path == "/LOGO_CF.png":
            self.serve_logo()
            
        elif parsed.path == "/health":
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(b'{"status": "ok", "app": "mira"}')
            
        elif parsed.path == "/api/valuation":
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            
            data = carregar_dados_mira_cached()
            self.wfile.write(json.dumps(data).encode("utf-8"))
            
        elif parsed.path == "/api/reload":
            global _VALUATION_CACHE
            _VALUATION_CACHE = None
            data = carregar_dados_mira_cached()
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            self.wfile.write(json.dumps({"status": "reloaded", "count": len(data)}).encode("utf-8"))
            
        else:
            self.send_response(404)
            self.end_headers()

    def serve_logo(self):
        paths_to_try = [
            os.path.join(PROJECT_DIR, "Mira", "LOGO_CF.png"),
            os.path.join(PROJECT_DIR, "LOGO_CF.png"),
            os.path.join(AGENTE_DIR, "LOGO_CF.png")
        ]
        for p in paths_to_try:
            if os.path.exists(p):
                self.send_response(200)
                self.send_header("Content-Type", "image/png")
                self.end_headers()
                with open(p, "rb") as f:
                    self.wfile.write(f.read())
                return
        self.send_response(404)
        self.end_headers()

    def render_mira_dashboard(self):
        html = """<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Monitor de Valuation (Mira) — Ceará Finance</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;600;700&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <style>
        :root {
            --bg-dark: #07080c;
            --bg-card: #0d0f17;
            --border-color: rgba(255, 255, 255, 0.08);
            --accent: #5b9bff;
            --text-main: #ffffff;
            --text-muted: #8e9bb0;
            --font-mono: 'JetBrains Mono', monospace;
        }
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            font-family: 'Outfit', sans-serif;
            background: var(--bg-dark);
            color: var(--text-main);
            padding: 20px 25px;
            min-height: 100vh;
        }
        .container {
            max-width: 1500px;
            margin: 0 auto;
        }
        header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding-bottom: 18px;
            border-bottom: 1px solid var(--border-color);
            margin-bottom: 20px;
        }
        .brand {
            display: flex;
            align-items: center;
            gap: 16px;
        }
        .brand img {
            height: 46px;
            object-fit: contain;
        }
        .brand-title h1 {
            font-size: 20px;
            font-weight: 800;
        }
        .brand-title p {
            font-size: 12px;
            color: var(--accent);
            text-transform: uppercase;
            letter-spacing: 0.1em;
            font-weight: 700;
        }
        .header-actions {
            display: flex;
            gap: 10px;
            align-items: center;
        }
        .btn-action {
            background: rgba(255, 255, 255, 0.06);
            border: 1px solid var(--border-color);
            color: #ffffff;
            padding: 8px 14px;
            border-radius: 8px;
            text-decoration: none;
            font-size: 13px;
            font-weight: 600;
            display: inline-flex;
            align-items: center;
            gap: 6px;
            cursor: pointer;
            transition: all 0.2s;
        }
        .btn-action:hover {
            background: rgba(255, 255, 255, 0.12);
            border-color: rgba(255, 255, 255, 0.2);
        }
        .btn-accent {
            background: rgba(91, 155, 255, 0.15);
            border-color: var(--accent);
            color: #93c5fd;
        }
        .filters-panel {
            background: var(--bg-card);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            padding: 18px;
            margin-bottom: 20px;
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 14px;
            align-items: end;
        }
        .filter-group label {
            display: block;
            font-size: 11px;
            font-weight: 600;
            color: var(--text-muted);
            margin-bottom: 6px;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }
        .filter-group input, .filter-group select {
            width: 100%;
            background: #141724;
            border: 1px solid var(--border-color);
            padding: 9px 12px;
            border-radius: 8px;
            color: #ffffff;
            font-size: 13px;
            outline: none;
            font-family: inherit;
        }
        .filter-group input:focus, .filter-group select:focus {
            border-color: var(--accent);
        }
        .summary-ribbon {
            display: flex;
            gap: 20px;
            margin-bottom: 16px;
            font-size: 13px;
            color: var(--text-muted);
            align-items: center;
            background: rgba(255, 255, 255, 0.02);
            padding: 10px 16px;
            border-radius: 8px;
            border: 1px solid var(--border-color);
        }
        .summary-ribbon span b {
            color: #ffffff;
            font-family: var(--font-mono);
        }
        .table-wrap {
            background: var(--bg-card);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            overflow-x: auto;
            max-height: 70vh;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
            text-align: left;
        }
        th {
            background: #121520;
            padding: 12px 14px;
            font-weight: 700;
            font-size: 11px;
            text-transform: uppercase;
            letter-spacing: 0.06em;
            color: var(--text-muted);
            border-bottom: 1px solid var(--border-color);
            position: sticky;
            top: 0;
            z-index: 10;
            cursor: pointer;
            user-select: none;
            white-space: nowrap;
        }
        th:hover {
            color: #ffffff;
            background: #181d2c;
        }
        td {
            padding: 10px 14px;
            border-bottom: 1px solid rgba(255, 255, 255, 0.04);
            white-space: nowrap;
        }
        tr:hover td {
            background: rgba(255, 255, 255, 0.03);
        }
        .ticker-badge {
            background: rgba(91, 155, 255, 0.12);
            color: #60a5fa;
            border: 1px solid rgba(91, 155, 255, 0.25);
            padding: 3px 8px;
            border-radius: 6px;
            font-family: var(--font-mono);
            font-weight: 700;
            font-size: 12px;
        }
        .num-cell {
            font-family: var(--font-mono);
            text-align: right;
        }
        .tag-positive {
            color: #46e0a0;
            font-weight: 600;
        }
        .tag-negative {
            color: #ff6b6b;
            font-weight: 600;
        }
    </style>
</head>
<body>
    <div class="container">
        <header>
            <div class="brand">
                <img src="/LOGO_CF.png" alt="Ceará Finance" onerror="this.style.display='none'">
                <div class="brand-title">
                    <h1>Monitor de Valuation (Mira)</h1>
                    <p>Filtros Dinâmicos Avançados & Múltiplos B3</p>
                </div>
            </div>
            <div class="header-actions">
                <a href="http://localhost:8000" class="btn-action"><i class="fas fa-arrow-left"></i> Hub Central</a>
                <button onclick="recarregarCache()" class="btn-action btn-accent"><i class="fas fa-bolt"></i> Recarregar Excel</button>
                <button onclick="exportarCSV()" class="btn-action"><i class="fas fa-file-csv"></i> Exportar CSV</button>
            </div>
        </header>

        <div class="filters-panel">
            <div class="filter-group">
                <label><i class="fas fa-search"></i> Busca Rápida</label>
                <input type="text" id="filtro-busca" placeholder="Ticker ou nome da ação..." oninput="aplicarFiltros()">
            </div>
            <div class="filter-group">
                <label><i class="fas fa-layer-group"></i> Setor de Atuação</label>
                <select id="filtro-setor" onchange="aplicarFiltros()">
                    <option value="">Todos os Setores</option>
                </select>
            </div>
            <div class="filter-group">
                <label><i class="fas fa-percent"></i> DY Mínimo (%)</label>
                <input type="number" id="filtro-dy-min" placeholder="Ex: 6.0" step="0.5" oninput="aplicarFiltros()">
            </div>
            <div class="filter-group">
                <label><i class="fas fa-arrow-trend-up"></i> ROE Mínimo (%)</label>
                <input type="number" id="filtro-roe-min" placeholder="Ex: 15.0" step="1" oninput="aplicarFiltros()">
            </div>
            <div class="filter-group">
                <label><i class="fas fa-clock"></i> P/L Máximo</label>
                <input type="number" id="filtro-pl-max" placeholder="Ex: 12.0" step="1" oninput="aplicarFiltros()">
            </div>
            <div class="filter-group">
                <label><i class="fas fa-shield"></i> Preço Teto Graham</label>
                <select id="filtro-graham" onchange="aplicarFiltros()">
                    <option value="">Todos</option>
                    <option value="desconto">Apenas Abaixo do Teto (Desconto)</option>
                </select>
            </div>
        </div>

        <div class="summary-ribbon">
            <span>Ações exibidas: <b id="sum-count">0</b></span>
            <span>P/L Médio: <b id="sum-pl">-</b></span>
            <span>ROE Médio: <b id="sum-roe">-</b></span>
            <span>DY Médio: <b id="sum-dy">-</b></span>
            <span style="margin-left:auto; color:#46e0a0;"><i class="fas fa-bolt"></i> Cache em RAM ativo (< 2ms)</span>
        </div>

        <div class="table-wrap">
            <table id="tabela-mira">
                <thead>
                    <tr>
                        <th onclick="ordenarPor('ticker')">Ticker <i class="fas fa-sort"></i></th>
                        <th onclick="ordenarPor('nome')">Empresa <i class="fas fa-sort"></i></th>
                        <th onclick="ordenarPor('setor')">Setor <i class="fas fa-sort"></i></th>
                        <th class="num-cell" onclick="ordenarPor('preco')">Cotação (R$) <i class="fas fa-sort"></i></th>
                        <th class="num-cell" onclick="ordenarPor('pl')">P/L <i class="fas fa-sort"></i></th>
                        <th class="num-cell" onclick="ordenarPor('dy')">DY (%) <i class="fas fa-sort"></i></th>
                        <th class="num-cell" onclick="ordenarPor('roe')">ROE (%) <i class="fas fa-sort"></i></th>
                        <th class="num-cell" onclick="ordenarPor('graham')">Teto Graham <i class="fas fa-sort"></i></th>
                        <th class="num-cell" onclick="ordenarPor('margem_graham')">Margem Graham <i class="fas fa-sort"></i></th>
                        <th class="num-cell" onclick="ordenarPor('bazin')">Teto Bazin <i class="fas fa-sort"></i></th>
                        <th class="num-cell" onclick="ordenarPor('margem_bazin')">Margem Bazin <i class="fas fa-sort"></i></th>
                    </tr>
                </thead>
                <tbody id="tabela-body">
                    <tr><td colspan="11" style="text-align:center; padding:30px;"><i class="fas fa-spinner fa-spin"></i> Carregando dados da planilha do Mira...</td></tr>
                </tbody>
            </table>
        </div>
    </div>

    <script>
        let todasAcoes = [];
        let acoesFiltradas = [];
        let colunaOrdenacao = 'ticker';
        let ordemAsc = true;

        async function carregarValuation() {
            try {
                const res = await fetch('/api/valuation');
                todasAcoes = await res.json();
                popularSetores();
                aplicarFiltros();
            } catch (err) {
                document.getElementById('tabela-body').innerHTML = `<tr><td colspan="11" style="color:#ff6b6b; text-align:center;">Erro ao carregar dados: ${err.message}</td></tr>`;
            }
        }

        async function recarregarCache() {
            document.getElementById('tabela-body').innerHTML = `<tr><td colspan="11" style="text-align:center;"><i class="fas fa-spinner fa-spin"></i> Atualizando cache da planilha...</td></tr>`;
            await fetch('/api/reload');
            await carregarValuation();
        }

        function popularSetores() {
            const select = document.getElementById('filtro-setor');
            const setores = [...new Set(todasAcoes.map(a => a.setor).filter(Boolean))].sort();
            select.innerHTML = '<option value="">Todos os Setores</option>' + setores.map(s => `<option value="${s}">${s}</option>`).join('');
        }

        function aplicarFiltros() {
            const busca = document.getElementById('filtro-busca').value.trim().toLowerCase();
            const setor = document.getElementById('filtro-setor').value;
            const dyMin = parseFloat(document.getElementById('filtro-dy-min').value) || 0;
            const roeMin = parseFloat(document.getElementById('filtro-roe-min').value) || -9999;
            const plMax = parseFloat(document.getElementById('filtro-pl-max').value) || 9999;
            const grahamFiltro = document.getElementById('filtro-graham').value;

            acoesFiltradas = todasAcoes.filter(a => {
                if (busca && !(`${a.ticker} ${a.nome}`.toLowerCase().includes(busca))) return false;
                if (setor && a.setor !== setor) return false;
                if (a.dy < dyMin) return false;
                if (a.roe < roeMin) return false;
                if (a.pl > plMax) return false;
                if (grahamFiltro === 'desconto' && (a.graham <= a.preco || a.graham <= 0)) return false;
                return true;
            });

            ordenarDados();
            renderizarTabela();
            atualizarSumario();
        }

        function ordenarPor(coluna) {
            if (colunaOrdenacao === coluna) {
                ordemAsc = !ordemAsc;
            } else {
                colunaOrdenacao = coluna;
                ordemAsc = true;
            }
            ordenarDados();
            renderizarTabela();
        }

        function ordenarDados() {
            acoesFiltradas.sort((a, b) => {
                let vA = a[colunaOrdenacao];
                let vB = b[colunaOrdenacao];
                if (typeof vA === 'string') {
                    return ordemAsc ? vA.localeCompare(vB) : vB.localeCompare(vA);
                }
                return ordemAsc ? (vA - vB) : (vB - vA);
            });
        }

        function atualizarSumario() {
            document.getElementById('sum-count').innerText = acoesFiltradas.length;
            if (acoesFiltradas.length > 0) {
                const mediaPL = acoesFiltradas.filter(a => a.pl > 0).reduce((acc, a) => acc + a.pl, 0) / (acoesFiltradas.filter(a => a.pl > 0).length || 1);
                const mediaROE = acoesFiltradas.reduce((acc, a) => acc + a.roe, 0) / acoesFiltradas.length;
                const mediaDY = acoesFiltradas.reduce((acc, a) => acc + a.dy, 0) / acoesFiltradas.length;

                document.getElementById('sum-pl').innerText = mediaPL.toFixed(1) + 'x';
                document.getElementById('sum-roe').innerText = mediaROE.toFixed(1) + '%';
                document.getElementById('sum-dy').innerText = mediaDY.toFixed(1) + '%';
            }
        }

        function renderizarTabela() {
            const tbody = document.getElementById('tabela-body');
            if (acoesFiltradas.length === 0) {
                tbody.innerHTML = `<tr><td colspan="11" style="text-align:center; padding:30px; color:var(--text-muted);">Nenhuma ação atende aos filtros especificados.</td></tr>`;
                return;
            }

            let html = '';
            for (const a of acoesFiltradas) {
                const margemGrahamClass = a.margem_graham > 0 ? 'tag-positive' : 'tag-negative';
                const margemBazinClass = a.margem_bazin > 0 ? 'tag-positive' : 'tag-negative';

                html += `
                <tr>
                    <td><span class="ticker-badge">${a.ticker}</span></td>
                    <td style="font-weight:600;">${a.nome}</td>
                    <td style="color:var(--text-muted);">${a.setor}</td>
                    <td class="num-cell">R$ ${a.preco.toFixed(2)}</td>
                    <td class="num-cell">${a.pl.toFixed(1)}</td>
                    <td class="num-cell">${a.dy.toFixed(2)}%</td>
                    <td class="num-cell">${a.roe.toFixed(1)}%</td>
                    <td class="num-cell">R$ ${a.graham > 0 ? a.graham.toFixed(2) : '-'}</td>
                    <td class="num-cell ${margemGrahamClass}">${a.graham > 0 ? a.margem_graham.toFixed(1) + '%' : '-'}</td>
                    <td class="num-cell">R$ ${a.bazin > 0 ? a.bazin.toFixed(2) : '-'}</td>
                    <td class="num-cell ${margemBazinClass}">${a.bazin > 0 ? a.margem_bazin.toFixed(1) + '%' : '-'}</td>
                </tr>
                `;
            }
            tbody.innerHTML = html;
        }

        function exportarCSV() {
            if (acoesFiltradas.length === 0) return;
            let csv = 'Ticker;Nome;Setor;Preco;PL;DY;ROE;Graham;Margem_Graham;Bazin;Margem_Bazin\n';
            for (const a of acoesFiltradas) {
                csv += `${a.ticker};${a.nome};${a.setor};${a.preco};${a.pl};${a.dy};${a.roe};${a.graham};${a.margem_graham};${a.bazin};${a.margem_bazin}\n`;
            }
            const blob = new Blob([csv], { type: 'text/csv;charset=utf-8;' });
            const link = document.createElement('a');
            link.href = URL.createObjectURL(blob);
            link.download = 'valuation_mira_filtrado.csv';
            link.click();
        }

        window.addEventListener('DOMContentLoaded', carregarValuation);
    </script>
</body>
</html>
"""
        return html

def iniciar_servidor(port=PORT):
    server_address = ("", port)
    httpd = ThreadingHTTPServer(server_address, MiraHandler)
    print(f"[Mira Server] Monitor de Valuation ativo em http://localhost:{port}")
    httpd.serve_forever()

if __name__ == "__main__":
    iniciar_servidor()
