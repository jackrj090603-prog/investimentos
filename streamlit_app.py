import os
import sys
import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import seaborn as sns

# Configuração da página Streamlit Cloud
st.set_page_config(
    page_title="Ceará Finance — Hub de Valuations & Backtest",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Header com Logotipo Oficial
col_logo, col_header = st.columns([0.15, 0.85])
with col_logo:
    for p in ["LOGO_CF.png", "Mira/LOGO_CF.png", "../LOGO_CF.png"]:
        if os.path.exists(p):
            st.image(p, width=110)
            break
with col_header:
    st.title("Ceará Finance — Inteligência de Mercado & Valuations")
    st.caption("Liga Acadêmica de Mercado Financeiro da UFC • Plataforma Quantitativa & Fundamentalista")

# Importar módulos locais com fallback
try:
    import dados
    import backtest
    import graficos
    import figuras
    HAS_BACKTEST_MODULES = True
except Exception as e:
    HAS_BACKTEST_MODULES = False

# ==============================================================================
# FUNÇÃO DE CACHE: MONITOR DE VALUATION DO MIRA
# ==============================================================================
@st.cache_data(show_spinner="Carregando e calculando planilha do Mira...")
def carregar_dados_mira():
    paths = [
        "Mira/MONITOR DE VALUATION DO MIRA (3).xlsx",
        "MONITOR DE VALUATION DO MIRA (3).xlsx",
        "../Mira/MONITOR DE VALUATION DO MIRA (3).xlsx"
    ]
    excel_path = None
    for p in paths:
        if os.path.exists(p):
            excel_path = p
            break
            
    if not excel_path:
        return pd.DataFrame()
        
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        b3_map = {}
        sheet_b3_name = 'Classificação B3' if 'Classificação B3' in wb.sheetnames else 'Classificacao B3'
        if sheet_b3_name in wb.sheetnames:
            sheet_b3 = wb[sheet_b3_name]
            for r in range(2, sheet_b3.max_row + 1):
                ticker = sheet_b3.cell(r, 1).value
                if ticker:
                    b3_map[str(ticker).strip().upper()] = {
                        "nome": sheet_b3.cell(r, 2).value or "",
                        "setor": sheet_b3.cell(r, 4).value or "Outros",
                        "segmento": sheet_b3.cell(r, 6).value or "Outros"
                    }
                    
        sheet_si = wb['Dados Status Invest']
        records = []
        
        for r in range(2, sheet_si.max_row + 1):
            ticker = sheet_si.cell(r, 1).value
            if not ticker:
                continue
            ticker = str(ticker).strip().upper()
            
            def get_val(col_idx):
                val = sheet_si.cell(r, col_idx).value
                if val is None or val == "-":
                    return 0.0
                try:
                    return float(val)
                except:
                    return 0.0
                    
            preco = get_val(2)
            dy_raw = get_val(3)
            pl = get_val(4)
            roe_raw = get_val(18)
            cagr_raw = get_val(25)
            vpa = get_val(27)
            lpa = get_val(28)
            
            graham = round((22.5 * vpa * lpa) ** 0.5, 2) if vpa > 0 and lpa > 0 else 0.0
            dpa = preco * dy_raw if dy_raw > 0 else 0.0
            bazin = round(dpa / 0.06, 2) if dpa > 0 else 0.0
            
            b3 = b3_map.get(ticker, {"nome": ticker, "setor": "Outros", "segmento": "Outros"})
            
            records.append({
                "Ticker": ticker,
                "Empresa": b3["nome"],
                "Setor": b3["setor"],
                "Preço (R$)": preco,
                "P/L": pl,
                "DY (%)": round(dy_raw * 100, 2),
                "ROE (%)": round(roe_raw * 100, 2),
                "Teto Graham (R$)": graham,
                "Margem Graham (%)": round(((graham / preco) - 1) * 100, 1) if preco > 0 and graham > 0 else 0.0,
                "Teto Bazin (R$)": bazin,
                "Margem Bazin (%)": round(((bazin / preco) - 1) * 100, 1) if preco > 0 and bazin > 0 else 0.0
            })
            
        wb.close()
        return pd.DataFrame(records)
    except Exception as err:
        st.error(f"Erro ao ler planilha: {err}")
        return pd.DataFrame()

# ==============================================================================
# ABAS DA APLICAÇÃO CLOUD
# ==============================================================================
tab_backtest, tab_valuation, tab_research, tab_sobre = st.tabs([
    "🛡️ Aegis Momentum Backtest",
    "📊 Monitor de Valuation (Mira)",
    "📋 Equity Research & Demonstrativos",
    "🏛️ Sobre a Plataforma"
])

# ------------------------------------------------------------------------------
# ABA 1: AEGIS MOMENTUM BACKTEST
# ------------------------------------------------------------------------------
with tab_backtest:
    st.subheader("Estratégia Quantitativa: Momentum Long-Short (B3 vs CDI)")
    
    if not HAS_BACKTEST_MODULES:
        st.warning("Módulos quantitativos locais carregando...")
    else:
        @st.cache_resource
        def carregar_dados_sistema():
            return dados.carregar_dados_completos()
            
        with st.spinner("Carregando cotações históricas da B3 e dados do CDI..."):
            data_dict = carregar_dados_sistema()
            
        col_ctrl1, col_ctrl2, col_ctrl3, col_ctrl4 = st.columns(4)
        with col_ctrl1:
            lookbacks = st.multiselect("Lookbacks (meses)", [3, 6, 9, 12, 18], default=[6, 12])
        with col_ctrl2:
            ativos = st.multiselect("Ativos por Perna (Top N)", [3, 5, 10, 15], default=[5, 10])
        with col_ctrl3:
            estilos = st.multiselect("Estilo da Operação", ["Long-Short", "Long-Only"], default=["Long-Short"])
        with col_ctrl4:
            custo_bps = st.selectbox("Custos Transação (bps)", [0, 5, 10, 20], index=2)
            
        if st.button("🚀 Executar Backtest Interativo", use_container_width=True):
            resultados = []
            progress_bar = st.progress(0)
            
            import itertools
            combinacoes = list(itertools.product(lookbacks, ativos, estilos))
            
            if not combinacoes:
                st.warning("Selecione pelo menos uma opção em cada filtro para simular.")
            else:
                for idx, (l, n, est) in enumerate(combinacoes):
                    res = backtest.rodar_backtest(
                        data_dict=data_dict,
                        lookback_meses=l,
                        top_n=n,
                        estrategia=est,
                        custos_bps=custo_bps
                    )
                    resultados.append(res)
                    progress_bar.progress((idx + 1) / len(combinacoes))
                    
                df_exp = pd.DataFrame(resultados)
                st.success(f"Backtest concluído para {len(combinacoes)} configurações!")
                
                # Exibir tabela comparativa de métricas
                cols_show = ["lookback_meses", "top_n", "estrategia", "retorno_acumulado", "retorno_anualizado", "volatilidade_anualizada", "sharpe", "max_drawdown", "taxa_acerto"]
                df_view = df_exp[[c for c in cols_show if c in df_exp.columns]].copy()
                st.dataframe(df_view.style.highlight_max(subset=["sharpe", "retorno_acumulado"], color="#1e3a8a"), use_container_width=True)

# ------------------------------------------------------------------------------
# ABA 2: MONITOR DE VALUATION DO MIRA
# ------------------------------------------------------------------------------
with tab_valuation:
    st.subheader("Planilha Inteligente de Valuation B3 (Base Mira)")
    df_mira = carregar_dados_mira()
    
    if df_mira.empty:
        st.info("Planilha não localizada ou vazia.")
    else:
        # Filtros Dinâmicos
        col_f1, col_f2, col_f3, col_f4 = st.columns(4)
        with col_f1:
            busca = st.text_input("🔍 Busca por Ticker / Empresa", "")
        with col_f2:
            setores_opcoes = ["Todos"] + sorted(list(df_mira["Setor"].dropna().unique()))
            setor_sel = st.selectbox("Setor de Atuação", setores_opcoes)
        with col_f3:
            dy_min = st.number_input("DY Mínimo (%)", min_value=0.0, value=0.0, step=1.0)
        with col_f4:
            apenas_graham = st.checkbox("Apenas com Desconto Graham", value=False)
            
        df_filtrado = df_mira.copy()
        if busca:
            df_filtrado = df_filtrado[
                df_filtrado["Ticker"].str.contains(busca.upper()) | 
                df_filtrado["Empresa"].str.contains(busca, case=False, na=False)
            ]
        if setor_sel != "Todos":
            df_filtrado = df_filtrado[df_filtrado["Setor"] == setor_sel]
        if dy_min > 0:
            df_filtrado = df_filtrado[df_filtrado["DY (%)"] >= dy_min]
        if apenas_graham:
            df_filtrado = df_filtrado[df_filtrado["Margem Graham (%)"] > 0]
            
        # Métricas Resumo
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Ações Selecionadas", len(df_filtrado))
        m2.metric("P/L Médio", f"{df_filtrado[df_filtrado['P/L'] > 0]['P/L'].mean():.1f}x" if not df_filtrado.empty else "-")
        m3.metric("DY Médio", f"{df_filtrado['DY (%)'].mean():.2f}%" if not df_filtrado.empty else "-")
        m4.metric("ROE Médio", f"{df_filtrado['ROE (%)'].mean():.1f}%" if not df_filtrado.empty else "-")
        
        # Tabela Interativa
        st.dataframe(
            df_filtrado.style.format({
                "Preço (R$)": "R$ {:.2f}",
                "P/L": "{:.1f}",
                "DY (%)": "{:.2f}%",
                "ROE (%)": "{:.1f}%",
                "Teto Graham (R$)": "R$ {:.2f}",
                "Margem Graham (%)": "{:+.1f}%",
                "Teto Bazin (R$)": "R$ {:.2f}",
                "Margem Bazin (%)": "{:+.1f}%"
            }),
            use_container_width=True,
            height=450
        )
        
        # Botão para download de CSV
        csv_data = df_filtrado.to_csv(index=False, sep=";").encode("utf-8")
        st.download_button(
            label="📥 Exportar Dados Filtrados (CSV)",
            data=csv_data,
            file_name="valuation_mira_ceara_finance.csv",
            mime="text/csv"
        )

# ------------------------------------------------------------------------------
# ABA 3: EQUITY RESEARCH & DEMONSTRATIVOS
# ------------------------------------------------------------------------------
with tab_research:
    st.subheader("Relatórios Acadêmicos & Lâminas de Tese (Modelo Poli USP)")
    
    empresa_sel = st.selectbox(
        "Selecione o Ativo Coberto:",
        ["DIRR3 (Direcional Engenharia)", "PETR4 (Petrobras)", "VALE3 (Vale S.A.)", "WEGE3 (WEG S.A.)"]
    )
    ticker = empresa_sel.split()[0]
    
    col_t1, col_t2 = st.columns([0.65, 0.35])
    with col_t1:
        st.markdown(f"### Tese de Investimento — **{ticker}**")
        st.write(f"A análise para **{ticker}** fundamenta-se na solidez do seu balanço patrimonial, na disciplina de alocação de capital e na resiliência operacional frente aos ciclos macroeconômicos brasileiros.")
        
        st.markdown("#### 🎯 Principais Gatilhos de Valor (Drivers)")
        st.markdown("""
        - **Expansão de Margem Operacional**: Ganhos de eficiência em custos e escala.
        - **Retorno Superior (ROIC vs WACC)**: Criação de valor econômico agregado (EVA).
        - **Geração de Caixa Livre**: Sustentabilidade na distribuição de proventos e desalavancagem.
        """)
        
    with col_t2:
        st.markdown("### Preço-Alvo & Recomendação")
        precos = {"DIRR3": (32.50, 24.80), "PETR4": (41.00, 37.20), "VALE3": (75.00, 58.40), "WEGE3": (62.00, 53.10)}
        alvo, atual = precos.get(ticker, (30.0, 20.0))
        upside = round(((alvo / atual) - 1) * 100, 1)
        
        st.metric(label="Recomendação", value="COMPRA", delta=f"+{upside}% Potencial")
        st.metric(label="Preço-Alvo Estimado", value=f"R$ {alvo:.2f}")
        st.metric(label="Cotação de Referência", value=f"R$ {atual:.2f}")

# ------------------------------------------------------------------------------
# ABA 4: SOBRE A PLATAFORMA
# ------------------------------------------------------------------------------
with tab_sobre:
    st.subheader("Sobre a Ceará Finance")
    st.write("""
    A **Ceará Finance** é a Liga Acadêmica de Mercado Financeiro da Universidade Federal do Ceará (UFC), 
    dedicada à formação de excelência em análise quantitativa, valuation fundamentalista, asset management e equity research.
    
    Esta plataforma congrega:
    1. **Modelagem Fundamentalista**: DCF com 5 anos de projeção, WACC via OLS contra o Ibovespa e Simulação de Monte Carlo.
    2. **Estratégias Quantitativas**: Fatores de Momentum (Long-Short) com controle estrito de risco e drawdown.
    3. **Monitoramento Regulatório**: Integração contínua com os dados oficiais da CVM.
    """)
    st.info("Repositório Oficial: https://github.com/jackrj090603-prog/investimentos")
